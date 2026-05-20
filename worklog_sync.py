#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
故障闭环 → 工作记录.xlsx 自动同步

当运维平台闭环一条故障后，自动将故障信息写入工作记录.xlsx。
字段映射：
  - 序号：当前年份区域自增
  - 时间：closed_at 转为中文日期格式
  - 变电站：stations.name（如"仙都变"）
  - 地点：stations.county（如"缙云"）
  - 故障描述：摄像机位置 + 故障类型 + 处理方式 自动拼接
  - 类型：system_type 映射（默认"图像监控"）
  - 甲供：equipment_type + equipment_quantity
  - 工作负责人：handler_name
"""

import os
import re
import logging
from datetime import datetime
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

# ---- 故障描述拼接 ----

# 处理方式关键词映射：从 handler_note 或 fault_type 推断
_ACTION_KEYWORDS = {
    "更换": "更换",
    "替换": "更换",
    "替换": "更换",
    "维修": "维修",
    "修复": "维修",
    "恢复": "恢复",
    "重做": "重做",
    "制作": "制作",
    "拆除": "拆除",
    "敷设": "敷设",
    "放线": "放线",
    "调试": "调试",
    "消缺": "消缺",
}

# system_type → 工作记录"类型"列映射
_SYSTEM_TYPE_MAP = {
    "image_monitoring": "图像监控",
    "图像监控": "图像监控",
    "smart_patrol": "智能巡视",
    "智能巡视": "智能巡视",
    "auxiliary_control": "辅控系统",
    "辅控系统": "辅控系统",
}

# ---- 工作记录读写 ----


def _find_worklog_path() -> str:
    """定位工作记录.xlsx路径。"""
    candidates = [
        r"E:\办公\工作记录\工作记录.xlsx",
        os.path.join(os.path.dirname(__file__), "..", "..", "办公", "工作记录", "工作记录.xlsx"),
    ]
    for p in candidates:
        resolved = Path(p).resolve()
        if resolved.exists():
            return str(resolved)
    raise FileNotFoundError("找不到工作记录.xlsx")


def _parse_closed_at(closed_at) -> datetime:
    """将 closed_at 字段转为 datetime 对象。"""
    if isinstance(closed_at, datetime):
        return closed_at
    text = str(closed_at).strip()
    # 尝试多种格式
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"无法解析 closed_at: {text}")


def _format_date(dt: datetime) -> str:
    """datetime → 中文日期格式，如 '2026年5月07日'。月和日补零。"""
    return f"{dt.year}年{dt.month}月{dt.day:02d}日"


def _infer_action(handler_note: str) -> str:
    """从处理备注中推断处理方式。"""
    for keyword, action in _ACTION_KEYWORDS.items():
        if keyword in handler_note:
            return action
    return "处理"


def _normalize_fault_type(fault_type: str) -> str:
    """将平台故障类型归一为更适合写入工作记录的短语。"""
    text = str(fault_type or "").strip()
    if not text:
        return ""
    if text == "摄像机/球机/枪机故障":
        return "摄像机故障"
    if text == "网络掉线/通信异常":
        return "通信异常"
    if text == "集中电源/供电异常":
        return "供电异常"
    return text


def _map_system_type(system_type) -> str:
    """将 system_type 映射为工作记录的类型列。"""
    if not system_type:
        return "图像监控"
    mapped = _SYSTEM_TYPE_MAP.get(str(system_type).strip())
    return mapped if mapped else str(system_type).strip()


_STATION_NAME_PREFIX_RE = re.compile(r'^(?:\d+kV)?', re.IGNORECASE)


def _clean_station_name(name: str) -> str:
    """去除站点名中的电压等级前缀，如 '220kV睦田变' → '睦田变'。"""
    return re.sub(r'^(\d+kV)', '', name)


def _format_equipment(equipment_type, equipment_quantity) -> str:
    """格式化甲供列，如 '1台摄像机'。"""
    qty = 0
    try:
        qty = int(equipment_quantity or 0)
    except (TypeError, ValueError):
        pass
    if qty <= 0 or not equipment_type:
        return ""
    etype = str(equipment_type).strip()
    # 如果 equipment_type 不含量词，自动加 '台'
    if not re.match(r'^\d*(台|套|个|块|根|条|卷|张)', etype):
        etype = f"台{etype}"
    return f"{qty}{etype}"


def _looks_generic_note(handler_note: str) -> bool:
    """判断处理备注是否过于笼统，需要自动补充点位信息。"""
    note = str(handler_note or "").strip()
    if not note:
        return True
    generic_notes = {
        "摄像机故障更换",
        "故障更换",
        "故障处理",
        "摄像机故障处理",
        "更换故障摄像机",
        "摄像机故障",
        "更换",
        "维修",
        "恢复",
        "处理",
    }
    if note in generic_notes:
        return True
    # 这些词一旦出现，通常已经属于较具体的现场描述
    if any(word in note for word in (
        "排查", "后恢复", "恢复正常", "水晶头", "光纤", "收发器", "接口", "网线",
        "集中电源", "空开", "制作", "敷设", "拆除", "松动", "异常", "断电", "未恢复"
    )):
        return False
    if len(note) <= 8:
        return True
    return False


def _build_default_tail(fault_type: str, action: str) -> str:
    """构建兜底尾部文案。"""
    if fault_type == "摄像机故障" and action == "更换":
        return "摄像机故障更换"
    if fault_type:
        return f"{fault_type}{action}" if action and action not in fault_type else fault_type
    return action or "故障处理"


def _labels_preview(labels: list[str], max_count: int = 3) -> str:
    if len(labels) <= max_count:
        return "、".join(labels)
    return f"{'、'.join(labels[:max_count])}等{len(labels)}处"


def _build_description(camera_labels: list, fault_type: str, handler_note: str) -> str:
    """
    自动拼接故障描述。

    规则尽量贴近历史工作记录写法：
    1. 备注已经很具体时，优先保留原备注
    2. 备注过于笼统时，自动补齐摄像机点位
    3. 多摄像机时保留前3个点位，避免描述过长
    """
    fault_type = _normalize_fault_type(fault_type)
    handler_note = str(handler_note or "").strip().strip('；;，,。')
    action = _infer_action(handler_note)
    default_tail = _build_default_tail(fault_type, action)
    generic_tail = handler_note or default_tail

    normalized_labels = []
    seen = set()
    for raw_label in camera_labels or []:
        label = str(raw_label or "").strip()
        if not label or label in seen:
            continue
        seen.add(label)
        normalized_labels.append(label)

    if not normalized_labels:
        if handler_note:
            return f"未识别具体摄像机，{handler_note}"
        return f"未识别具体摄像机，{default_tail}"

    # 备注已包含任一点位，直接保留原始丰富写法
    if handler_note and any(label in handler_note for label in normalized_labels):
        return handler_note

    # 单摄像机：优先写成“点位 + 具体备注”
    if len(normalized_labels) == 1:
        label = normalized_labels[0]
        if handler_note and not _looks_generic_note(handler_note):
            return f"{label}{handler_note}"
        return f"{label}{generic_tail}"

    # 多摄像机：优先保留点位预览，再接丰富备注；过于笼统时走兜底写法
    preview = _labels_preview(normalized_labels)
    if handler_note and not _looks_generic_note(handler_note):
        return f"{preview}{handler_note}"
    return f"{preview}{generic_tail}"


def _find_or_create_year_section(ws, year: int) -> tuple:
    """
    在工作记录.xlsx中找到指定年份的起始行和结束行。
    返回 (data_start_row, data_end_row)，不含年份标题行。
    如果年份不存在，在末尾创建。
    """
    year_str = f"{year}年"

    # 扫描所有年份标题行（格式如 '2022年'）
    all_year_rows = {}  # row_num -> year_string
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val is not None:
            val_str = str(val).strip()
            if re.match(r'^\d{4}年$', val_str):
                all_year_rows[row_idx] = val_str

    target_found = year_str in all_year_rows.values()

    if not all_year_rows and ws.max_row <= 2:
        # 空表或只有表头
        ws.cell(row=2, column=1, value=year_str)
        return 3, 3

    if not target_found:
        # 年份不存在，在所有数据的最后一个年份区域之后追加
        sorted_year_positions = sorted(all_year_rows.keys())
        last_year_row = sorted_year_positions[-1]
        # 找这个年份区域的数据结束行
        end_row = last_year_row + 1
        while end_row <= ws.max_row:
            val = ws.cell(row=end_row, column=1).value
            if val is not None and str(val).strip().endswith("年"):
                break
            if val is not None:
                end_row += 1
            else:
                end_row += 1
        # 在 end_row 插入新年份
        ws.cell(row=end_row, column=1, value=year_str)
        return end_row + 1, end_row  # 空区域，start > end 表示无数据

    # 年份已存在
    target_row = None
    for yr_row, yr_str in all_year_rows.items():
        if yr_str == year_str:
            target_row = yr_row
            break

    if target_row is None:
        # fallback
        return ws.max_row + 1, ws.max_row

    # 找下一个年份行
    sorted_year_rows = sorted(all_year_rows.keys())
    next_year_row = ws.max_row + 1
    for yr_row in sorted_year_rows:
        if yr_row > target_row:
            next_year_row = yr_row
            break

    data_start = target_row + 1
    data_end = next_year_row - 1

    return data_start, data_end


def _get_next_seq(ws, data_start: int, data_end: int) -> int:
    """获取当前年份区域的有效最大序号+1。只统计时间列有值的行。"""
    max_seq = 0
    for row_idx in range(data_start, data_end + 1):
        # 只有序号和时间都有值的行才算有效数据
        time_val = ws.cell(row=row_idx, column=2).value
        if time_val is None or str(time_val).strip() == '':
            continue
        val = ws.cell(row=row_idx, column=1).value
        if val is not None:
            try:
                seq = int(val)
                if seq > max_seq:
                    max_seq = seq
            except (TypeError, ValueError):
                pass
    return max_seq + 1


def _find_first_empty_row(ws, data_start: int, data_end: int) -> int:
    """在年份区域中找到第一个时间列（第2列）为空的行。"""
    for row_idx in range(data_start, data_end + 1):
        val = ws.cell(row=row_idx, column=2).value
        if val is None or str(val).strip() == '':
            return row_idx
    return data_end + 1


def sync_fault_to_worklog(
    db,
    fault_id: int,
    worklog_path: str = None,
) -> bool:
    """
    将一条已闭环的故障写入工作记录.xlsx。

    参数：
      db: 数据库连接
      fault_id: 故障ID
      worklog_path: 工作记录.xlsx路径（可选，自动检测）

    返回：
      True 表示成功写入，False 表示跳过
    """
    # 1. 查询故障记录 + 站点信息
    fault = db.execute(
        """
        SELECT
            fr.id, fr.station_id, fr.camera_id, fr.fault_type, fr.handler_name,
            fr.handler_note, fr.equipment_type, fr.equipment_quantity,
            fr.closed_at, fr.system_type, fr.status,
            s.name AS station_name, s.county AS station_county
        FROM fault_reports fr
        JOIN stations s ON fr.station_id = s.id
        WHERE fr.id = ?
        """,
        (fault_id,),
    ).fetchone()

    if not fault:
        logger.warning("worklog_sync: 故障 %d 不存在", fault_id)
        return False

    if fault["status"] != "closed":
        logger.warning("worklog_sync: 故障 %d 未闭环，跳过", fault_id)
        return False

    # 2. 查询关联的摄像机信息
    camera_labels = []
    try:
        cameras = db.execute(
            """
            SELECT c.location_desc, c.area, c.camera_index
            FROM fault_report_cameras frc
            JOIN cameras c ON frc.camera_id = c.id
            WHERE frc.fault_report_id = ?
            ORDER BY c.id
            """,
            (fault_id,),
        ).fetchall()
        for cam in cameras:
            # 优先用 location_desc，其次 area + camera_index
            label = str(cam["location_desc"] or "").strip()
            if not label:
                area = str(cam["area"] or "").strip()
                idx = str(cam["camera_index"] or "").strip()
                if area and idx:
                    label = f"{area}{idx}"
                elif area:
                    label = area
                elif idx:
                    label = f"#{idx}"
            if label:
                camera_labels.append(label)
    except Exception:
        # fault_report_cameras 表可能不存在（老数据）
        pass

    # 兼容老数据：如果明细表没取到，再尝试主摄像机
    if not camera_labels and fault.get("camera_id"):
        try:
            cam = db.execute(
                """
                SELECT location_desc, area, camera_index
                FROM cameras
                WHERE id = ?
                """,
                (fault["camera_id"],),
            ).fetchone()
            if cam:
                label = str(cam["location_desc"] or "").strip()
                if not label:
                    area = str(cam["area"] or "").strip()
                    idx = str(cam["camera_index"] or "").strip()
                    if area and idx:
                        label = f"{area}{idx}"
                    elif area:
                        label = area
                    elif idx:
                        label = f"#{idx}"
                if label:
                    camera_labels.append(label)
        except Exception:
            pass

    # 3. 构建各字段
    closed_dt = _parse_closed_at(fault["closed_at"])
    date_str = _format_date(closed_dt)
    station_name = _clean_station_name(str(fault["station_name"] or "").strip())
    county = str(fault["station_county"] or "").strip()
    fault_type = str(fault["fault_type"] or "").strip()
    handler_note = str(fault["handler_note"] or "").strip()
    description = _build_description(camera_labels, fault_type, handler_note)
    work_type = _map_system_type(fault["system_type"])
    equipment = _format_equipment(fault["equipment_type"], fault["equipment_quantity"])
    handler = str(fault["handler_name"] or "").strip()

    year = closed_dt.year

    # 4. 定位工作记录.xlsx
    if not worklog_path:
        worklog_path = _find_worklog_path()

    logger.info(
        "worklog_sync: 写入故障 %d → %s %s %s",
        fault_id, date_str, station_name, description,
    )

    # 5. 读写 Excel
    wb = openpyxl.load_workbook(worklog_path)
    ws = wb["Sheet1"]

    data_start, data_end = _find_or_create_year_section(ws, year)
    seq = _get_next_seq(ws, data_start, data_end)
    target_row = _find_first_empty_row(ws, data_start, data_end)

    # 从上一行（最后一条有效数据行）复制样式
    style_source_row = target_row - 1
    values = [seq, date_str, station_name, county, description, work_type,
              equipment if equipment else None, handler if handler else None]
    for col_idx, val in enumerate(values, start=1):
        target_cell = ws.cell(row=target_row, column=col_idx)
        source_cell = ws.cell(row=style_source_row, column=col_idx)
        # 复制样式
        if source_cell.has_style:
            target_cell.font = source_cell.font.copy()
            target_cell.alignment = source_cell.alignment.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.fill = source_cell.fill.copy()
            target_cell.number_format = source_cell.number_format
        # 写入值
        target_cell.value = val

    wb.save(worklog_path)
    wb.close()

    logger.info("worklog_sync: 已写入工作记录第 %d 行，序号 %d", target_row, seq)
    return True
