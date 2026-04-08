# parse_excel.py — Excel解析模块
import os
import re
from pathlib import Path

from openpyxl import load_workbook
try:
    import xlrd
except ImportError:  # pragma: no cover - production dependency, exercised via explicit error path
    xlrd = None

class ExcelParseError(Exception):
    """Excel解析异常"""
    pass


def validate_station_inventory_data(data, filepath=''):
    cameras = data.get('cameras') if isinstance(data, dict) else None
    if isinstance(cameras, list) and cameras:
        return data

    prefix = f"{filepath}: " if filepath else ''
    raise ExcelParseError(f"{prefix}未识别到摄像头，请确认上传的是摄像头台账而不是监控日报或其他报表")


def _load_xlsx_rows(filepath):
    try:
        workbook = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelParseError(f"无法打开Excel文件: {filepath}, 错误: {exc}") from exc

    try:
        worksheet = workbook.active
    except Exception as exc:
        workbook.close()
        raise ExcelParseError(f"无法读取工作表: {filepath}, 错误: {exc}") from exc

    try:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ExcelParseError(f"Excel文件为空: {filepath}")
        return rows
    except ExcelParseError:
        raise
    except Exception as exc:
        raise ExcelParseError(f"解析Excel失败: {filepath}, 错误: {exc}") from exc
    finally:
        workbook.close()


def _load_xls_rows(filepath):
    if xlrd is None:
        raise ExcelParseError(f"无法打开Excel文件: {filepath}, 错误: 当前环境未安装xlrd，无法读取.xls文件")

    try:
        workbook = xlrd.open_workbook(filepath, on_demand=True)
    except Exception as exc:
        raise ExcelParseError(f"无法打开Excel文件: {filepath}, 错误: {exc}") from exc

    try:
        worksheet = workbook.sheet_by_index(0)
    except Exception as exc:
        raise ExcelParseError(f"无法读取工作表: {filepath}, 错误: {exc}") from exc

    try:
        rows = [worksheet.row_values(index) for index in range(worksheet.nrows)]
        if not rows:
            raise ExcelParseError(f"Excel文件为空: {filepath}")
        return rows
    except ExcelParseError:
        raise
    except Exception as exc:
        raise ExcelParseError(f"解析Excel失败: {filepath}, 错误: {exc}") from exc
    finally:
        release_resources = getattr(workbook, 'release_resources', None)
        if callable(release_resources):
            release_resources()


def _load_excel_rows(filepath):
    suffix = Path(filepath).suffix.lower()
    if suffix == '.xls':
        return _load_xls_rows(filepath)
    return _load_xlsx_rows(filepath)


def _normalize_header_value(value):
    return str(value or '').strip()


def _is_flat_inventory_format(rows):
    if not rows:
        return False
    header_row = rows[0] or []
    normalized_headers = {_normalize_header_value(cell) for cell in header_row if cell}
    required_headers = {'序号', '变电站', '设备名称'}
    return required_headers.issubset(normalized_headers)


def _extract_voltage_from_station_name(station_name):
    match = re.search(r'(\d+)\s*[kK][vV]', str(station_name or ''))
    if match:
        return f"{match.group(1)}kV"
    return ''


DEVICE_TYPE_SUFFIXES = (
    '测温球机',
    '全景机',
    '卡片机',
    '球机',
    '云台',
    '枪机',
)

VIDEO_TYPE_SUFFIXES = (
    '可见光',
    '测温',
    '热成像',
)


def _strip_inventory_prefix(text):
    value = str(text or '').strip()
    value = re.sub(r'^[A-Za-z]+\d+-', '', value)
    value = re.sub(r'^\d+-', '', value)
    return value


def _strip_device_type_suffix(text):
    value = str(text or '').strip().rstrip('-').rstrip()
    for suffix in VIDEO_TYPE_SUFFIXES:
        value = re.sub(rf'-?{suffix}\s*$', '', value)
    for suffix in DEVICE_TYPE_SUFFIXES:
        value = re.sub(rf'{suffix}\s*$', '', value)
    return value.rstrip('-').strip()


def _derive_location_from_device_name(device_name):
    text = _strip_inventory_prefix(device_name)
    if not text:
        return ''
    parts = [part.strip() for part in text.split('-') if str(part).strip()]
    if len(parts) >= 5 and parts[1].isdigit() and parts[-1].isdigit():
        return '-'.join(parts[2:-2]).strip()
    if len(parts) >= 3:
        semantic_prefixes = {'红外', '可见光', '测温', '热成像'}
        normalized_tail = _strip_device_type_suffix(parts[-1])
        middle = parts[1:-1]
        if parts[0] in semantic_prefixes:
            middle = [parts[0], *middle]
        if normalized_tail:
            middle.append(normalized_tail)
        return '-'.join(part for part in middle if part).strip('-').strip()

    stripped = _strip_device_type_suffix(text)
    return stripped.strip('-').strip()


def _extract_camera_index_from_device_name(device_name):
    text = _strip_inventory_prefix(device_name)
    stripped = _strip_device_type_suffix(text)

    match = re.search(r'#\s*(\d+)\s*$', stripped)
    if match:
        return match.group(1)

    match = re.search(r'-(\d+)\s*#?\s*$', stripped)
    if match:
        return match.group(1)

    hash_matches = re.findall(r'#\s*(\d+)', stripped)
    if hash_matches:
        return hash_matches[-1]

    match = re.search(r'(\d+)\s*#?\s*$', stripped)
    if match:
        return match.group(1)
    return ''


def _parse_flat_inventory_rows(rows, filepath):
    header_row = rows[0] or []
    header_index = {}
    for index, cell in enumerate(header_row):
        header = _normalize_header_value(cell)
        if header:
            header_index[header] = index

    station_col = header_index.get('变电站')
    device_name_col = header_index.get('设备名称')
    model_col = header_index.get('设备型号')
    use_type_col = header_index.get('使用类型')
    video_type_col = header_index.get('视频类型')
    install_col = header_index.get('安装位置')

    if station_col is None or device_name_col is None:
        raise ExcelParseError(f"缺少智慧巡视台账关键列: {filepath}")

    station_name = ''
    for row in rows[1:]:
        if row and len(row) > station_col and row[station_col]:
            station_name = str(row[station_col]).strip()
            break
    if not station_name:
        raise ExcelParseError(f"变电站名称为空: {filepath}")

    cameras = []
    for row in rows[1:]:
        if not row or len(row) <= device_name_col:
            continue

        row_station_name = str(row[station_col]).strip() if len(row) > station_col and row[station_col] else station_name
        device_name = str(row[device_name_col]).strip() if row[device_name_col] else ''
        if not row_station_name or not device_name:
            continue

        location = ''
        if install_col is not None and len(row) > install_col and row[install_col]:
            location = str(row[install_col]).strip()
        if not location:
            location = _derive_location_from_device_name(device_name)

        camera_index = _extract_camera_index_from_device_name(device_name)
        use_type = str(row[use_type_col]).strip() if use_type_col is not None and len(row) > use_type_col and row[use_type_col] else ''
        video_type = str(row[video_type_col]).strip() if video_type_col is not None and len(row) > video_type_col and row[video_type_col] else ''
        model = str(row[model_col]).strip() if model_col is not None and len(row) > model_col and row[model_col] else ''
        area = use_type
        if video_type:
            area = f"{use_type}/{video_type}" if use_type else video_type

        cameras.append({
            'camera_index': camera_index,
            'area': area,
            'location': location,
            'location_desc': location,
            'ip_address': '',
            'channel_port': None,
            'channel_number': int(camera_index) if camera_index.isdigit() else None,
            'slot_code': '',
            'project_camera_code': device_name,
            'device_model': model,
        })

    county = Path(filepath).parent.name
    return {
        'station': {
            'name': station_name,
            'voltage_level': _extract_voltage_from_station_name(station_name),
            'county': county,
            'location': '',
            'ip_range': '',
            'nvr_ip': '',
            'nvr_port': None,
        },
        'cameras': cameras
    }

def parse_station_excel(filepath):
    """
    解析变电站Excel文件

    返回结构:
    {
        'station': {'name': str, 'voltage_level': str, 'county': str, ...},
        'cameras': [{'camera_index': str, 'area': str, 'location': str, 'ip': str, 'port': int, 'channel': int}, ...]
    }

    Excel结构（根据数据源清单）:
    - 第1行：变电站名称
    - 第2-5行：站内设备
    - 第6-8行：摄像头区域
    - 第9-12行：通道端口
    - 第13行后：通道详情
    """
    if not os.path.exists(filepath):
        raise ExcelParseError(f"文件不存在: {filepath}")

    try:
        rows = _load_excel_rows(filepath)

        if _is_flat_inventory_format(rows):
            return _parse_flat_inventory_rows(rows, filepath)

        # 解析变电站信息（第1行）
        station_name = rows[0][0] if rows[0] and rows[0][0] else None
        if not station_name:
            raise ExcelParseError(f"变电站名称为空: {filepath}")

        # 解析电压等级（从文件名或内容推断，这里先留空让import时填充）
        voltage_level = extract_voltage_from_content(rows)

        # 解析县区（从文件路径推断）
        county = extract_county_from_path(filepath)

        # 解析摄像头信息
        cameras = parse_camera_rows(rows)

        return {
            'station': {
                'name': str(station_name).strip(),
                'voltage_level': voltage_level,
                'county': county,
                'location': '',
                'ip_range': '',
                'nvr_ip': '',
                'nvr_port': None,
            },
            'cameras': cameras
        }

    except ExcelParseError:
        raise
    except Exception as e:
        raise ExcelParseError(f"解析Excel失败: {filepath}, 错误: {e}")

def extract_voltage_from_content(rows):
    """从内容中提取电压等级"""
    # 遍历前几行查找电压等级信息
    for row in rows[:10]:
        for cell in row:
            if cell:
                cell_str = str(cell).strip()
                # 匹配电压等级格式：220kV, 110kV, 35kV等
                if 'kV' in cell_str or 'kv' in cell_str.lower():
                    # 提取电压值
                    import re
                    match = re.search(r'(\d+)\s*[kK][vV]', cell_str)
                    if match:
                        return match.group(1) + 'kV'
    return ''

def extract_county_from_path(filepath):
    """从文件路径提取县区名称"""
    # 路径格式: .../图像监控设备资料/县名/变电站Excel
    parts = filepath.split(os.sep)
    for i, part in enumerate(parts):
        if part == '图像监控设备资料' and i + 1 < len(parts):
            return parts[i + 1]
    return ''

def _detect_camera_header_map(header_row):
    header_map = {
        'channel_col': 0,
        'location_col': 1,
        'ip_col': None,
        'area_col': None,
        'slot_code_col': None,
        'project_camera_code_col': None,
    }
    for index, cell in enumerate(header_row or []):
        if not cell:
            continue
        cell_str = str(cell).strip().lower()
        if '通道' in cell_str:
            header_map['channel_col'] = index
        elif '位置' in cell_str:
            header_map['location_col'] = index
        elif 'ip' in cell_str:
            header_map['ip_col'] = index
        elif '区域' in cell_str:
            header_map['area_col'] = index
        elif '槽位' in cell_str or 'slot' in cell_str or '点位' in cell_str:
            header_map['slot_code_col'] = index
        elif '设备编号' in cell_str or '摄像机编号' in cell_str or 'camera code' in cell_str:
            header_map['project_camera_code_col'] = index
    return header_map


def parse_camera_rows(rows):
    """
    解析摄像头区域和通道信息

    Excel有两种格式：
    格式A (丽水变.xlsx): 表头在Row 17，IP在Col 3
    格式B (四都变.xlsx): 表头在Row 16，IP在Col 4

    动态查找IP列位置：
    - 找到包含'通道'和'IP地址'的行作为表头
    - 确定IP列的索引位置
    """
    cameras = []

    # 查找表头行（包含'通道'和'IP地址'的行）
    header_row_idx = None
    header_map = None
    for idx, row in enumerate(rows):
        if not row:
            continue
        # 查找包含'通道'的列
        for col_idx, cell in enumerate(row):
            if cell and '通道' in str(cell).strip():
                # 找到通道列后，在同行的其他列中找IP地址
                for col_idx2, cell2 in enumerate(row):
                    cell2_str = str(cell2).strip().upper() if cell2 else ''
                    if cell2 and ('IP' in cell2_str):
                        header_row_idx = idx
                        header_map = _detect_camera_header_map(row)
                        header_map['ip_col'] = col_idx2
                        break
                if header_row_idx is not None and header_map is not None:
                    break
        if header_row_idx is not None and header_map is not None:
            break

    if header_row_idx is None or header_map is None or header_map['ip_col'] is None:
        return cameras

    # 从表头下一行开始解析摄像头数据
    for row_idx in range(header_row_idx + 1, len(rows)):
        row = rows[row_idx]
        if not row:
            continue

        camera_info = extract_camera_from_row(row, row_idx, header_map)
        if camera_info:
            cameras.append(camera_info)

    return cameras

def extract_camera_from_row(row, row_idx, header_map=None):
    """从单行数据中提取摄像头信息

    Args:
        row: 行数据
        row_idx: 行索引
        header_map: 列映射
    """
    if isinstance(header_map, int):
        header_map = {
            'channel_col': 0,
            'location_col': 1,
            'ip_col': header_map,
            'area_col': None,
            'slot_code_col': None,
            'project_camera_code_col': None,
        }
    if header_map is None:
        header_map = {
            'channel_col': 0,
            'location_col': 1,
            'ip_col': 3,
            'area_col': None,
            'slot_code_col': None,
            'project_camera_code_col': None,
        }
    ip_col_idx = header_map['ip_col']
    if not row or len(row) <= ip_col_idx:
        return None

    # 兼容“表头通道列在 C 列，但数据写在 A 列”的历史格式。
    channel_candidates = []
    preferred_index = header_map.get('channel_col')
    if preferred_index is not None and len(row) > preferred_index:
        channel_candidates.append(row[preferred_index])
    channel_candidates.extend(cell for cell in row if cell not in channel_candidates)

    channel_str = ''
    for candidate in channel_candidates:
        if candidate and '通道' in str(candidate).strip():
            channel_str = str(candidate).strip()
            break
    if not channel_str:
        return None
    if '通道' not in channel_str:
        return None

    # 提取通道号数字
    match = re.search(r'通道(\d+)', channel_str)
    camera_index = match.group(1) if match else ''

    # IP在动态确定的列
    ip_cell = row[ip_col_idx]
    if not ip_cell or not isinstance(ip_cell, str):
        return None

    ip = ip_cell.strip()
    if not re.match(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', ip):
        return None

    location_col = header_map['location_col']
    location = str(row[location_col]).strip() if len(row) > location_col and row[location_col] else ''
    area_col = header_map['area_col']
    area = str(row[area_col]).strip() if area_col is not None and len(row) > area_col and row[area_col] else ''
    slot_code_col = header_map['slot_code_col']
    slot_code = str(row[slot_code_col]).strip() if slot_code_col is not None and len(row) > slot_code_col and row[slot_code_col] else ''
    device_code_col = header_map['project_camera_code_col']
    project_camera_code = str(row[device_code_col]).strip() if device_code_col is not None and len(row) > device_code_col and row[device_code_col] else ''

    return {
        'camera_index': camera_index,
        'area': area,
        'location': location,
        'location_desc': location,
        'ip_address': ip,
        'channel_port': None,
        'channel_number': int(camera_index) if camera_index.isdigit() else None,
        'slot_code': slot_code,
        'project_camera_code': project_camera_code,
    }

def validate_excel_structure(filepath):
    """
    预检Excel结构
    返回: {'valid': bool, 'rows': int, 'errors': list}
    """
    result = {'valid': True, 'rows': 0, 'errors': []}

    try:
        rows = _load_excel_rows(filepath)
        result['rows'] = len(rows)

        if _is_flat_inventory_format(rows):
            station_values = [
                str(row[1]).strip()
                for row in rows[1:]
                if row and len(row) > 1 and row[1]
            ]
            if not station_values:
                result['valid'] = False
                result['errors'].append("智慧巡视台账缺少变电站名称")
            return result

        # 基本验证
        if result['rows'] < 10:
            result['errors'].append(f"数据行数过少: {result['rows']}")

        if not rows or not rows[0][0]:
            result['valid'] = False
            result['errors'].append("变电站名称为空")

    except Exception as e:
        result['valid'] = False
        result['errors'].append(str(e))

    return result

if __name__ == '__main__':
    # 测试
    import sys
    if len(sys.argv) > 1:
        result = validate_excel_structure(sys.argv[1])
        print(f"预检结果: {result}")
