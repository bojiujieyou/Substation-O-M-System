# test_parse_excel.py — Excel解析模块测试
import os
import sys
from types import SimpleNamespace
import pytest
import tempfile

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import parse_excel as parse_excel_module
from parse_excel import (
    parse_station_excel,
    validate_excel_structure,
    extract_voltage_from_content,
    extract_county_from_path,
    parse_camera_rows,
    extract_camera_from_row,
    ExcelParseError
)

class TestExtractVoltageFromContent:
    """测试电压等级提取"""

    def test_220kv(self):
        rows = [[None, '220kV变电站'], [None], [None]]
        assert extract_voltage_from_content(rows) == '220kV'

    def test_110kv(self):
        rows = [[None, None, '电压:110kV'], [None]]
        assert extract_voltage_from_content(rows) == '110kV'

    def test_no_voltage(self):
        rows = [[None, '测试'], [None]]
        assert extract_voltage_from_content(rows) == ''

    def test_empty_rows(self):
        assert extract_voltage_from_content([]) == ''
        assert extract_voltage_from_content([[None, None]]) == ''


class TestExtractCountyFromPath:
    """测试县区提取"""

    def test_lishui(self):
        path = r'e:\办公\图像监控\图像监控设备资料\丽水\某变电站.xlsx'
        assert extract_county_from_path(path) == '丽水'

    def test_longyan(self):
        path = r'e:\办公\图像监控\图像监控设备资料\龙泉\某变电站.xlsx'
        assert extract_county_from_path(path) == '龙泉'

    def test_no_county(self):
        path = r'e:\其他路径\某文件.xlsx'
        assert extract_county_from_path(path) == ''


class TestExtractCameraFromRow:
    """测试摄像头信息提取"""

    def test_valid_camera_row(self):
        """正确格式: 通道1, 位置描述, IP"""
        row = ['通道1', '220kV场地北侧-1#球', None, '192.168.1.100', None, '2018年']
        result = extract_camera_from_row(row, 20)
        assert result is not None
        assert result['ip_address'] == '192.168.1.100'
        assert result['camera_index'] == '1'
        assert result['location'] == '220kV场地北侧-1#球'

    def test_channel_10(self):
        """通道10的解析"""
        row = ['通道10', '35kV开关室1-10#球', None, '192.168.1.10', None, '2018年']
        result = extract_camera_from_row(row, 20)
        assert result is not None
        assert result['camera_index'] == '10'
        assert result['channel_number'] == 10

    def test_no_channel_header(self):
        """没有'通道'关键字的行应返回None"""
        row = ['设备', '某设备', None, '192.168.1.100', None, None]
        result = extract_camera_from_row(row, 20)
        assert result is None

    def test_invalid_ip(self):
        """IP格式错误"""
        row = ['通道1', '位置描述', None, 'invalid-ip', None, None]
        result = extract_camera_from_row(row, 20)
        assert result is None

    def test_empty_row(self):
        row = [None, None, None, None, None]
        result = extract_camera_from_row(row, 20)
        assert result is None

    def test_no_ip_column(self):
        """没有IP列"""
        row = ['通道1', '位置描述', '一些数据']
        result = extract_camera_from_row(row, 20)
        assert result is None


class TestValidateExcelStructure:
    """测试Excel结构验证"""

    def test_file_not_found(self):
        result = validate_excel_structure(r'C:\nonexistent\file.xlsx')
        assert result['valid'] == False
        assert len(result['errors']) > 0


class TestExcelParseError:
    """测试异常类型"""

    def test_is_exception(self):
        assert issubclass(ExcelParseError, Exception)

    def test_message(self):
        error = ExcelParseError('测试错误')
        assert str(error) == '测试错误'


class TestParseStationExcelFull:
    """端到端Excel解析测试"""

    def test_parse_valid_excel(self, tmp_path):
        """完整Excel解析：创建真实xlsx并解析"""
        from openpyxl import Workbook

        # 创建测试Excel文件（符合实际格式）
        wb = Workbook()
        ws = wb.active

        # 第1行：变电站名称
        ws['A1'] = '测试变电站'

        # 添加电压等级信息
        ws['A2'] = '220kV变电站设备清单'

        # 添加摄像头表头（第17行，与实际格式一致）
        ws['A17'] = '序号'
        ws['B17'] = '位置'
        ws['C17'] = '通道'
        ws['D17'] = 'IP地址'
        ws['E17'] = '备注'

        # 添加摄像头数据（格式：col A='通道N', col B=位置, col D=IP）
        ws['A18'] = '通道1'
        ws['B18'] = '220kV场地北侧-1#球'
        ws['D18'] = '192.168.1.100'

        ws['A19'] = '通道10'
        ws['B19'] = '35kV开关室1-10#球'
        ws['D19'] = '192.168.1.110'

        filepath = tmp_path / '测试变电站.xlsx'
        wb.save(filepath)

        # 解析
        result = parse_station_excel(str(filepath))

        # 验证结果结构
        assert 'station' in result
        assert 'cameras' in result
        assert result['station']['name'] == '测试变电站'
        assert len(result['cameras']) == 2

        # 验证第一个摄像头
        cam1 = result['cameras'][0]
        assert cam1['ip_address'] == '192.168.1.100'
        assert cam1['camera_index'] == '1'
        assert cam1['location'] == '220kV场地北侧-1#球'

    def test_parse_missing_station_name(self, tmp_path):
        """空变电站名称应抛出异常"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws['A1'] = None  # 变电站名称为空

        filepath = tmp_path / '空名称.xlsx'
        wb.save(filepath)

        with pytest.raises(ExcelParseError) as exc_info:
            parse_station_excel(str(filepath))
        assert '变电站名称为空' in str(exc_info.value)

    def test_parse_nonexistent_file(self):
        """不存在的文件应抛出异常"""
        with pytest.raises(ExcelParseError) as exc_info:
            parse_station_excel(r'C:\nonexistent\file.xlsx')
        assert '文件不存在' in str(exc_info.value)

    def test_parse_missing_header_row(self, tmp_path):
        """缺少表头行时返回空摄像头列表"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws['A1'] = '测试变电站'
        # 没有摄像头表头行

        filepath = tmp_path / '无表头.xlsx'
        wb.save(filepath)

        result = parse_station_excel(str(filepath))
        assert result['station']['name'] == '测试变电站'
        assert result['cameras'] == []

    def test_parse_flat_inventory_excel(self, tmp_path):
        """智慧巡视按设备逐行台账也能解析"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(['序号', '变电站', '设备名称', '设备型号', '生产厂家', '使用单位', '使用类型', '视频类型', '安装位置', '运行状态', '设备来源'])
        ws.append(['1', '110kV四都变电站', '四都变1-18-大门口-球机-1', 'DS-1', '海康威视', '', '球机', '可见光', '', '在线', '南瑞'])
        ws.append(['2', '110kV四都变电站', '四都变1-12-#1主变南侧-云台-33', 'DS-2', '海康威视', '', '云台', '可见光', '', '在线', '南瑞'])

        county_dir = tmp_path / '智慧巡视样本'
        county_dir.mkdir()
        filepath = county_dir / '110kV四都变电站.xlsx'
        wb.save(filepath)

        result = parse_station_excel(str(filepath))

        assert result['station']['name'] == '110kV四都变电站'
        assert result['station']['voltage_level'] == '110kV'
        assert result['station']['county'] == '智慧巡视样本'
        assert len(result['cameras']) == 2
        assert result['cameras'][0]['project_camera_code'] == '四都变1-18-大门口-球机-1'
        assert result['cameras'][0]['location_desc'] == '大门口'
        assert result['cameras'][0]['camera_index'] == '1'
        assert result['cameras'][1]['location_desc'] == '#1主变南侧'
        assert result['cameras'][1]['camera_index'] == '33'

    def test_parse_flat_inventory_excel_with_compact_device_names(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(['序号', '变电站', '设备名称', '设备型号', '使用类型'])
        ws.append(['1', '220kV仙都变电站', '220kV区域#1主变220kV开关旁#1云台', 'DS-1', '云台'])
        ws.append(['2', '220kV金亭变电站', 'A1-金亭变-110kV区域#1球机', 'DS-2', '球机'])

        county_dir = tmp_path / '智慧巡视样本'
        county_dir.mkdir()
        filepath = county_dir / 'inspection_compact.xlsx'
        wb.save(filepath)

        result = parse_station_excel(str(filepath))

        first = result['cameras'][0]
        second = result['cameras'][1]
        assert first['location_desc'] == '220kV区域#1主变220kV开关旁#1'
        assert first['camera_index'] == '1'
        assert second['location_desc'] == '金亭变-110kV区域#1'
        assert second['camera_index'] == '1'

    def test_parse_flat_inventory_keeps_video_type_and_semantic_prefix(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(['序号', '变电站', '设备名称', '设备型号', '使用类型', '视频类型'])
        ws.append(['1', '220kV松阳变电站', '红外-松阳变-#2主变西北侧#50测温球机', 'DS-1', '球机', '测温'])
        ws.append(['2', '220kV青田变电站', '#1电容器旁#1热成像球机-可见光', 'DS-2', '球机', '可见光'])

        county_dir = tmp_path / '智慧巡视样本'
        county_dir.mkdir()
        filepath = county_dir / 'inspection_video_type.xlsx'
        wb.save(filepath)

        result = parse_station_excel(str(filepath))
        first = result['cameras'][0]
        second = result['cameras'][1]

        assert first['location_desc'] == '红外-松阳变-#2主变西北侧#50'
        assert first['camera_index'] == '50'
        assert first['area'] == '球机/测温'
        assert second['location_desc'] == '#1电容器旁#1热成像'
        assert second['camera_index'] == '1'
        assert second['area'] == '球机/可见光'

    def test_validate_flat_inventory_excel(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(['序号', '变电站', '设备名称'])
        ws.append(['1', '220kV演练变电站', '演练变1-主变东-球机-1'])
        filepath = tmp_path / 'inspection_flat.xlsx'
        wb.save(filepath)

        result = validate_excel_structure(str(filepath))
        assert result['valid'] is True
        assert result['errors'] == []

    def test_parse_valid_legacy_xls(self, tmp_path, monkeypatch):
        class FakeSheet:
            def __init__(self, rows):
                self._rows = rows
                self.nrows = len(rows)

            def row_values(self, index):
                return self._rows[index]

        class FakeBook:
            def __init__(self, rows):
                self._sheet = FakeSheet(rows)
                self.released = False

            def sheet_by_index(self, index):
                assert index == 0
                return self._sheet

            def release_resources(self):
                self.released = True

        rows = [
            ['测试旧版变电站'],
            ['220kV变电站设备清单'],
            [],
            ['序号', '位置', '通道', 'IP地址'],
            ['通道1', '主变北侧', '', '192.168.10.1'],
            ['通道2', '主变南侧', '', '192.168.10.2'],
        ]
        workbook = FakeBook(rows)
        filepath = tmp_path / 'legacy_station.xls'
        filepath.write_bytes(b'legacy-xls')

        monkeypatch.setattr(
            parse_excel_module,
            'xlrd',
            SimpleNamespace(open_workbook=lambda _path, on_demand=True: workbook),
        )

        result = parse_station_excel(str(filepath))

        assert result['station']['name'] == '测试旧版变电站'
        assert result['station']['voltage_level'] == '220kV'
        assert len(result['cameras']) == 2
        assert result['cameras'][0]['ip_address'] == '192.168.10.1'
        assert result['cameras'][1]['camera_index'] == '2'
        assert workbook.released is True

    def test_validate_legacy_xls_structure(self, tmp_path, monkeypatch):
        class FakeSheet:
            def __init__(self, rows):
                self._rows = rows
                self.nrows = len(rows)

            def row_values(self, index):
                return self._rows[index]

        class FakeBook:
            def __init__(self, rows):
                self._sheet = FakeSheet(rows)

            def sheet_by_index(self, index):
                assert index == 0
                return self._sheet

            def release_resources(self):
                return None

        rows = [
            ['序号', '变电站', '设备名称'],
            ['1', '220kV演练变电站', '演练变1-主变东-球机-1'],
        ]
        filepath = tmp_path / 'legacy_inventory.xls'
        filepath.write_bytes(b'legacy-inventory')

        monkeypatch.setattr(
            parse_excel_module,
            'xlrd',
            SimpleNamespace(open_workbook=lambda _path, on_demand=True: FakeBook(rows)),
        )

        result = validate_excel_structure(str(filepath))

        assert result['valid'] is True
        assert result['rows'] == 2
        assert result['errors'] == []

    def test_parse_legacy_xls_requires_xlrd(self, tmp_path, monkeypatch):
        filepath = tmp_path / 'missing-xlrd.xls'
        filepath.write_bytes(b'legacy-xls')
        monkeypatch.setattr(parse_excel_module, 'xlrd', None)

        with pytest.raises(ExcelParseError) as exc_info:
            parse_station_excel(str(filepath))

        assert '未安装xlrd' in str(exc_info.value)
