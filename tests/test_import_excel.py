import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook

from app import app
from import_excel import run_batch_import
from init_db import init_db


@pytest.fixture
def import_excel_db(tmp_path):
    return str(tmp_path / "test_import_excel.db")


@pytest.fixture
def seeded_import_excel_schema(import_excel_db):
    import config as config_module

    original_path = config_module.Config.DATABASE_PATH
    app_original_db_path = app.config["DATABASE_PATH"]
    config_module.Config.DATABASE_PATH = import_excel_db
    app.config["DATABASE_PATH"] = import_excel_db
    init_db(force=True)

    conn = sqlite3.connect(import_excel_db)
    conn.executescript(
        """
        CREATE TABLE projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            short_name TEXT NOT NULL,
            color TEXT NOT NULL DEFAULT '#1a73e8',
            fault_type_version_id INTEGER,
            sort_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE camera_slots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slot_code TEXT NOT NULL,
            station_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            location_desc TEXT NOT NULL DEFAULT '',
            area TEXT NOT NULL DEFAULT '',
            channel_number INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(station_id, project_id, slot_code)
        );

        ALTER TABLE cameras ADD COLUMN slot_id INTEGER;
        ALTER TABLE cameras ADD COLUMN project_id INTEGER;
        ALTER TABLE cameras ADD COLUMN project_camera_code TEXT;
        ALTER TABLE cameras ADD COLUMN status TEXT DEFAULT 'active';
        ALTER TABLE cameras ADD COLUMN replaced_by_camera_id INTEGER;
        ALTER TABLE cameras ADD COLUMN retired_at TIMESTAMP;

        DROP TABLE cameras;
        CREATE TABLE cameras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slot_id INTEGER,
            station_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            project_camera_code TEXT,
            camera_index TEXT,
            area TEXT,
            location_desc TEXT,
            ip_address TEXT,
            channel_port INTEGER,
            channel_number INTEGER,
            status TEXT DEFAULT 'active',
            replaced_by_camera_id INTEGER,
            retired_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE UNIQUE INDEX idx_cameras_one_active_per_slot
        ON cameras(slot_id)
        WHERE status = 'active';

        CREATE TABLE import_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            source_type TEXT NOT NULL,
            mode TEXT NOT NULL,
            file_count INTEGER,
            success_count INTEGER,
            fail_count INTEGER,
            report_path TEXT,
            operator_id INTEGER,
            timezone_default_used TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    conn.execute(
        """
        INSERT INTO projects (id, code, name, short_name, color, sort_order, is_active)
        VALUES
            (1, 'unified', 'Unified', 'UNI', '#1a73e8', 1, 1),
            (2, 'inspection', 'Inspection', 'INSP', '#34a853', 2, 1)
        """
    )
    conn.commit()
    conn.close()

    yield import_excel_db

    config_module.Config.DATABASE_PATH = original_path
    app.config["DATABASE_PATH"] = app_original_db_path


def build_inventory_workbook(path: Path, station_name: str, cameras: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = station_name
    ws["A2"] = "110kV变电站"
    ws["A17"] = "通道"
    ws["B17"] = "位置"
    ws["C17"] = "区域"
    ws["D17"] = "IP地址"
    ws["E17"] = "槽位编码"
    ws["F17"] = "设备编号"

    start_row = 18
    for offset, camera in enumerate(cameras):
        row = start_row + offset
        ws[f"A{row}"] = f"通道{camera['channel']}"
        ws[f"B{row}"] = camera["location"]
        ws[f"C{row}"] = camera.get("area", "")
        ws[f"D{row}"] = camera["ip"]
        ws[f"E{row}"] = camera.get("slot_code", "")
        ws[f"F{row}"] = camera.get("project_camera_code", "")

    wb.save(path)


def test_run_batch_import_uses_multi_project_slot_model(seeded_import_excel_schema, tmp_path):
    source_root = tmp_path / "source"
    county_dir = source_root / "丽水"
    county_dir.mkdir(parents=True)
    workbook = county_dir / "station_a.xlsx"
    build_inventory_workbook(
        workbook,
        "Station A",
        [
            {"channel": 1, "location": "A-1", "area": "Area A", "ip": "10.0.0.1", "slot_code": "SLOT_A_1", "project_camera_code": "CAM-A-1"},
            {"channel": 2, "location": "A-2", "area": "Area A", "ip": "10.0.0.2", "slot_code": "SLOT_A_2", "project_camera_code": "CAM-A-2"},
        ],
    )
    report_path = tmp_path / "report.json"

    report = run_batch_import(
        database=seeded_import_excel_schema,
        source_root=str(source_root),
        project_code="inspection",
        report_path=report_path,
    )

    assert report["success_count"] == 1
    assert report["fail_count"] == 0
    assert report_path.exists()

    conn = sqlite3.connect(seeded_import_excel_schema)
    try:
        station = conn.execute(
            "SELECT id, county FROM stations WHERE name = 'Station A' AND voltage_level = '110kV'"
        ).fetchone()
        slot_count = conn.execute("SELECT COUNT(*) FROM camera_slots WHERE project_id = 2").fetchone()[0]
        camera_rows = conn.execute(
            "SELECT project_id, project_camera_code, status FROM cameras ORDER BY id"
        ).fetchall()
        batch = conn.execute(
            "SELECT project_id, source_type, mode, file_count, success_count, fail_count, report_path FROM import_batches"
        ).fetchone()
    finally:
        conn.close()

    assert station[1] == "丽水"
    assert slot_count == 2
    assert camera_rows == [(2, "CAM-A-1", "active"), (2, "CAM-A-2", "active")]
    assert batch == (2, "import_excel", "best-effort", 1, 1, 0, str(report_path))


def test_run_batch_import_replaces_camera_within_same_slot(seeded_import_excel_schema, tmp_path):
    source_root = tmp_path / "source"
    county_dir = source_root / "丽水"
    county_dir.mkdir(parents=True)
    workbook = county_dir / "station_replace.xlsx"

    build_inventory_workbook(
        workbook,
        "Station Replace",
        [
            {"channel": 1, "location": "Replace Slot", "area": "Area", "ip": "10.0.0.1", "slot_code": "SLOT_R_1", "project_camera_code": "OLD-CAM"},
        ],
    )
    run_batch_import(
        database=seeded_import_excel_schema,
        source_root=str(source_root),
        project_code="inspection",
    )

    build_inventory_workbook(
        workbook,
        "Station Replace",
        [
            {"channel": 1, "location": "Replace Slot", "area": "Area", "ip": "10.0.0.9", "slot_code": "SLOT_R_1", "project_camera_code": "NEW-CAM"},
        ],
    )
    report = run_batch_import(
        database=seeded_import_excel_schema,
        source_root=str(source_root),
        project_code="inspection",
    )

    assert report["success_count"] == 1

    conn = sqlite3.connect(seeded_import_excel_schema)
    try:
        rows = conn.execute(
            """
            SELECT project_camera_code, ip_address, status, replaced_by_camera_id
            FROM cameras
            WHERE station_id = (SELECT id FROM stations WHERE name = 'Station Replace')
            ORDER BY id
            """
        ).fetchall()
    finally:
        conn.close()

    assert rows[0][0] == "OLD-CAM"
    assert rows[0][2] == "replaced"
    assert rows[0][3] == 2
    assert rows[1] == ("NEW-CAM", "10.0.0.9", "active", None)


def test_run_batch_import_dry_run_rolls_back(seeded_import_excel_schema, tmp_path):
    source_root = tmp_path / "source"
    county_dir = source_root / "丽水"
    county_dir.mkdir(parents=True)
    workbook = county_dir / "station_dry_run.xlsx"
    build_inventory_workbook(
        workbook,
        "Station Dry Run",
        [{"channel": 1, "location": "Dry Slot", "area": "Area", "ip": "10.0.1.1", "slot_code": "SLOT_D_1"}],
    )

    report = run_batch_import(
        database=seeded_import_excel_schema,
        source_root=str(source_root),
        project_code="inspection",
        dry_run=True,
    )

    assert report["dry_run"] is True
    assert report["success_count"] == 1

    conn = sqlite3.connect(seeded_import_excel_schema)
    try:
        station_count = conn.execute("SELECT COUNT(*) FROM stations WHERE name = 'Station Dry Run'").fetchone()[0]
        batch_count = conn.execute("SELECT COUNT(*) FROM import_batches").fetchone()[0]
    finally:
        conn.close()

    assert station_count == 0
    assert batch_count == 0


def test_run_batch_import_full_rollback_aborts_on_error(seeded_import_excel_schema, tmp_path):
    source_root = tmp_path / "source"
    county_dir = source_root / "丽水"
    county_dir.mkdir(parents=True)
    good = county_dir / "good.xlsx"
    bad = county_dir / "bad.xlsx"
    build_inventory_workbook(
        good,
        "Station Good",
        [{"channel": 1, "location": "Good Slot", "area": "Area", "ip": "10.0.2.1", "slot_code": "SLOT_G_1"}],
    )
    Workbook().save(bad)

    report = run_batch_import(
        database=seeded_import_excel_schema,
        source_root=str(source_root),
        project_code="inspection",
        mode="full-rollback",
    )

    assert report["aborted"] is True
    assert report["fail_count"] == 1

    conn = sqlite3.connect(seeded_import_excel_schema)
    try:
        station_count = conn.execute("SELECT COUNT(*) FROM stations WHERE name = 'Station Good'").fetchone()[0]
        camera_count = conn.execute("SELECT COUNT(*) FROM cameras").fetchone()[0]
        batch_count = conn.execute("SELECT COUNT(*) FROM import_batches").fetchone()[0]
    finally:
        conn.close()

    assert station_count == 0
    assert camera_count == 0
    assert batch_count == 0
