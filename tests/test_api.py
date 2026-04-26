# test_api.py — Flask API集成测试
import os
import sys
import pytest

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import sqlite3
from io import BytesIO
from openpyxl import load_workbook
from app import app, get_db, ensure_fault_report_multi_camera_schema

from config import Config



def _with_csrf(client, kwargs):
    """自动为状态变更请求注入 CSRF token。"""
    headers = dict(kwargs.pop("headers", {}) or {})
    with client.session_transaction() as sess:
        token = sess.get("csrf_token")
    if token:
        headers.setdefault("X-CSRF-Token", token)
    kwargs["headers"] = headers
    return kwargs


def _patch_client(client):
    """给 test client 的状态变更方法自动注入 CSRF token。"""
    _orig_post = client.post
    _orig_put = client.put
    _orig_delete = client.delete

    def _post(url, **kwargs):
        return _orig_post(url, **_with_csrf(client, kwargs))

    def _put(url, **kwargs):
        return _orig_put(url, **_with_csrf(client, kwargs))

    def _delete(url, **kwargs):
        return _orig_delete(url, **_with_csrf(client, kwargs))

    client.post = _post
    client.put = _put
    client.delete = _delete
    return client


def login_admin_session(client):
    with client.session_transaction() as session:
        session['user_id'] = 1
        session['username'] = 'admin'
        session['role'] = 'admin'
        session['csrf_token'] = 'test-csrf-token'


def ensure_unified_project(test_db):
    conn = sqlite3.connect(test_db)
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1
            )
            """
        )
        conn.execute(
            """
            INSERT OR IGNORE INTO projects (id, code, name, is_active)
            VALUES (1, 'unified', '统一项目', 1)
            """
        )
        conn.commit()
    finally:
        conn.close()



@pytest.fixture
def client(test_db):
    """测试客户端"""
    # Override app.config directly since app.config.from_object(Config)
    # copies DATABASE_PATH at import time (before monkeypatch runs)
    original_db_path = app.config['DATABASE_PATH']
    app.config['DATABASE_PATH'] = test_db
    app.config['TESTING'] = True
    with app.test_client() as client:
        _patch_client(client)
        yield client
    app.config['DATABASE_PATH'] = original_db_path

@pytest.fixture
def init_db(test_db):
    """初始化测试数据库"""
    from init_db import init_db
    # DATABASE_PATH已通过conftest.py的use_test_db自动指向test_db
    # force=True确保即使数据库已存在也能重建（test_db fixture已删除旧文件）
    init_db(force=True)
    yield

def get_token():
    """获取认证token"""
    return Config.API_TOKEN

class TestHealthEndpoint:
    """健康检查端点"""

    def test_health(self, client, init_db):
        response = client.get('/health')
        assert response.status_code == 200
        data = response.get_json()
        assert data['status'] == 'ok'

class TestStatsEndpoint:
    """统计端点"""

    def test_get_stats_empty(self, client, init_db):
        login_admin_session(client)
        response = client.get('/api/stats')
        assert response.status_code == 200
        data = response.get_json()
        assert data['stations'] == 0
        assert data['cameras'] == 0
        assert data['faults'] == 0
        assert data['fault_events'] == 0
        assert len(data['monthly_trend']) == 12
        assert all(item['event_count'] == 0 for item in data['monthly_trend'])
        assert data['county_distribution'] == []


    def test_get_stats_batch_impact_deduplicates_fault_group(self, client, init_db, test_db):
        login_admin_session(client)
        with app.app_context():
            ensure_fault_report_multi_camera_schema(get_db())

        conn = sqlite3.connect(test_db)
        try:
            conn.execute(
                """
                INSERT INTO stations (id, name, voltage_level, county)
                VALUES (1, '测试变电站', '220kV', '测试县')
                """
            )
            conn.execute(
                """
                INSERT INTO cameras (id, station_id, camera_index, area, location_desc, ip_address, channel_port, channel_number)
                VALUES (1, 1, 'CAM-01', '主变区', '主变区1号位', '192.168.1.10', 8000, 1)
                """
            )

            conn.executemany(
                """
                INSERT INTO fault_reports (
                    id, station_id, camera_id, fault_type, reporter_name, status,
                    fault_group_key, fault_owner_type, root_cause_type,
                    is_batch_impact, impact_camera_count, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,

                [
                    (1, 1, 1, '无图像', '张三', 'closed', 'group-a', 'camera', 'camera', 1, 2, '2026-04-25 10:00:00', '2026-04-25 10:30:00'),
                    (2, 1, 1, '无图像', '张三', 'closed', 'group-a', 'camera', 'camera', 1, 2, '2026-04-25 10:00:00', '2026-04-25 10:30:00'),
                    (3, 1, 1, '无图像', '李四', 'open', None, 'camera', 'camera', 0, 1, '2026-04-25 11:00:00', '2026-04-25 11:00:00')


                ]

            )
            conn.commit()
        finally:
            conn.close()

        response = client.get('/api/stats?year=2026')
        assert response.status_code == 200
        data = response.get_json()

        assert data['faults'] == 3
        assert data['fault_events'] == 2
        assert data['batch_impact_event_count'] == 1
        assert data['batch_impact']['batch_fault_count'] == 1
        assert data['batch_impact']['avg_impact_cameras'] == 2.0
        assert data['batch_impact']['total_impact_cameras'] == 2
        assert data['kpi']['open_count'] == 1
        assert data['kpi']['closed_count'] == 2
        assert data['root_cause_distribution'][0]['cause_type'] == 'camera'
        assert data['root_cause_distribution'][0]['count'] == 2
        assert data['fault_type_distribution'][0]['fault_label'] == '无图像'
        assert data['fault_type_distribution'][0]['count'] == 3


        monthly_map = {item['month']: item for item in data['monthly_trend']}

        assert monthly_map['2026-04']['count'] == 3
        assert monthly_map['2026-04']['event_count'] == 2

        assert len(data['county_distribution']) == 1
        assert data['county_distribution'][0]['county'] == '测试县'
        assert data['county_distribution'][0]['count'] == 3
        assert data['county_distribution'][0]['event_count'] == 2

        assert len(data['station_ranking']) == 1

        assert data['station_ranking'][0]['fault_count'] == 3
        assert data['station_ranking'][0]['fault_event_count'] == 2
        assert data['station_ranking'][0]['unresolved_count'] == 1

        assert len(data['camera_ranking']) == 1
        assert data['camera_ranking'][0]['fault_count'] == 3
        assert data['camera_ranking'][0]['fault_event_count'] == 2








def test_get_stats_camera_ranking_uses_fault_report_cameras_for_aggregated_faults(client, init_db, test_db):
    login_admin_session(client)
    with app.app_context():
        ensure_fault_report_multi_camera_schema(get_db())

    conn = sqlite3.connect(test_db)
    try:
        conn.execute(
            """
            INSERT INTO stations (id, name, voltage_level, county)
            VALUES (1, '测试变电站', '220kV', '测试县')
            """
        )
        conn.executemany(
            """
            INSERT INTO cameras (id, station_id, camera_index, area, location_desc, ip_address, channel_port, channel_number)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (1, 1, 'CAM-01', '主变区', '主变区1号位', '192.168.1.10', 8000, 1),
                (2, 1, 'CAM-02', '主变区', '主变区2号位', '192.168.1.11', 8000, 2),
            ],
        )
        conn.execute(
            """
            INSERT INTO fault_reports (
                id, station_id, camera_id, fault_type, reporter_name, status,
                fault_group_key, fault_owner_type, root_cause_type,
                is_batch_impact, impact_camera_count, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (1, 1, 1, '无图像', '张三', 'closed', 'group-a', 'camera', 'camera', 0, 2, '2026-04-25 10:00:00', '2026-04-25 10:30:00'),
        )
        conn.executemany(
            """
            INSERT INTO fault_report_cameras (
                fault_report_id, camera_id, camera_label, recovery_state,
                detail_fault_reason, detail_resolution, detail_note
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (1, 1, '主变区1号位', 'resolved', '镜头起雾', '已擦拭', '恢复正常'),
                (1, 2, '主变区2号位', 'self_recovered', '瞬时抖动', '观察恢复', '自动恢复'),
            ],
        )
        conn.commit()
    finally:
        conn.close()

    response = client.get('/api/stats?year=2026')
    assert response.status_code == 200
    data = response.get_json()

    ranking = data['camera_ranking']
    assert len(ranking) == 2
    assert ranking[0]['camera_location'] == '主变区1号位'
    assert ranking[0]['fault_count'] == 1
    assert ranking[0]['fault_event_count'] == 1
    assert ranking[1]['camera_location'] == '主变区2号位'
    assert ranking[1]['fault_count'] == 1
    assert ranking[1]['fault_event_count'] == 1




def test_export_statistics_county_sheet_contains_record_and_event_counts(client, init_db, test_db):
    login_admin_session(client)
    with app.app_context():
        ensure_fault_report_multi_camera_schema(get_db())


    conn = sqlite3.connect(test_db)
    try:
        conn.execute(
            """
            INSERT INTO stations (id, name, voltage_level, county)
            VALUES (1, '测试变电站', '220kV', '测试县')
            """
        )
        conn.executemany(
            """
            INSERT INTO fault_reports (
                id, station_id, fault_type, reporter_name, status,
                fault_group_key, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (1, 1, '无图像', '张三', 'closed', 'group-a', '2026-04-25 10:00:00', '2026-04-25 10:30:00'),
                (2, 1, '无图像', '张三', 'closed', 'group-a', '2026-04-25 10:05:00', '2026-04-25 10:35:00'),
                (3, 1, '无图像', '李四', 'open', None, '2026-04-25 11:00:00', '2026-04-25 11:00:00'),
            ]
        )
        conn.commit()
    finally:
        conn.close()

    response = client.get('/api/statistics/export?year=2026')
    assert response.status_code == 200

    workbook = load_workbook(filename=BytesIO(response.data))

    overview_sheet = workbook['概览']
    assert overview_sheet['A1'].value == '指标'
    assert overview_sheet['B1'].value == '数值'
    assert overview_sheet['A4'].value == '故障记录数'
    assert overview_sheet['B4'].value == 3
    assert overview_sheet['A5'].value == '独立故障事件数'
    assert overview_sheet['B5'].value == 2

    monthly_sheet = workbook['月度趋势']
    assert monthly_sheet['A1'].value == '月份'
    assert monthly_sheet['B1'].value == '故障记录数'
    assert monthly_sheet['C1'].value == '独立事件数'
    assert monthly_sheet['A5'].value == '2026-04'
    assert monthly_sheet['B5'].value == 3
    assert monthly_sheet['C5'].value == 2

    county_sheet = workbook['县区统计']
    assert county_sheet['A1'].value == '县区'
    assert county_sheet['B1'].value == '故障记录数'
    assert county_sheet['C1'].value == '独立事件数'
    assert county_sheet['A2'].value == '测试县'
    assert county_sheet['B2'].value == 3
    assert county_sheet['C2'].value == 2

    voltage_sheet = workbook['电压等级统计']
    assert voltage_sheet['A1'].value == '电压等级'
    assert voltage_sheet['B1'].value == '故障记录数'
    assert voltage_sheet['C1'].value == '独立事件数'
    assert voltage_sheet['A2'].value == '220kV'
    assert voltage_sheet['B2'].value == 3
    assert voltage_sheet['C2'].value == 2





def test_export_statistics_detail_sheet_contains_multi_camera_summary(client, init_db, test_db):
    login_admin_session(client)
    with app.app_context():
        ensure_fault_report_multi_camera_schema(get_db())

    conn = sqlite3.connect(test_db)
    try:
        conn.execute(
            """
            INSERT INTO stations (id, name, voltage_level, county)
            VALUES (1, '测试变电站', '220kV', '测试县')
            """
        )
        conn.executemany(
            """
            INSERT INTO cameras (id, station_id, camera_index, area, location_desc, ip_address, channel_port, channel_number)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (1, 1, 'CAM-01', '主变区', '主变区1号位', '192.168.1.10', 8000, 1),
                (2, 1, 'CAM-02', '主变区', '主变区2号位', '192.168.1.11', 8000, 2),
            ],
        )
        conn.execute(
            """
            INSERT INTO fault_reports (
                id, station_id, camera_id, fault_type, reporter_name, reporter_contact, status,
                fault_group_key, fault_owner_type, root_cause_type, is_batch_impact,
                impact_camera_count, created_at, updated_at, closed_at, handler_name, handler_note
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                1, 1, 1, '无图像', '张三', '13800000000', 'closed',
                'group-a', 'camera', 'camera', 1,
                2, '2026-04-25 10:00:00', '2026-04-25 10:30:00', '2026-04-25 10:30:00', '李四', '现场处理完成'
            ),
        )
        conn.executemany(
            """
            INSERT INTO fault_report_cameras (
                fault_report_id, camera_id, camera_label, recovery_state,
                detail_fault_reason, detail_resolution, detail_note
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (1, 1, '主变区1号位', 'resolved', '镜头起雾', '已擦拭', '恢复正常'),
                (1, 2, '主变区2号位', 'self_recovered', '瞬时抖动', '观察恢复', '自动恢复'),
            ],
        )
        conn.commit()
    finally:
        conn.close()

    response = client.get('/api/statistics/export?year=2026')
    assert response.status_code == 200

    workbook = load_workbook(filename=BytesIO(response.data))
    detail_sheet = workbook['故障明细']
    headers = [cell.value for cell in detail_sheet[1]]
    row = [cell.value for cell in detail_sheet[2]]
    row_map = dict(zip(headers, row))

    assert row_map['摄像头位置'] == '主变区1号位、主变区2号位'
    assert row_map['摄像头明细'] == '主变区1号位、主变区2号位'
    assert row_map['逐路恢复摘要'] == '主变区1号位（已修复）；主变区2号位（自恢复）'
    assert row_map['自恢复路数'] == 1
    assert row_map['影响摄像头数'] == 2


    def test_get_cameras_empty(self, client, init_db):
        response = client.get('/api/cameras')
        assert response.status_code == 200
        data = response.get_json()
        assert data['cameras'] == []
        assert data['total'] == 0

    def test_get_camera_by_ip_not_found(self, client, init_db):
        response = client.get('/api/cameras/by-ip?ip=192.168.1.1')
        assert response.status_code == 404
        data = response.get_json()
        assert '该IP暂未录入系统' in data['error']

    def test_get_camera_by_ip_no_param(self, client, init_db):
        response = client.get('/api/cameras/by-ip')
        assert response.status_code == 400

class TestFaultsEndpoint:
    """故障端点"""

    def test_get_faults_empty(self, client, init_db):
        response = client.get('/api/faults')
        assert response.status_code == 200
        data = response.get_json()
        assert data['faults'] == []
        assert data['total'] == 0

    def test_create_fault_missing_fields(self, client, init_db):
        """缺少必填字段"""
        response = client.post('/api/faults', json={})
        assert response.status_code == 400

        response = client.post('/api/faults', json={'station_id': 1})
        assert response.status_code == 400

    def test_create_fault_station_not_found(self, client, init_db):
        """变电站不存在"""
        response = client.post('/api/faults', json={
            'station_id': 999,
            'fault_type': '无图像',
            'reporter_name': '张三'
        })
        assert response.status_code == 404

    def test_get_faults_prioritizes_unclosed_before_closed(self, client, init_db, test_db):
        conn = sqlite3.connect(test_db)
        conn.execute(
            """
            INSERT INTO stations (id, name, voltage_level, county)
            VALUES (1, '测试变电站', '110kV', '测试县')
            """
        )
        conn.execute(
            """
            INSERT INTO fault_reports (id, station_id, fault_type, reporter_name, status, created_at, updated_at)
            VALUES (1, 1, '设备故障', 'A', 'closed', '2026-04-08 10:00:00', '2026-04-08 10:00:00')
            """
        )
        conn.execute(
            """
            INSERT INTO fault_reports (id, station_id, fault_type, reporter_name, status, created_at, updated_at)
            VALUES (2, 1, '设备故障', 'B', 'open', '2026-04-08 09:00:00', '2026-04-08 09:00:00')
            """
        )
        conn.execute(
            """
            INSERT INTO fault_reports (id, station_id, fault_type, reporter_name, status, created_at, updated_at)
            VALUES (3, 1, '设备故障', 'C', 'handling', '2026-04-08 08:00:00', '2026-04-08 08:00:00')
            """
        )
        conn.commit()
        conn.close()

        response = client.get('/api/faults')
        assert response.status_code == 200
        faults = response.get_json()['faults']
        assert [fault['id'] for fault in faults] == [2, 3, 1]

    def test_update_fault_detail_fields(self, client, init_db, test_db):
        conn = sqlite3.connect(test_db)
        conn.execute(
            """
            INSERT INTO stations (id, name, voltage_level, county)
            VALUES (1, 'Test Station', '110kV', 'Test County')
            """
        )
        conn.execute(
            """
            INSERT INTO fault_reports (
                id, station_id, fault_type, description, reporter_name, reporter_contact,
                handler_name, handler_note, status, created_at, updated_at
            )
            VALUES (
                1, 1, 'Legacy Type', 'legacy description', 'legacy reporter', '10086',
                'legacy handler', 'legacy note', 'closed', '2026-04-08 10:00:00', '2026-04-08 10:00:00'
            )
            """
        )
        conn.commit()
        conn.close()

        response = client.put('/api/faults/1', json={
            'fault_type': 'Updated Type',
            'description': 'updated description',
            'camera_location_text': 'updated location',
            'reporter_name': 'updated reporter',
            'reporter_contact': '10010',
            'handler_name': 'updated handler',
            'handler_note': 'updated note',
        })
        assert response.status_code == 200

        conn = sqlite3.connect(test_db)
        try:
            row = conn.execute(
                """
                SELECT fault_type, description, camera_location_text, reporter_name, reporter_contact, handler_name, handler_note
                FROM fault_reports
                WHERE id = 1
                """
            ).fetchone()
        finally:
            conn.close()

        assert row == (
            'Updated Type',
            'updated description',
            'updated location',
            'updated reporter',
            '10010',
            'updated handler',
            'updated note',
        )

    def test_fault_camera_location_falls_back_to_description(self, client, init_db, test_db):
        conn = sqlite3.connect(test_db)
        conn.execute(
            """
            INSERT INTO stations (id, name, voltage_level, county)
            VALUES (1, '汤公变', '220kV', '遂昌')
            """
        )
        conn.execute(
            """
            INSERT INTO fault_reports (
                id, station_id, fault_type, description, reporter_name, status, created_at, updated_at
            )
            VALUES (
                1, 1, '设备故障', '西南角13#摄像机排查维修后恢复 | 地点: 遂昌', '工作记录导入',
                'closed', '2026-04-08 10:00:00', '2026-04-08 10:00:00'
            )
            """
        )
        conn.commit()
        conn.close()

        list_response = client.get('/api/faults')
        assert list_response.status_code == 200
        list_fault = list_response.get_json()['faults'][0]
        assert list_fault['camera_location'] == '西南角13#摄像机'

        detail_response = client.get('/api/faults/1/detail')
        assert detail_response.status_code == 200
        detail_fault = detail_response.get_json()['fault']
        assert detail_fault['camera_location'] == '西南角13#摄像机'
        assert detail_fault['camera_location_text'] == '西南角13#摄像机'

    def test_delete_fault_moves_record_to_trash_and_can_restore(self, client, init_db, test_db):
        login_admin_session(client)

        conn = sqlite3.connect(test_db)
        conn.execute(
            """
            INSERT INTO stations (id, name, voltage_level, county)
            VALUES (1, 'Test Station', '110kV', 'Test County')
            """
        )
        conn.execute(
            """
            INSERT INTO fault_reports (id, station_id, fault_type, reporter_name, status, created_at, updated_at)
            VALUES (1, 1, 'No Image', 'tester', 'open', '2026-04-09 09:00:00', '2026-04-09 09:00:00')
            """
        )
        conn.commit()
        conn.close()

        delete_response = client.delete('/api/faults/1')
        assert delete_response.status_code == 200

        active_list = client.get('/api/faults')
        assert active_list.status_code == 200
        assert active_list.get_json()['faults'] == []

        deleted_detail = client.get('/api/faults/1/detail?deleted=only')
        assert deleted_detail.status_code == 200
        assert deleted_detail.get_json()['fault']['deleted_at'] is not None

        trash_list = client.get('/api/faults?deleted=only')
        assert trash_list.status_code == 200
        trash_faults = trash_list.get_json()['faults']
        assert len(trash_faults) == 1
        assert trash_faults[0]['id'] == 1
        assert trash_faults[0]['deleted_at'] is not None

        restore_response = client.post('/api/faults/1/restore')
        assert restore_response.status_code == 200

        restored_list = client.get('/api/faults')
        assert restored_list.status_code == 200
        restored_faults = restored_list.get_json()['faults']
        assert len(restored_faults) == 1
        assert restored_faults[0]['id'] == 1
        assert restored_faults[0]['deleted_at'] is None

    def test_delete_station_cleans_related_records_and_unlinks_photos(self, client, init_db, test_db):
        login_admin_session(client)

        conn = sqlite3.connect(test_db)
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS camera_slots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                slot_code TEXT NOT NULL,
                station_id INTEGER NOT NULL,
                project_id INTEGER NOT NULL,
                location_desc TEXT NOT NULL DEFAULT '',
                area TEXT NOT NULL DEFAULT '',
                channel_number INTEGER,
                FOREIGN KEY (station_id) REFERENCES stations(id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS station_external_names (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                station_id INTEGER NOT NULL,
                source_system TEXT NOT NULL,
                external_name TEXT NOT NULL,
                normalized_name TEXT NOT NULL,
                is_primary INTEGER DEFAULT 0,
                FOREIGN KEY (station_id) REFERENCES stations(id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS station_recorders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                station_id INTEGER NOT NULL,
                project_id INTEGER NOT NULL,
                recorder_name TEXT NOT NULL,
                ip_address TEXT,
                port INTEGER,
                description TEXT,
                status TEXT NOT NULL DEFAULT 'active',
                FOREIGN KEY (station_id) REFERENCES stations(id)
            )
            """
        )
        conn.execute(
            """
            INSERT INTO stations (id, name, voltage_level, county)
            VALUES (1, '110kV寿元变', '110kV', '测试县')
            """
        )
        conn.execute(
            """
            INSERT INTO camera_slots (id, slot_code, station_id, project_id, location_desc, area, channel_number)
            VALUES (1, 'A-01', 1, 1, '主变区域', '主变区', 1)
            """
        )
        conn.execute(
            """
            INSERT INTO cameras (id, station_id, camera_index, area, location_desc, ip_address, channel_port, channel_number)
            VALUES (1, 1, 'CAM-01', '主变区', '主变区域', '10.0.0.1', 8000, 1)
            """
        )
        conn.execute(
            """
            INSERT INTO fault_reports (id, station_id, camera_id, fault_type, reporter_name, status)
            VALUES (1, 1, 1, 'OFFLINE', 'tester', 'open')
            """
        )
        conn.execute(
            """
            INSERT INTO photos (id, rel_path, abs_path, filename, ext, station_id, match_status, match_method)
            VALUES (1, 'photos/shouyuan.jpg', 'E:/photos/shouyuan.jpg', 'shouyuan.jpg', '.jpg', 1, 'matched', 'manual')
            """
        )
        conn.execute(
            """
            INSERT INTO station_aliases (id, station_id, alias, source)
            VALUES (1, 1, '寿元变', 'manual')
            """
        )
        conn.execute(
            """
            INSERT INTO station_external_names (id, station_id, source_system, external_name, normalized_name, is_primary)
            VALUES (1, 1, 'inventory', '110kV寿元变', '110kv寿元变', 1)
            """
        )
        conn.execute(
            """
            INSERT INTO station_recorders (id, station_id, project_id, recorder_name, ip_address, port, description, status)
            VALUES (1, 1, 1, 'NVR-01', '10.0.0.10', 8000, '测试录像机', 'active')
            """
        )
        conn.commit()
        conn.close()

        response = client.delete('/admin/stations/1')
        assert response.status_code == 200
        assert '已删除变电站' in response.get_json()['message']

        conn = sqlite3.connect(test_db)
        assert conn.execute("SELECT COUNT(*) FROM stations WHERE id = 1").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM cameras WHERE station_id = 1").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM fault_reports WHERE station_id = 1").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM camera_slots WHERE station_id = 1").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM station_aliases WHERE station_id = 1").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM station_external_names WHERE station_id = 1").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM station_recorders WHERE station_id = 1").fetchone()[0] == 0
        assert conn.execute("SELECT station_id FROM photos WHERE id = 1").fetchone()[0] is None
        conn.close()

class TestTokenAuth:
    """Token认证测试（决策#1, #2）"""

    def test_missing_token(self, client, init_db):
        """缺少token - 401"""
        response = client.get('/api/stations/1/password')
        assert response.status_code == 401
        data = response.get_json()
        assert '未提供认证令牌' in data['error']

    def test_invalid_token_format(self, client, init_db):
        """token格式错误 - 401"""
        response = client.get('/api/stations/1/password',
                              headers={'Authorization': 'InvalidFormat'})
        assert response.status_code == 401

    def test_wrong_token(self, client, init_db):
        """错误的token - 403"""
        response = client.get('/api/stations/1/password',
                              headers={'Authorization': 'Bearer wrong_token'})
        assert response.status_code == 403
        data = response.get_json()
        assert '令牌无效' in data['error']

class TestFaultStatusTransition:
    """故障状态转换测试（决策#7）"""

    @pytest.fixture
    def setup_fault(self, client, test_db):
        """创建测试故障记录"""
        # 确保使用测试数据库（与test_import.py相同的模式）
        import config
        original_path = config.Config.DATABASE_PATH
        config.Config.DATABASE_PATH = test_db

        # 调用init_db确保schema正确
        from init_db import init_db
        init_db()

        # 通过直接连接插入测试数据
        import sqlite3
        conn = sqlite3.connect(test_db)
        conn.execute("""
            INSERT INTO stations (id, name, voltage_level, county)
            VALUES (1, '测试变电站', '220kV', '测试县')
        """)
        conn.execute("""
            INSERT INTO fault_reports (id, station_id, fault_type, reporter_name, status)
            VALUES (1, 1, '无图像', '测试人', 'open')
        """)
        conn.commit()
        conn.close()

        # 恢复原始路径
        config.Config.DATABASE_PATH = original_path
        return 1

    def test_open_to_handling(self, client, init_db, setup_fault):
        """open -> handling 合法"""
        response = client.put('/api/faults/1/status',
                              json={'status': 'handling'})
        assert response.status_code == 200

    def test_open_to_closed(self, client, init_db, setup_fault):
        """open -> closed 合法"""
        response = client.put('/api/faults/1/status',
                              json={'status': 'closed'})
        assert response.status_code == 200

    def test_handling_to_closed_requires_note(self, client, init_db, setup_fault):
        """handling -> closed 需要处理人和备注"""
        # 先转为handling
        client.put('/api/faults/1/status', json={'status': 'handling'})

        # 不带处理信息
        response = client.put('/api/faults/1/status', json={'status': 'closed'})
        assert response.status_code == 400

        # 带处理信息
        response = client.put('/api/faults/1/status', json={
            'status': 'closed',
            'handler_name': '李四',
            'handler_note': '已修复'
        })
        assert response.status_code == 200

    def test_handling_to_closed_can_update_fault_type(self, client, init_db, setup_fault, test_db):
        client.put('/api/faults/1/status', json={'status': 'handling'})

        response = client.put('/api/faults/1/status', json={
            'status': 'closed',
            'handler_name': '李四',
            'handler_note': '现场确认并修复',
            'fault_type': '网络掉线/通信异常',
        })
        assert response.status_code == 200

        conn = sqlite3.connect(test_db)
        try:
            row = conn.execute(
                "SELECT status, fault_type, handler_name, handler_note FROM fault_reports WHERE id = 1"
            ).fetchone()
        finally:
            conn.close()

        assert row == ('closed', '网络掉线/通信异常', '李四', '现场确认并修复')

    def test_invalid_status(self, client, init_db, setup_fault):
        """无效状态值"""
        response = client.put('/api/faults/1/status',
                              json={'status': 'invalid'})
        assert response.status_code == 400

    def test_not_found(self, client, init_db):
        """故障不存在"""
        response = client.put('/api/faults/999/status',
                              json={'status': 'handling'})
        assert response.status_code == 404

    def test_closed_to_handling_rejected(self, client, init_db, setup_fault):
        """closed -> handling 应该被拒绝"""
        # 先将故障关闭
        client.put('/api/faults/1/status', json={
            'status': 'closed',
            'handler_name': '张三',
            'handler_note': '已处理'
        })

        # 尝试从closed转为handling应该失败
        response = client.put('/api/faults/1/status',
                              json={'status': 'handling'})
        assert response.status_code == 400
        data = response.get_json()
        assert '不能从 closed 转换为 handling' in data['error']

    def test_closed_to_open_rejected(self, client, init_db, setup_fault):
        """closed -> open 应该被拒绝"""
        # 先将故障关闭
        client.put('/api/faults/1/status', json={
            'status': 'closed',
            'handler_name': '张三',
            'handler_note': '已处理'
        })

        # 尝试从closed转为open应该失败
        response = client.put('/api/faults/1/status',
                              json={'status': 'open'})
        assert response.status_code == 400
        data = response.get_json()
        assert '不能从 closed 转换为 open' in data['error']


class TestIdempotency:
    """幂等键测试（决策#7）"""

    @pytest.fixture
    def setup_with_camera(self, client, test_db):
        """创建带摄像头的测试数据"""
        # 确保使用测试数据库（与test_import.py相同的模式）
        import config
        original_path = config.Config.DATABASE_PATH
        config.Config.DATABASE_PATH = test_db

        # 调用init_db确保schema正确
        from init_db import init_db
        init_db()

        # 通过直接连接插入测试数据
        import sqlite3
        conn = sqlite3.connect(test_db)
        conn.execute("""
            INSERT INTO stations (id, name, voltage_level, county)
            VALUES (1, '测试变电站', '220kV', '测试县')
        """)
        conn.execute("""
            INSERT INTO cameras (id, station_id, camera_index, ip_address)
            VALUES (1, 1, '1', '192.168.1.100')
        """)
        conn.execute("""
            INSERT INTO cameras (id, station_id, camera_index, ip_address)
            VALUES (2, 1, '2', '192.168.1.101')
        """)
        conn.commit()
        conn.close()

        # 恢复原始路径
        config.Config.DATABASE_PATH = original_path

    def test_duplicate_submission_conflict(self, client, init_db, setup_with_camera):
        """5分钟内重复提交返回409"""
        # 第一次提交
        response = client.post('/api/faults', json={
            'station_id': 1,
            'camera_id': 1,
            'fault_type': '无图像',
            'reporter_name': '张三'
        })
        assert response.status_code == 201

        # 立即重复提交（同一窗口）
        response = client.post('/api/faults', json={
            'station_id': 1,
            'camera_id': 1,
            'fault_type': '无图像',
            'reporter_name': '张三'
        })
        assert response.status_code == 409
        data = response.get_json()
        assert '5分钟内有报修记录' in data['error']

    def test_without_camera_id(self, client, init_db, setup_with_camera):
        """无camera_id时使用IP文本哈希"""
        response = client.post('/api/faults', json={
            'station_id': 1,
            'fault_type': '无图像',
            'reporter_name': '张三',
            'camera_ip_free_text': '192.168.1.200'
        })
        # 没有IP匹配，应该能成功（不会触发幂等）
        assert response.status_code in [201, 400]

    def test_allows_missing_fault_type_and_defaults_to_pending(self, client, init_db, setup_with_camera, test_db):
        response = client.post('/api/faults', json={
            'station_id': 1,
            'camera_id': 1,
            'reporter_name': '张三'
        })

        assert response.status_code == 201
        fault_id = response.get_json()['fault_id']

        conn = sqlite3.connect(test_db)
        try:
            row = conn.execute(
                "SELECT fault_type FROM fault_reports WHERE id = ?",
                (fault_id,)
            ).fetchone()
        finally:
            conn.close()

        assert row == ('待现场确认',)

    def test_idempotency_window_boundary(self, client, init_db, setup_with_camera):
        """5分钟窗口边界：14:59和15:01属于不同窗口，都应创建记录"""
        import time
        import math

        # 计算两个不同5分钟窗口的时间戳
        # 窗口1: floor(t / 300) = W
        # 窗口2: floor((t + 301) / 300) = W + 1
        base_time = int(time.time())
        window1 = math.floor(base_time / 300) * 300  # 窗口1的起始时间（对齐到5分钟）
        # 确保两个时间落在不同的5分钟窗口
        time_in_window1 = window1 + 60   # 窗口1的第60秒 (14:59附近)
        time_in_window2 = window1 + 301  # 下一个窗口的第1秒 (15:01附近)

        # 第一次提交（窗口1）
        response = client.post('/api/faults', json={
            'station_id': 1,
            'camera_id': 1,
            'fault_type': '无图像',
            'reporter_name': '张三',
            'report_time': str(time_in_window1)
        })
        assert response.status_code == 201
        fault_id_1 = response.get_json()['fault_id']

        # 第二次提交（窗口2，相隔301秒）应该成功（新窗口）
        response = client.post('/api/faults', json={
            'station_id': 1,
            'camera_id': 1,
            'fault_type': '无图像',
            'reporter_name': '张三',
            'report_time': str(time_in_window2)
        })
        assert response.status_code == 201
        fault_id_2 = response.get_json()['fault_id']

        # 两次提交应该创建不同的故障记录
        assert fault_id_1 != fault_id_2

    def test_create_fault_ignores_owner_fields_during_creation(self, client, init_db, setup_with_camera, test_db):
        ensure_unified_project(test_db)

        with client.session_transaction() as session:
            session['user_id'] = 1
            session['username'] = 'admin'
            session['role'] = 'admin'
            session['csrf_token'] = 'test-csrf-token'


        payload = {
            'station_id': 1,
            'project': 'unified',
            'camera_ids': [1, 2],
            'fault_type': '无图像',
            'reporter_name': '张三',
            'fault_owner_type': 'switch',
            'is_batch_impact': 1,
            'fault_owner_confirmed_by': 99,
            'fault_owner_confirmed_at': '2026-04-25 12:00:00',
        }

        response = client.post('/api/faults', json=payload)

        assert response.status_code == 201
        payload = response.get_json()

        conn = sqlite3.connect(test_db)
        try:
            rows = conn.execute(
                """
                SELECT fault_owner_type, is_batch_impact, impact_camera_count
                FROM fault_reports
                WHERE id IN (?, ?)
                ORDER BY id
                """,
                tuple(payload['fault_ids'])
            ).fetchall()
        finally:
            conn.close()

        assert rows == [
            (None, None, 2),
            (None, None, 2),
        ]

    def test_close_multi_camera_fault_requires_owner_type(self, client, init_db, test_db):
        login_admin_session(client)
        ensure_unified_project(test_db)


        conn = sqlite3.connect(test_db)
        try:
            conn.execute(
                "INSERT INTO stations (id, name, voltage_level, county) VALUES (1, '测试变电站', '220kV', '测试县')"
            )
            conn.execute(
                """
                INSERT INTO fault_reports (
                    id, station_id, fault_type, reporter_name, status,
                    impact_camera_count, created_at, updated_at
                )
                VALUES (
                    1, 1, '无图像', '张三', 'handling',
                    2, '2026-04-25 10:00:00', '2026-04-25 10:00:00'
                )
                """
            )
            conn.commit()
        finally:
            conn.close()

        response = client.put('/api/faults/1/status', json={
            'status': 'closed',
            'handler_name': '李四',
            'handler_note': '已恢复',
            'is_batch_impact': 1,
        })

        assert response.status_code == 400
        assert '必须填写故障归属' in response.get_json()['error']

    def test_close_multi_camera_fault_writes_owner_confirmation_metadata(self, client, init_db, test_db):
        login_admin_session(client)
        ensure_unified_project(test_db)


        conn = sqlite3.connect(test_db)
        try:
            conn.execute(
                "INSERT INTO stations (id, name, voltage_level, county) VALUES (1, '测试变电站', '220kV', '测试县')"
            )
            conn.execute(
                """
                INSERT INTO fault_reports (
                    id, station_id, fault_type, reporter_name, status,
                    impact_camera_count, created_at, updated_at
                )
                VALUES (
                    1, 1, '无图像', '张三', 'handling',
                    2, '2026-04-25 10:00:00', '2026-04-25 10:00:00'
                )
                """
            )
            conn.commit()
        finally:
            conn.close()

        response = client.put('/api/faults/1/status', json={
            'status': 'closed',
            'handler_name': '李四',
            'handler_note': '已恢复',
            'fault_owner_type': 'switch',
            'root_cause_type': 'switch',
            'is_batch_impact': 1,
        })

        assert response.status_code == 200

        conn = sqlite3.connect(test_db)
        try:
            row = conn.execute(
                """
                SELECT status, fault_owner_type, root_cause_type, is_batch_impact,
                       fault_owner_confirmed_by, fault_owner_confirmed_at
                FROM fault_reports
                WHERE id = 1
                """
            ).fetchone()
        finally:
            conn.close()

        assert row[0] == 'closed'
        assert row[1] == 'switch'
        assert row[2] == 'switch'
        assert row[3] == 1
        assert row[4] == 1
        assert row[5]











