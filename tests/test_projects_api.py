# test_projects_api.py - project scope integration tests
import sqlite3
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

import app as app_module
from app import app
from auth import hash_password
from init_db import init_db


@pytest.fixture
def project_test_db(tmp_path):
    return str(tmp_path / "test_projects.db")


@pytest.fixture
def client(project_test_db):
    import config as config_module

    original_path = config_module.Config.DATABASE_PATH
    app_original_db_path = app.config["DATABASE_PATH"]
    config_module.Config.DATABASE_PATH = project_test_db
    app.config["DATABASE_PATH"] = project_test_db
    app.config["TESTING"] = True

    init_db(force=True)
    with app.test_client() as c:
        _patch_client(c)
        yield c

    config_module.Config.DATABASE_PATH = original_path
    app.config["DATABASE_PATH"] = app_original_db_path


@pytest.fixture
def legacy_upload_client(project_test_db):
    import config as config_module

    original_path = config_module.Config.DATABASE_PATH
    app_original_db_path = app.config["DATABASE_PATH"]
    config_module.Config.DATABASE_PATH = project_test_db
    app.config["DATABASE_PATH"] = project_test_db
    app.config["TESTING"] = True

    init_db(force=True)
    with app.test_client() as c:
        _patch_client(c)
        yield c

    config_module.Config.DATABASE_PATH = original_path
    app.config["DATABASE_PATH"] = app_original_db_path


@pytest.fixture
def seeded_project_schema(project_test_db):
    conn = sqlite3.connect(project_test_db)
    conn.executescript(
        """
        CREATE TABLE projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            short_name TEXT NOT NULL,
            color TEXT NOT NULL DEFAULT '#1a73e8',
            fault_type_version_id INTEGER,
            sort_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE project_fault_type_versions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            version INTEGER NOT NULL,
            description TEXT,
            is_published INTEGER DEFAULT 0,
            published_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(project_id, version)
        );

        CREATE TABLE project_fault_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            version_id INTEGER NOT NULL,
            type_code TEXT NOT NULL,
            type_label TEXT NOT NULL,
            semantic_group TEXT,
            sort_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(version_id, type_code)
        );

        CREATE TABLE user_project_scopes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            can_write INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, project_id)
        );

        CREATE TABLE camera_slots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slot_code TEXT NOT NULL,
            station_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            location_desc TEXT NOT NULL DEFAULT '',
            area TEXT NOT NULL DEFAULT '',
            channel_number INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(station_id, project_id, slot_code)
        );

        ALTER TABLE fault_reports ADD COLUMN project_id INTEGER;
        ALTER TABLE fault_reports ADD COLUMN camera_slot_id INTEGER;
        ALTER TABLE fault_reports ADD COLUMN fault_type_code TEXT;
        ALTER TABLE fault_reports ADD COLUMN fault_type_label_snapshot TEXT;
        ALTER TABLE fault_reports ADD COLUMN fault_type_version_id INTEGER;
        ALTER TABLE fault_reports ADD COLUMN source_type TEXT DEFAULT 'manual';
        ALTER TABLE fault_reports ADD COLUMN source_time_raw TEXT;
        ALTER TABLE fault_reports ADD COLUMN source_timezone TEXT;
        ALTER TABLE fault_reports ADD COLUMN handling_started_at TIMESTAMP;
        ALTER TABLE fault_reports ADD COLUMN assigned_to INTEGER;
        ALTER TABLE fault_reports ADD COLUMN tags_json TEXT;
        ALTER TABLE fault_reports ADD COLUMN project_device_code TEXT;
        ALTER TABLE photos ADD COLUMN project_id INTEGER;
        ALTER TABLE photos ADD COLUMN project_hint TEXT;
        ALTER TABLE cameras ADD COLUMN slot_id INTEGER;
        ALTER TABLE cameras ADD COLUMN project_id INTEGER;
        ALTER TABLE cameras ADD COLUMN project_camera_code TEXT;
        ALTER TABLE cameras ADD COLUMN status TEXT DEFAULT 'active';
        ALTER TABLE cameras ADD COLUMN replaced_by_camera_id INTEGER;
        ALTER TABLE cameras ADD COLUMN retired_at TIMESTAMP;

        DROP TABLE cameras;
        CREATE TABLE cameras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slot_id INTEGER,
            station_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            project_camera_code TEXT,
            camera_index TEXT,
            area TEXT,
            location_desc TEXT,
            ip_address TEXT,
            channel_port INTEGER,
            channel_number INTEGER,
            status TEXT DEFAULT 'active',
            replaced_by_camera_id INTEGER,
            retired_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE import_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            source_type TEXT NOT NULL,
            mode TEXT NOT NULL,
            file_count INTEGER,
            success_count INTEGER,
            fail_count INTEGER,
            report_path TEXT,
            operator_id INTEGER,
            timezone_default_used TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE UNIQUE INDEX idx_cameras_one_active_per_slot
        ON cameras(slot_id)
        WHERE status = 'active';
        """
    )

    conn.execute(
        """
        INSERT INTO projects (id, code, name, short_name, color, sort_order, is_active)
        VALUES
            (1, 'unified', 'Unified', 'UNI', '#1a73e8', 1, 1),
            (2, 'inspection', 'Inspection', 'INSP', '#34a853', 2, 1),
            (3, 'auxiliary', 'Auxiliary', 'AUX', '#ea4335', 3, 0)
        """
    )
    conn.execute(
        """
        INSERT INTO project_fault_type_versions (id, project_id, version, description, is_published, published_at)
        VALUES (10, 2, 3, 'inspection current', 1, '2026-04-02T00:00:00')
        """
    )
    conn.execute("UPDATE projects SET fault_type_version_id = 10 WHERE id = 2")
    conn.execute(
        """
        INSERT INTO project_fault_types (version_id, type_code, type_label, semantic_group, sort_order, is_active)
        VALUES
            (10, 'NO_IMAGE', 'No Image', 'NO_IMAGE', 1, 1),
            (10, 'BLUR', 'Blur', 'BLUR', 2, 1)
        """
    )

    password_hash_admin = hash_password("adminpass")
    password_hash_operator = hash_password("operatorpass")
    conn.execute(
        "INSERT INTO users (id, username, password_hash, role) VALUES (1, 'admin1', ?, 'admin')",
        (password_hash_admin,),
    )
    conn.execute(
        "INSERT INTO users (id, username, password_hash, role) VALUES (2, 'operator1', ?, 'operator')",
        (password_hash_operator,),
    )
    conn.execute(
        "INSERT INTO user_project_scopes (user_id, project_id, can_write) VALUES (2, 2, 1)"
    )

    conn.execute(
        "INSERT INTO stations (id, name, voltage_level, county) VALUES (1, 'Station A', '110kV', 'County A')"
    )
    conn.execute(
        "INSERT INTO stations (id, name, voltage_level, county) VALUES (2, 'Station B', '220kV', 'County B')"
    )
    conn.execute(
        """
        INSERT INTO camera_slots (id, slot_code, station_id, project_id, location_desc, area, channel_number)
        VALUES
            (1, 'INSPECT_1', 1, 2, 'inspection-slot', '', 1),
            (2, 'UNIFIED_1', 2, 1, 'unified-slot', '', 1)
        """
    )
    conn.execute(
        """
        INSERT INTO cameras
            (id, station_id, camera_index, area, location_desc, ip_address, channel_number, slot_id, project_id, status)
        VALUES
            (1, 1, '1', '', 'inspection-slot', '10.0.0.2', 1, 1, 2, 'active'),
            (2, 2, '1', '', 'unified-slot', '10.0.0.1', 1, 2, 1, 'active')
        """
    )
    conn.execute(
        """
        INSERT INTO fault_reports
            (id, station_id, camera_id, fault_type, reporter_name, status, project_id, camera_slot_id, fault_type_label_snapshot, tags_json)
        VALUES
            (1, 2, 2, 'No Image', 'Alice', 'open', 1, 2, 'No Image', '["台风", "统一平台"]'),
            (2, 1, 1, 'Blur', 'Bob', 'open', 2, 1, 'Blur', '["台风", "巡视重点"]')
        """
    )
    conn.execute("UPDATE fault_reports SET source_type = 'manual' WHERE id = 1")
    conn.execute(
        """
        UPDATE fault_reports
        SET source_type = 'import_excel',
            project_device_code = 'INSPECT-CAM-001',
            description = 'Lens blur on east yard camera',
            handler_note = 'Cleaned housing and restored focus',
            source_time_raw = '2026-04-02 08:00:00',
            source_timezone = 'Asia/Shanghai'
        WHERE id = 2
        """
    )
    conn.commit()
    conn.close()
    yield


@pytest.fixture
def seeded_legacy_upload_schema(project_test_db):
    conn = sqlite3.connect(project_test_db)
    conn.executescript(
        """
        CREATE TABLE projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            short_name TEXT NOT NULL,
            color TEXT NOT NULL DEFAULT '#1a73e8',
            sort_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE import_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            source_type TEXT NOT NULL,
            mode TEXT NOT NULL,
            file_count INTEGER,
            success_count INTEGER,
            fail_count INTEGER,
            report_path TEXT,
            operator_id INTEGER,
            timezone_default_used TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    conn.execute(
        """
        INSERT INTO projects (id, code, name, short_name, color, sort_order, is_active)
        VALUES (1, 'unified', 'Unified', 'UNI', '#1a73e8', 1, 1)
        """
    )
    password_hash_admin = hash_password("adminpass")
    conn.execute(
        "INSERT INTO users (id, username, password_hash, role) VALUES (1, 'admin1', ?, 'admin')",
        (password_hash_admin,),
    )
    conn.commit()
    conn.close()
    yield


def login(client, username, password):
    response = client.post("/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    return response.get_json()


def _with_csrf(client, kwargs):
    """自动为状态变更请求注入 CSRF token。"""
    headers = dict(kwargs.pop("headers", {}) or {})
    with client.session_transaction() as sess:
        token = sess.get("csrf_token")
    if token:
        headers.setdefault("X-CSRF-Token", token)
    kwargs["headers"] = headers
    return kwargs


def _patch_client(client):
    """给 test client 的状态变更方法自动注入 CSRF token。"""
    _orig_post = client.post
    _orig_put = client.put
    _orig_delete = client.delete

    def _post(url, **kwargs):
        return _orig_post(url, **_with_csrf(client, kwargs))

    def _put(url, **kwargs):
        return _orig_put(url, **_with_csrf(client, kwargs))

    def _delete(url, **kwargs):
        return _orig_delete(url, **_with_csrf(client, kwargs))

    client.post = _post
    client.put = _put
    client.delete = _delete
    return client


def test_api_projects_anonymous_returns_active_projects(client, seeded_project_schema):
    """匿名用户访问 /api/projects 现在需要登录（安全收口后不再公开）。"""
    response = client.get("/api/projects")
    assert response.status_code in (401, 302)


def test_api_projects_anonymous_returns_401_json(client, seeded_project_schema):
    response = client.get("/api/projects")
    assert response.status_code == 401
    data = response.get_json()
    assert data["error"] == "请先登录"


def test_login_response_contains_visible_projects(client, seeded_project_schema):

    data = login(client, "operator1", "operatorpass")
    assert data["user"]["default_project_code"] == "inspection"
    assert [p["code"] for p in data["user"]["projects"]] == ["inspection"]
    assert data["user"]["projects"][0]["can_write"] is True


def test_me_returns_project_scope(client, seeded_project_schema):
    login(client, "operator1", "operatorpass")
    response = client.get("/auth/me")
    assert response.status_code == 200
    data = response.get_json()
    assert data["user"]["default_project_code"] == "inspection"
    assert len(data["user"]["projects"]) == 1
    assert data["user"]["projects"][0]["code"] == "inspection"


def test_admin_can_get_and_update_user_project_scopes(client, seeded_project_schema):
    login(client, "admin1", "adminpass")

    response = client.get("/auth/users/2/projects")
    assert response.status_code == 200
    data = response.get_json()
    assert data["user"]["username"] == "operator1"
    assigned_codes = [p["code"] for p in data["projects"] if p["assigned"]]
    assert assigned_codes == ["inspection"]

    response = client.put(
        "/auth/users/2/projects",
        json={
            "projects": [
                {"project_code": "unified", "can_write": False},
                {"project_code": "inspection", "can_write": True},
            ]
        },
    )
    assert response.status_code == 200
    updated = response.get_json()
    assert updated["message"]
    assert len(updated["projects"]) == 2

    response = client.get("/auth/users/2/projects")
    data = response.get_json()
    scope_map = {p["code"]: p for p in data["projects"]}
    assert scope_map["unified"]["assigned"] is True
    assert scope_map["unified"]["can_write"] is False
    assert scope_map["inspection"]["assigned"] is True
    assert scope_map["inspection"]["can_write"] is True


def test_user_access_center_page_renders_for_admin(client, seeded_project_schema):
    login(client, "admin1", "adminpass")

    response = client.get("/admin/user-access-center")

    assert response.status_code == 200
    assert "用户项目授权".encode("utf-8") in response.data


def test_project_fault_types_endpoint_returns_current_published_types(client, seeded_project_schema):
    login(client, "operator1", "operatorpass")
    response = client.get("/api/projects/inspection/fault-types")
    assert response.status_code == 200
    data = response.get_json()

    assert data["project"]["code"] == "inspection"
    assert data["project"]["current_fault_type_version"]["version"] == 3
    assert [item["type_code"] for item in data["fault_types"]] == ["NO_IMAGE", "BLUR"]


def test_create_fault_uses_project_fault_type_version_and_camera_slot(client, seeded_project_schema, project_test_db):
    login(client, "operator1", "operatorpass")

    response = client.post(
        "/api/faults",
        json={
            "station_id": 1,
            "camera_id": 1,
            "project": "inspection",
            "fault_type": "Blur",
            "fault_type_code": "BLUR",
            "description": "camera blur detected",
            "reporter_name": "Operator One",
            "reporter_contact": "10086",
        },
    )

    assert response.status_code == 201
    fault_id = response.get_json()["fault_id"]

    conn = sqlite3.connect(project_test_db)
    try:
        fault = conn.execute(
            """
            SELECT project_id, camera_slot_id, fault_type_code, fault_type_label_snapshot,
                   fault_type_version_id, project_device_code
            FROM fault_reports
            WHERE id = ?
            """,
            (fault_id,),
        ).fetchone()
    finally:
        conn.close()

    assert fault == (2, 1, "BLUR", "Blur", 10, "1")


def test_create_fault_allows_missing_fault_type_and_defaults_to_pending(client, seeded_project_schema, project_test_db):
    login(client, "operator1", "operatorpass")

    response = client.post(
        "/api/faults",
        json={
            "station_id": 1,
            "camera_id": 1,
            "project": "inspection",
            "description": "type unknown before onsite inspection",
            "reporter_name": "Operator One",
            "reporter_contact": "10086",
        },
    )

    assert response.status_code == 201
    fault_id = response.get_json()["fault_id"]

    conn = sqlite3.connect(project_test_db)
    try:
        fault = conn.execute(
            """
            SELECT fault_type, fault_type_code, fault_type_label_snapshot, fault_type_version_id
            FROM fault_reports
            WHERE id = ?
            """,
            (fault_id,),
        ).fetchone()
    finally:
        conn.close()

    assert fault == ("待现场确认", None, "待现场确认", None)


def test_create_fault_with_multiple_cameras_creates_project_scoped_group(client, seeded_project_schema, project_test_db):
    conn = sqlite3.connect(project_test_db)
    conn.execute(
        """
        INSERT INTO camera_slots (id, slot_code, station_id, project_id, location_desc, area, channel_number)
        VALUES (4, 'INSPECT_2', 1, 2, 'inspection-slot-2', '', 2)
        """
    )
    conn.execute(
        """
        INSERT INTO cameras
            (id, station_id, camera_index, area, location_desc, ip_address, channel_number, slot_id, project_id, status)
        VALUES
            (4, 1, '2', '', 'inspection-slot-2', '10.0.0.4', 2, 4, 2, 'active')
        """
    )
    conn.commit()
    conn.close()

    login(client, "operator1", "operatorpass")

    response = client.post(
        "/api/faults",
        json={
            "station_id": 1,
            "camera_ids": [1, 4],
            "project": "inspection",
            "fault_type": "Blur",
            "fault_type_code": "BLUR",
            "description": "multi camera blur detected",
            "reporter_name": "Operator One",
            "reporter_contact": "10086",
        },
    )

    assert response.status_code == 201
    payload = response.get_json()
    assert payload["fault_count"] == 2
    assert len(payload["fault_ids"]) == 2
    assert payload["fault_group_key"]

    conn = sqlite3.connect(project_test_db)
    try:
        rows = conn.execute(
            """
            SELECT camera_id, project_id, camera_slot_id, fault_type_code, fault_type_version_id, fault_group_key
            FROM fault_reports
            WHERE id IN (?, ?)
            ORDER BY camera_id
            """,
            tuple(payload["fault_ids"]),
        ).fetchall()
    finally:
        conn.close()

    assert rows == [
        (1, 2, 1, "BLUR", 10, payload["fault_group_key"]),
        (4, 2, 4, "BLUR", 10, payload["fault_group_key"]),
    ]


def test_create_fault_rejects_camera_project_mismatch(client, seeded_project_schema):
    login(client, "admin1", "adminpass")

    response = client.post(
        "/api/faults",
        json={
            "station_id": 1,
            "camera_id": 1,
            "project": "unified",
            "fault_type": "No Image",
            "fault_type_code": "NO_IMAGE",
            "reporter_name": "Admin One",
        },
    )

    assert response.status_code == 400
    assert "摄像头与项目不匹配" in response.get_json()["error"]


def test_create_fault_requires_project_write_scope(client, seeded_project_schema, project_test_db):
    conn = sqlite3.connect(project_test_db)
    conn.execute(
        "INSERT INTO user_project_scopes (user_id, project_id, can_write) VALUES (2, 1, 0)"
    )
    conn.commit()
    conn.close()

    login(client, "operator1", "operatorpass")

    response = client.post(
        "/api/faults",
        json={
            "station_id": 2,
            "camera_id": 2,
            "project": "unified",
            "fault_type": "No Image",
            "fault_type_code": "NO_IMAGE",
            "reporter_name": "Operator One",
        },
    )

    assert response.status_code == 403
    denied = response.get_json()
    assert denied["code"] == "PROJECT_ACCESS_DENIED"


def test_create_fault_rejects_inactive_project_inferred_from_camera(client, seeded_project_schema, project_test_db):
    conn = sqlite3.connect(project_test_db)
    conn.execute(
        """
        INSERT INTO camera_slots (id, slot_code, station_id, project_id, location_desc, area, channel_number)
        VALUES (3, 'AUX_1', 1, 3, 'aux-slot', '', 3)
        """
    )
    conn.execute(
        """
        INSERT INTO cameras
            (id, station_id, camera_index, area, location_desc, ip_address, channel_number, slot_id, project_id, status)
        VALUES
            (3, 1, '3', '', 'aux-slot', '10.0.0.3', 3, 3, 3, 'active')
        """
    )
    conn.commit()
    conn.close()

    login(client, "admin1", "adminpass")

    response = client.post(
        "/api/faults",
        json={
            "station_id": 1,
            "camera_id": 3,
            "fault_type": "Blur",
            "reporter_name": "Admin One",
        },
    )

    assert response.status_code == 404
    assert "椤圭洰" in response.get_json()["error"]


def test_faults_project_filter_enforces_scope(client, seeded_project_schema):
    login(client, "operator1", "operatorpass")

    response = client.get("/api/faults?project=inspection")
    assert response.status_code == 200
    data = response.get_json()
    assert data["total"] == 1
    assert data["faults"][0]["project_id"] == 2

    response = client.get("/api/faults?project=unified")
    assert response.status_code == 403
    denied = response.get_json()
    assert denied["code"] == "PROJECT_ACCESS_DENIED"


def test_faults_default_scope_only_returns_visible_projects(client, seeded_project_schema):
    login(client, "operator1", "operatorpass")

    response = client.get("/api/faults")
    assert response.status_code == 200
    data = response.get_json()

    assert data["total"] == 1
    assert {fault["project_id"] for fault in data["faults"]} == {2}


def test_faults_source_type_filter_returns_matching_records(client, seeded_project_schema):
    login(client, "admin1", "adminpass")

    response = client.get("/api/faults?source_type=import_excel")
    assert response.status_code == 200
    data = response.get_json()

    assert data["total"] == 1
    assert data["faults"][0]["id"] == 2
    assert data["faults"][0]["source_type"] == "import_excel"


def test_stations_and_cameras_default_scope_follow_visible_projects(client, seeded_project_schema):
    login(client, "operator1", "operatorpass")

    stations_response = client.get("/api/stations")
    assert stations_response.status_code == 200
    stations_data = stations_response.get_json()
    assert stations_data["total"] == 1
    assert stations_data["stations"][0]["id"] == 1

    station_detail = client.get("/api/stations/1")
    assert station_detail.status_code == 200
    station_payload = station_detail.get_json()
    assert station_payload["station"]["camera_count"] == 1
    assert station_payload["station"]["fault_count"] == 1
    assert len(station_payload["cameras"]) == 1
    assert station_payload["cameras"][0]["project_id"] == 2

    hidden_station = client.get("/api/stations/2")
    assert hidden_station.status_code == 404

    cameras_response = client.get("/api/cameras")
    assert cameras_response.status_code == 200
    cameras_data = cameras_response.get_json()
    assert cameras_data["total"] == 1
    assert cameras_data["cameras"][0]["project_id"] == 2

    by_ip_allowed = client.get("/api/cameras/by-ip?ip=10.0.0.2")
    assert by_ip_allowed.status_code == 200
    assert by_ip_allowed.get_json()["camera"]["project_id"] == 2

    by_ip_hidden = client.get("/api/cameras/by-ip?ip=10.0.0.1")
    assert by_ip_hidden.status_code == 404


def test_station_slots_endpoint_returns_project_scoped_slot_view(client, seeded_project_schema):
    conn = sqlite3.connect(app.config["DATABASE_PATH"])
    conn.execute(
        """
        CREATE TABLE station_recorders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            station_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            recorder_name TEXT,
            ip_address TEXT,
            port INTEGER,
            description TEXT,
            source_type TEXT DEFAULT 'manual',
            source_key TEXT,
            status TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        INSERT INTO station_recorders
            (id, station_id, project_id, recorder_name, ip_address, port, description, status)
        VALUES
            (1, 1, 2, 'Inspection Recorder 1', '10.0.0.2', 8000, 'Main inspection recorder', 'active')
        """
    )
    conn.commit()
    conn.close()

    login(client, "operator1", "operatorpass")

    response = client.get("/api/stations/1/slots?project=inspection")
    assert response.status_code == 200
    data = response.get_json()

    assert data["total"] == 1
    assert len(data["recorders"]) == 1
    assert data["recorders"][0]["recorder_name"] == "Inspection Recorder 1"
    slot = data["slots"][0]
    assert slot["project_code"] == "inspection"
    assert slot["slot_id"] == 1
    assert slot["fault_count"] == 1
    assert slot["current_camera"]["id"] == 1
    assert slot["history_camera_count"] == 0
    assert slot["recorder"]["recorder_name"] == "Inspection Recorder 1"
    assert slot["recent_faults"][0]["fault_label"] == "Blur"
    assert slot["recent_faults"][0]["description"] == "Lens blur on east yard camera"
    assert slot["recent_faults"][0]["handler_note"] == "Cleaned housing and restored focus"


def test_station_slots_endpoint_dedupes_semantic_duplicate_slots(client, seeded_project_schema):
    conn = sqlite3.connect(app.config["DATABASE_PATH"])
    conn.execute(
        """
        INSERT INTO camera_slots (id, slot_code, station_id, project_id, location_desc, area, channel_number)
        VALUES
            (3, 'MIGRATED_119', 1, 2, '2#主变北侧', '', 9),
            (4, 'LEGACY_unified_1_9_2#主变北侧-9#球机_dup', 1, 2, '2#主变北侧-9#球机', '', 9),
            (5, 'LEGACY_unified_1_9_枪机_2#主变北侧-9#球机_dup', 1, 2, '2#主变北侧-9#球机', '枪机', 9)
        """
    )
    conn.execute(
        """
        INSERT INTO cameras
            (id, station_id, camera_index, project_camera_code, area, location_desc, ip_address, channel_number, slot_id, project_id, status)
        VALUES
            (3, 1, '9', '', '', '2#主变北侧', '10.0.0.9', 9, 3, 2, 'retired'),
            (4, 1, '9', '2#主变北侧-9#球机', '', '2#主变北侧-9#球机', '10.0.0.253', 9, 4, 2, 'retired'),
            (5, 1, '9', '', '枪机', '2#主变北侧-9#球机', '', 9, 5, 2, 'active')
        """
    )
    conn.commit()
    conn.close()

    login(client, "operator1", "operatorpass")

    response = client.get("/api/stations/1/slots?project=inspection")
    assert response.status_code == 200
    data = response.get_json()

    assert data["total"] == 2
    duplicate_slot = next(
        slot for slot in data["slots"] if slot["channel_number"] == 9
    )
    assert duplicate_slot["location_desc"] == "2#主变北侧"
    assert duplicate_slot["current_camera"]["id"] == 5
    assert duplicate_slot["history_camera_count"] == 1
    assert [camera["id"] for camera in duplicate_slot["history_cameras"]] == [3]


def test_station_slots_endpoint_drops_shadow_history_from_merged_slot(client, seeded_project_schema):
    conn = sqlite3.connect(app.config["DATABASE_PATH"])
    conn.execute(
        """
        INSERT INTO camera_slots (id, slot_code, station_id, project_id, location_desc, area, channel_number)
        VALUES
            (3, 'MIGRATED_113', 1, 2, '110kV场地西侧-3#球机', '', 3),
            (4, 'LEGACY_unified_1_3_枪机_110kV场地西侧-3#球机_dup', 1, 2, '110kV场地西侧-3#球机', '枪机', 3)
        """
    )
    conn.execute(
        """
        INSERT INTO cameras
            (id, station_id, camera_index, project_camera_code, area, location_desc, ip_address, channel_number, slot_id, project_id, status, replaced_by_camera_id)
        VALUES
            (3, 1, '3', NULL, '', '110kV场地西侧-3#球机', '10.0.0.131', 3, 3, 2, 'replaced', 4),
            (4, 1, '3', '110kV场地西侧-3#球机', '', '110kV场地西侧-3#球机', '10.0.0.253', 3, 3, 2, 'retired', NULL),
            (5, 1, '3', '', '枪机', '110kV场地西侧-3#球机', '', 3, 4, 2, 'active', NULL)
        """
    )
    conn.commit()
    conn.close()

    login(client, "operator1", "operatorpass")

    response = client.get("/api/stations/1/slots?project=inspection")
    assert response.status_code == 200
    data = response.get_json()

    duplicate_slot = next(
        slot for slot in data["slots"] if slot["channel_number"] == 3
    )
    assert duplicate_slot["current_camera"]["id"] == 5
    assert duplicate_slot["history_camera_count"] == 1
    assert [camera["id"] for camera in duplicate_slot["history_cameras"]] == [3]


def test_station_slots_endpoint_shows_history_after_replacement(client, seeded_project_schema):
    login(client, "admin1", "adminpass")
    replace_response = client.post(
        "/admin/cameras/1/replace",
        json={
            "project": "inspection",
            "ip_address": "10.0.0.88",
            "project_camera_code": "INS-NEW-001",
        },
    )
    assert replace_response.status_code == 201
    new_camera_id = replace_response.get_json()["new_camera_id"]

    response = client.get("/api/stations/1/slots?project=inspection")
    assert response.status_code == 200
    slot = response.get_json()["slots"][0]

    assert slot["current_camera"]["id"] == new_camera_id
    assert slot["current_camera"]["ip_address"] == "10.0.0.88"
    assert slot["history_camera_count"] == 1
    assert slot["history_cameras"][0]["id"] == 1
    assert slot["history_cameras"][0]["status"] == "replaced"
    assert slot["history_cameras"][0]["replaced_by_camera_id"] == new_camera_id


def test_fault_status_update_requires_project_write_scope(client, seeded_project_schema):
    login(client, "operator1", "operatorpass")

    allowed = client.put("/api/faults/2/status", json={"status": "handling"})
    assert allowed.status_code == 200

    conn = sqlite3.connect(app.config["DATABASE_PATH"])
    try:
        row = conn.execute(
            "SELECT status, assigned_to, handling_started_at FROM fault_reports WHERE id = 2"
        ).fetchone()
    finally:
        conn.close()

    assert row[0] == "handling"
    assert row[1] == 2
    assert row[2] is not None

    denied = client.put("/api/faults/1/status", json={"status": "handling"})
    assert denied.status_code == 403
    assert denied.get_json()["code"] == "PROJECT_ACCESS_DENIED"


def test_fault_status_close_updates_catalog_fault_type(client, seeded_project_schema, project_test_db):
    login(client, "operator1", "operatorpass")

    handling = client.put("/api/faults/2/status", json={"status": "handling"})
    assert handling.status_code == 200

    closed = client.put(
        "/api/faults/2/status",
        json={
            "status": "closed",
            "handler_name": "Operator One",
            "handler_note": "onsite diagnosis confirmed blur fault",
            "fault_type": "待现场确认",
            "fault_type_code": "BLUR",
        },
    )
    assert closed.status_code == 200

    conn = sqlite3.connect(project_test_db)
    try:
        row = conn.execute(
            """
            SELECT status, fault_type, fault_type_code, fault_type_label_snapshot, fault_type_version_id
            FROM fault_reports
            WHERE id = 2
            """
        ).fetchone()
    finally:
        conn.close()

    assert row == ("closed", "Blur", "BLUR", "Blur", 10)


def test_fault_status_close_accepts_multiple_catalog_fault_types(client, seeded_project_schema, project_test_db):
    login(client, "operator1", "operatorpass")

    handling = client.put("/api/faults/2/status", json={"status": "handling"})
    assert handling.status_code == 200

    closed = client.put(
        "/api/faults/2/status",
        json={
            "status": "closed",
            "handler_name": "Operator One",
            "handler_note": "onsite diagnosis found multiple issues",
            "fault_type": "待现场确认",
            "fault_type_code": "BLUR,NO_IMAGE",
        },
    )
    assert closed.status_code == 200

    conn = sqlite3.connect(project_test_db)
    try:
        row = conn.execute(
            """
            SELECT status, fault_type, fault_type_code, fault_type_label_snapshot, fault_type_version_id
            FROM fault_reports
            WHERE id = 2
            """
        ).fetchone()
    finally:
        conn.close()

    assert row == ("closed", "Blur | No Image", "BLUR,NO_IMAGE", "Blur | No Image", 10)


def test_fault_status_close_accepts_catalog_labels_in_fault_type_code(client, seeded_project_schema, project_test_db):
    login(client, "operator1", "operatorpass")

    handling = client.put("/api/faults/2/status", json={"status": "handling"})
    assert handling.status_code == 200

    closed = client.put(
        "/api/faults/2/status",
        json={
            "status": "closed",
            "handler_name": "Operator One",
            "handler_note": "frontend submitted labels instead of codes",
            "fault_type": "待现场确认",
            "fault_type_code": "Blur,No Image",
        },
    )
    assert closed.status_code == 200

    conn = sqlite3.connect(project_test_db)
    try:
        row = conn.execute(
            """
            SELECT status, fault_type, fault_type_code, fault_type_label_snapshot, fault_type_version_id
            FROM fault_reports
            WHERE id = 2
            """
        ).fetchone()
    finally:
        conn.close()

    assert row == ("closed", "Blur | No Image", "BLUR,NO_IMAGE", "Blur | No Image", 10)


def test_fault_status_close_autoresolves_codes_from_catalog_labels(client, seeded_project_schema, project_test_db):
    login(client, "operator1", "operatorpass")

    handling = client.put("/api/faults/2/status", json={"status": "handling"})
    assert handling.status_code == 200

    closed = client.put(
        "/api/faults/2/status",
        json={
            "status": "closed",
            "handler_name": "Operator One",
            "handler_note": "labels should map back to published codes",
            "fault_type": "Blur | No Image",
            "fault_type_code": None,
        },
    )
    assert closed.status_code == 200

    conn = sqlite3.connect(project_test_db)
    try:
        row = conn.execute(
            """
            SELECT status, fault_type, fault_type_code, fault_type_label_snapshot, fault_type_version_id
            FROM fault_reports
            WHERE id = 2
            """
        ).fetchone()
    finally:
        conn.close()

    assert row == ("closed", "Blur | No Image", "BLUR,NO_IMAGE", "Blur | No Image", 10)


def test_stats_counts_camera_replacements_from_closed_faults(client, seeded_project_schema, project_test_db):
    login(client, "operator1", "operatorpass")

    conn = sqlite3.connect(project_test_db)
    try:
        conn.execute(
            """
            INSERT INTO fault_reports (
                id, station_id, camera_id, fault_type, reporter_name, status, project_id, camera_slot_id,
                equipment_type, equipment_quantity, created_at, closed_at, updated_at
            )
            VALUES
                (20, 1, 1, 'Blur', 'Bob', 'closed', 2, 1, '摄像机', 2, '2026-04-02 08:00:00', '2026-04-03 09:00:00', '2026-04-03 09:00:00'),
                (21, 1, 1, 'Blur', 'Bob', 'closed', 2, 1, '球机', 0, '2026-04-06 08:00:00', '2026-04-06 10:00:00', '2026-04-06 10:00:00'),
                (22, 2, 2, 'No Image', 'Alice', 'closed', 1, 2, '摄像机', 9, '2026-04-04 08:00:00', '2026-04-04 10:00:00', '2026-04-04 10:00:00')
            """
        )
        conn.commit()
    finally:
        conn.close()

    response = client.get("/api/stats?project=inspection&year=2026")
    assert response.status_code == 200
    payload = response.get_json()
    kpi = payload["kpi"]

    assert kpi["camera_replacement_count"] == 3
    assert kpi["camera_replacement_record_count"] == 2
    assert kpi["camera_replacement_inferred_record_count"] == 1


def test_stats_payload_includes_distribution_and_coverage_metrics(client, seeded_project_schema, project_test_db):
    login(client, "operator1", "operatorpass")

    conn = sqlite3.connect(project_test_db)
    try:
        conn.execute(
            """
            INSERT INTO camera_slots (id, slot_code, station_id, project_id, location_desc, area, channel_number)
            VALUES (3, 'INSPECT_2', 2, 2, 'inspection-slot-b', '', 2)
            """
        )
        conn.execute(
            """
            INSERT INTO cameras
                (id, station_id, camera_index, area, location_desc, ip_address, channel_number, slot_id, project_id, status)
            VALUES
                (3, 2, '2', '', 'inspection-slot-b', '10.0.0.3', 2, 3, 2, 'active')
            """
        )
        conn.execute(
            """
            INSERT INTO fault_reports (
                id, station_id, camera_id, fault_type, reporter_name, status, project_id, camera_slot_id,
                fault_type_label_snapshot, source_type, created_at, handling_started_at, closed_at, updated_at
            )
            VALUES
                (30, 1, 1, 'Blur', 'Bob', 'closed', 2, 1, 'Blur', 'manual', '2026-04-02 08:00:00', '2026-04-02 09:30:00', '2026-04-03 10:00:00', '2026-04-03 10:00:00'),
                (31, 2, 3, 'No Image', 'Carol', 'handling', 2, 3, 'No Image', 'manual', '2026-04-01 08:00:00', '2026-04-02 10:00:00', NULL, '2026-04-02 10:00:00'),
                (32, 2, 3, 'Blur', 'Carol', 'closed', 2, 3, 'Blur', 'manual', '2026-04-05 08:00:00', '2026-04-05 09:00:00', '2026-04-10 09:00:00', '2026-04-10 09:00:00')
            """
        )
        conn.execute(
            """
            INSERT INTO photos
                (rel_path, abs_path, filename, ext, station_id, match_status, match_method, project_id, project_hint)
            VALUES
                ('coverage-station-a.jpg', 'coverage-station-a.jpg', 'coverage-station-a.jpg', '.jpg', 1, 'matched', 'manual', 2, 'inspection')
            """
        )
        conn.commit()
    finally:
        conn.close()

    response = client.get("/api/stats?project=inspection")
    assert response.status_code == 200
    payload = response.get_json()
    kpi = payload["kpi"]

    assert kpi["overdue_threshold_days"] == 7
    assert kpi["overdue_unresolved_count"] == 1
    assert kpi["overdue_unresolved_ratio"] == 50.0

    assert payload["response_buckets"] == [
        {"label": "2小时内", "count": 2},
        {"label": "2-8小时", "count": 0},
        {"label": "8-24小时", "count": 0},
        {"label": "24小时以上", "count": 1},
    ]
    assert payload["close_buckets"] == [
        {"label": "当天闭环", "count": 0},
        {"label": "1-3天", "count": 1},
        {"label": "3-7天", "count": 1},
        {"label": "7天以上", "count": 0},
    ]

    assert payload["station_ranking"][0] == {
        "station_id": 1,
        "station_name": "Station A",
        "county": "County A",
        "fault_count": 2,
        "unresolved_count": 1,
    }
    assert payload["photo_coverage"] == {
        "fault_station_count": 2,
        "covered_station_count": 1,
        "uncovered_station_count": 1,
        "coverage_ratio": 50.0,
        "uncovered_stations": [
            {
                "station_id": 2,
                "station_name": "Station B",
                "county": "County B",
                "fault_count": 2,
                "unresolved_count": 1,
            }
        ],
    }


def test_expand_fault_type_distribution_splits_multi_fault_type_rows():
    rows = [
        {"semantic_group": "BLUR,NO_IMAGE", "fault_label": "Blur | No Image", "count": 2},
        {"semantic_group": "BLUR", "fault_label": "Blur", "count": 1},
    ]

    distribution = app_module.expand_fault_type_distribution(rows)

    assert distribution == [
        {"semantic_group": "BLUR", "fault_label": "Blur", "count": 3},
        {"semantic_group": "NO_IMAGE", "fault_label": "No Image", "count": 2},
    ]


def test_fault_tag_suggestions_follow_project_scope(client, seeded_project_schema):
    login(client, "operator1", "operatorpass")

    response = client.get("/api/fault-tags?project=inspection")
    assert response.status_code == 200
    data = response.get_json()
    assert data["tags"] == ["台风", "巡视重点"]


def test_fault_tag_update_requires_project_write_scope(client, seeded_project_schema, project_test_db):
    login(client, "operator1", "operatorpass")

    allowed = client.put("/api/faults/2/tags", json={"tags": ["台风", "复核"]})
    assert allowed.status_code == 200
    assert allowed.get_json()["tags"] == ["台风", "复核"]

    denied = client.put("/api/faults/1/tags", json={"tags": ["不应成功"]})
    assert denied.status_code == 403
    assert denied.get_json()["code"] == "PROJECT_ACCESS_DENIED"

    conn = sqlite3.connect(project_test_db)
    try:
        inspection_tags = conn.execute(
            "SELECT tags_json FROM fault_reports WHERE id = 2"
        ).fetchone()[0]
        unified_tags = conn.execute(
            "SELECT tags_json FROM fault_reports WHERE id = 1"
        ).fetchone()[0]
    finally:
        conn.close()

    assert inspection_tags == '["台风", "复核"]'
    assert unified_tags == '["台风", "统一平台"]'


def test_fault_detail_update_requires_project_write_scope(client, seeded_project_schema, project_test_db):
    login(client, "operator1", "operatorpass")

    allowed = client.put(
        "/api/faults/2",
        json={
            "fault_type": "Blur Updated",
            "description": "updated description",
            "camera_location_text": "inspection-slot-updated",
            "handler_note": "updated note",
        },
    )
    assert allowed.status_code == 200

    denied = client.put(
        "/api/faults/1",
        json={
            "fault_type": "No Image Updated",
            "description": "should fail",
        },
    )
    assert denied.status_code == 403
    assert denied.get_json()["code"] == "PROJECT_ACCESS_DENIED"

    conn = sqlite3.connect(project_test_db)
    try:
        inspection_fault = conn.execute(
            """
            SELECT fault_type, fault_type_label_snapshot, description, camera_location_text, handler_note
            FROM fault_reports
            WHERE id = 2
            """
        ).fetchone()
        unified_fault = conn.execute(
            """
            SELECT fault_type, description
            FROM fault_reports
            WHERE id = 1
            """
        ).fetchone()
    finally:
        conn.close()

    assert inspection_fault == (
        "Blur Updated",
        "Blur Updated",
        "updated description",
        "inspection-slot-updated",
        "updated note",
    )
    assert unified_fault == ("No Image", None)


def test_admin_add_camera_requires_project_in_multi_project_mode(client, seeded_project_schema):
    login(client, "admin1", "adminpass")

    response = client.post(
        "/admin/cameras",
        json={
            "station_id": 1,
            "ip_address": "10.0.0.99",
            "location": "new-slot",
            "channel_number": 9,
        },
    )

    assert response.status_code == 400
    assert "project" in response.get_json()["error"]


def test_admin_add_camera_creates_project_scoped_slot_and_camera(client, seeded_project_schema, project_test_db):
    login(client, "admin1", "adminpass")

    response = client.post(
        "/admin/cameras",
        json={
            "station_id": 1,
            "project": "inspection",
            "ip_address": "10.0.0.99",
            "location": "yard-east",
            "area": "yard",
            "channel_number": 9,
            "project_camera_code": "INS-0099",
        },
    )

    assert response.status_code == 201
    payload = response.get_json()
    assert payload["slot_id"]
    assert payload["slot_code"]

    conn = sqlite3.connect(project_test_db)
    try:
        camera = conn.execute(
            """
            SELECT station_id, project_id, slot_id, ip_address, location_desc, status, project_camera_code
            FROM cameras
            WHERE id = ?
            """,
            (payload["camera_id"],),
        ).fetchone()
        slot_row = conn.execute(
            """
            SELECT station_id, project_id, location_desc, area, channel_number
            FROM camera_slots
            WHERE id = ?
            """,
            (payload["slot_id"],),
        ).fetchone()
    finally:
        conn.close()

    assert camera == (1, 2, payload["slot_id"], "10.0.0.99", "yard-east", "active", "INS-0099")
    assert slot_row == (1, 2, "yard-east", "yard", 9)


def test_admin_replace_camera_preserves_slot_history(client, seeded_project_schema, project_test_db):
    login(client, "admin1", "adminpass")

    response = client.post(
        "/admin/cameras/1/replace",
        json={
            "project": "inspection",
            "ip_address": "10.0.0.88",
            "project_camera_code": "INS-NEW-001",
            "camera_index": "1A",
            "channel_number": 1,
        },
    )

    assert response.status_code == 201
    payload = response.get_json()
    assert payload["old_camera_id"] == 1
    assert payload["project"] == "inspection"

    conn = sqlite3.connect(project_test_db)
    try:
        old_camera = conn.execute(
            """
            SELECT status, slot_id, project_id, replaced_by_camera_id, retired_at
            FROM cameras
            WHERE id = 1
            """
        ).fetchone()
        new_camera = conn.execute(
            """
            SELECT id, status, slot_id, project_id, ip_address, project_camera_code, camera_index
            FROM cameras
            WHERE id = ?
            """,
            (payload["new_camera_id"],),
        ).fetchone()
        active_count = conn.execute(
            "SELECT COUNT(*) FROM cameras WHERE slot_id = 1 AND status = 'active'"
        ).fetchone()[0]
    finally:
        conn.close()

    assert old_camera[0] == "replaced"
    assert old_camera[1] == 1
    assert old_camera[2] == 2
    assert old_camera[3] == payload["new_camera_id"]
    assert old_camera[4] is not None
    assert new_camera == (
        payload["new_camera_id"],
        "active",
        1,
        2,
        "10.0.0.88",
        "INS-NEW-001",
        "1A",
    )
    assert active_count == 1


def test_admin_replace_camera_rejects_non_active_camera(client, seeded_project_schema):
    login(client, "admin1", "adminpass")

    first = client.post(
        "/admin/cameras/1/replace",
        json={"project": "inspection", "ip_address": "10.0.0.88"},
    )
    assert first.status_code == 201

    second = client.post(
        "/admin/cameras/1/replace",
        json={"project": "inspection", "ip_address": "10.0.0.89"},
    )
    assert second.status_code == 409
    assert "active" in second.get_json()["error"]


def test_admin_upload_excel_requires_project_in_multi_project_mode(client, seeded_project_schema):
    login(client, "admin1", "adminpass")

    response = client.post(
        "/admin/upload",
        data={
            "county": "County A",
            "file": (BytesIO(b"fake excel"), "import.xlsx"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert "project" in response.get_json()["error"]


def test_admin_upload_excel_assigns_selected_project(client, seeded_project_schema, project_test_db, monkeypatch):
    login(client, "admin1", "adminpass")

    def fake_parse_excel(_filepath):
        return {
            "station": {
                "name": "Station Upload",
                "voltage_level": "35kV",
                "county": "",
                "location": "Upload Yard",
                "ip_range": "",
                "nvr_ip": "",
                "nvr_port": None,
            },
            "cameras": [
                {
                    "camera_index": "3",
                    "area": "Upload Area",
                    "location": "Upload Slot",
                    "ip_address": "10.2.0.3",
                    "channel_port": None,
                    "channel_number": 3,
                }
            ],
        }

    monkeypatch.setattr("admin.parse_excel_admin", fake_parse_excel)

    response = client.post(
        "/admin/upload",
        data={
            "county": "County Upload",
            "project": "inspection",
            "file": (BytesIO(b"fake excel"), "import.xlsx"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["project"] == "inspection"
    assert payload["batch_id"] > 0
    assert payload["result_url"] == f"/admin/import-batches/{payload['batch_id']}"

    page = client.get("/admin")
    assert page.status_code == 200
    assert b"admin-section-header" in page.data
    assert b"admin-section-header-title" in page.data
    assert b"admin-section-title" in page.data
    assert b"renderBlockMessage(" in page.data
    assert "导入结果摘要页".encode("utf-8") in page.data
    assert "导入审查中心".encode("utf-8") in page.data
    assert b"upload-result" in page.data
    assert b"upload-feedback-card" in page.data
    assert b"form-input-file" in page.data
    assert b"admin-side-note-steps" in page.data
    assert b"station-search-input" in page.data
    assert b"is-hidden" in page.data

    conn = sqlite3.connect(project_test_db)
    try:
        station = conn.execute(
            "SELECT id, county FROM stations WHERE name = 'Station Upload' AND voltage_level = '35kV'"
        ).fetchone()
        slot_row = conn.execute(
            """
            SELECT project_id, location_desc, area, channel_number
            FROM camera_slots
            WHERE station_id = ?
            """,
            (station[0],),
        ).fetchone()
        camera = conn.execute(
            """
            SELECT project_id, slot_id, ip_address, location_desc, area, status
            FROM cameras
            WHERE station_id = ?
            """,
            (station[0],),
        ).fetchone()
        batch = conn.execute(
            """
            SELECT project_id, source_type, mode, file_count, success_count, fail_count, report_path
            FROM import_batches
            WHERE id = ?
            """,
            (payload["batch_id"],),
        ).fetchone()
    finally:
        conn.close()

    assert station == (payload["station_id"], "County Upload")
    assert slot_row == (2, "Upload Slot", "Upload Area", 3)
    assert camera == (2, 3, "10.2.0.3", "Upload Slot", "Upload Area", "active")
    assert batch[:6] == (2, "import_excel", "best-effort", 1, 1, 0)
    assert batch[6]


def test_admin_upload_excel_uses_county_from_excel_when_not_selected(client, seeded_project_schema, project_test_db, monkeypatch):
    login(client, "admin1", "adminpass")

    def fake_parse_excel(_filepath):
        return {
            "station": {
                "name": "Station Inferred County",
                "voltage_level": "110kV",
                "county": "Excel County",
                "location": "Upload Yard",
                "ip_range": "",
                "nvr_ip": "",
                "nvr_port": None,
            },
            "cameras": [
                {
                    "camera_index": "5",
                    "area": "Upload Area",
                    "location": "Upload Slot",
                    "ip_address": "10.2.0.5",
                    "channel_port": None,
                    "channel_number": 5,
                }
            ],
        }

    monkeypatch.setattr("admin.parse_excel_admin", fake_parse_excel)

    response = client.post(
        "/admin/upload",
        data={
            "project": "inspection",
            "file": (BytesIO(b"fake excel"), "import.xlsx"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200

    conn = sqlite3.connect(project_test_db)
    try:
        station = conn.execute(
            "SELECT county FROM stations WHERE name = 'Station Inferred County' AND voltage_level = '110kV'"
        ).fetchone()
    finally:
        conn.close()

    assert station == ("Excel County",)


def test_admin_upload_excel_requires_county_when_missing_from_input_and_excel(client, seeded_project_schema, monkeypatch):
    login(client, "admin1", "adminpass")

    def fake_parse_excel(_filepath):
        return {
            "station": {
                "name": "Station Missing County",
                "voltage_level": "110kV",
                "county": "",
                "location": "Upload Yard",
                "ip_range": "",
                "nvr_ip": "",
                "nvr_port": None,
            },
            "cameras": [],
        }

    monkeypatch.setattr("admin.parse_excel_admin", fake_parse_excel)

    response = client.post(
        "/admin/upload",
        data={
            "project": "inspection",
            "file": (BytesIO(b"fake excel"), "import.xlsx"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert "确定县区" in response.get_json()["error"]


def test_admin_upload_excel_rejects_inventory_without_cameras(client, seeded_project_schema, monkeypatch):
    login(client, "admin1", "adminpass")

    def fake_parse_excel(_filepath):
        return {
            "station": {
                "name": "Station No Cameras",
                "voltage_level": "110kV",
                "county": "County Upload",
                "location": "Upload Yard",
                "ip_range": "",
                "nvr_ip": "",
                "nvr_port": None,
            },
            "cameras": [],
        }

    monkeypatch.setattr("admin.parse_excel_admin", fake_parse_excel)

    response = client.post(
        "/admin/upload",
        data={
            "project": "inspection",
            "county": "County Upload",
            "file": (BytesIO(b"fake excel"), "import.xlsx"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert "未识别到摄像头" in response.get_json()["error"]


def test_admin_upload_daily_fault_summary_creates_fault_records_and_review_items(client, seeded_project_schema, project_test_db):
    login(client, "admin1", "adminpass")

    conn = sqlite3.connect(project_test_db)
    conn.executescript(
        """
        ALTER TABLE fault_reports ADD COLUMN source_batch_id TEXT;
        ALTER TABLE fault_reports ADD COLUMN source_record_key TEXT;

        CREATE TABLE station_name_mapping_proposals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_batch_id INTEGER,
            project_id INTEGER,
            source_system TEXT NOT NULL,
            external_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            candidate_station_id INTEGER,
            confidence_score REAL,
            raw_context_json TEXT,
            status TEXT NOT NULL DEFAULT 'pending',
            reviewer_id INTEGER,
            reviewed_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE fault_import_review_queue (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_batch_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            source_type TEXT NOT NULL,
            source_record_key_candidate TEXT,
            raw_payload_json TEXT NOT NULL,
            issue_type TEXT NOT NULL,
            issue_detail TEXT,
            status TEXT NOT NULL DEFAULT 'pending',
            resolved_fault_id INTEGER,
            reviewer_id INTEGER,
            reviewed_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    conn.commit()
    conn.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["变电站视频系统监控日报", "", ""])
    sheet.append(["时间：  04 月 07 日", "", "检查人员：正好科技"])
    sheet.append(["检查发现问题情况", "", ""])
    sheet.append(["变电站", "问题描述", ""])
    sheet.append(["省公司平台离线摄像头", "", ""])
    sheet.append(["Station A", "主变西北侧球机离线", ""])
    sheet.append(["Unknown Station", "大门口球机离线", ""])

    file_stream = BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    response = client.post(
        "/admin/upload",
        data={
            "project": "unified",
            "import_type": "daily_fault_summary",
            "file": (file_stream, "变电站视频系统监控日报20260407.xlsx"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["import_type"] == "daily_fault_summary"
    assert payload["project"] == "unified"
    assert payload["faults_added"] == 1
    assert payload["queued_count"] == 1
    assert payload["proposal_count"] == 1
    assert payload["batch_id"] > 0
    assert payload["source_date"] == "2026-04-07"
    assert "ai_status" in payload
    assert "status" in payload["ai_status"]
    assert "message" in payload["ai_status"]

    conn = sqlite3.connect(project_test_db)
    try:
        fault = conn.execute(
            """
            SELECT project_id, source_type, source_batch_id, source_record_key, fault_type_label_snapshot, status
            FROM fault_reports
            WHERE source_type = 'import_daily_fault_summary'
            """
        ).fetchone()
        review_count = conn.execute(
            "SELECT COUNT(*) FROM fault_import_review_queue WHERE source_type = 'import_daily_fault_summary'"
        ).fetchone()[0]
        proposal_count = conn.execute(
            "SELECT COUNT(*) FROM station_name_mapping_proposals WHERE source_system = 'daily_fault_summary'"
        ).fetchone()[0]
        batch = conn.execute(
            """
            SELECT project_id, source_type, mode, success_count, fail_count, report_path
            FROM import_batches
            WHERE id = ?
            """,
            (payload["batch_id"],),
        ).fetchone()
    finally:
        conn.close()

    assert fault[0] == 1
    assert fault[1] == "import_daily_fault_summary"
    assert fault[2] == str(payload["batch_id"])
    assert fault[3].startswith("unified:import_daily_fault_summary:")
    assert fault[4] == "摄像头离线"
    assert fault[5] == "open"
    assert review_count == 1
    assert proposal_count == 1
    assert batch[:5] == (1, "import_daily_fault_summary", "best-effort", 1, 1)
    assert batch[5]


def test_admin_upload_excel_legacy_returns_batch_result_redirect(legacy_upload_client, seeded_legacy_upload_schema, project_test_db, monkeypatch):
    login(legacy_upload_client, "admin1", "adminpass")

    def fake_parse_excel(_filepath):
        return {
            "station": {
                "name": "Legacy Upload Station",
                "voltage_level": "110kV",
                "county": "Legacy County",
                "location": "Legacy Yard",
                "ip_range": "",
                "nvr_ip": "",
                "nvr_port": None,
            },
            "cameras": [
                {
                    "camera_index": "1",
                    "area": "Legacy Area",
                    "location": "Legacy Slot",
                    "ip_address": "10.3.0.1",
                    "channel_port": None,
                    "channel_number": 1,
                },
                {
                    "camera_index": "2",
                    "area": "Legacy Area",
                    "location": "Legacy Slot 2",
                    "ip_address": "10.3.0.2",
                    "channel_port": None,
                    "channel_number": 2,
                },
            ],
        }

    monkeypatch.setattr("admin.parse_excel_admin", fake_parse_excel)

    response = legacy_upload_client.post(
        "/admin/upload",
        data={
            "file": (BytesIO(b"fake excel"), "legacy-import.xlsx"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["cameras_added"] == 2
    assert payload["project"] == "unified"
    assert payload["batch_id"] > 0
    assert payload["result_url"] == f"/admin/import-batches/{payload['batch_id']}"

    conn = sqlite3.connect(project_test_db)
    try:
        station = conn.execute(
            "SELECT id, county FROM stations WHERE name = 'Legacy Upload Station' AND voltage_level = '110kV'"
        ).fetchone()
        camera_count = conn.execute(
            "SELECT COUNT(*) FROM cameras WHERE station_id = ?",
            (station[0],),
        ).fetchone()[0]
        batch = conn.execute(
            """
            SELECT project_id, source_type, mode, file_count, success_count, fail_count, report_path
            FROM import_batches
            WHERE id = ?
            """,
            (payload["batch_id"],),
        ).fetchone()
    finally:
        conn.close()

    assert station == (payload["station_id"], "Legacy County")
    assert camera_count == 2
    assert batch[:6] == (1, "import_excel", "best-effort", 1, 1, 0)
    assert batch[6]


def test_statistics_export_respects_project_scope(client, seeded_project_schema):
    login(client, "operator1", "operatorpass")

    response = client.get("/api/statistics/export")
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    overview_sheet = workbook["概览"]
    assert overview_sheet["B4"].value == 1
    assert overview_sheet["B5"].value == 1
    assert overview_sheet["B6"].value == 1


def test_statistics_export_operator_hides_audit_columns(client, seeded_project_schema):
    login(client, "operator1", "operatorpass")

    response = client.get("/api/statistics/export")
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    detail_sheet = workbook["故障明细"]
    headers = [cell.value for cell in detail_sheet[1]]
    values = [cell.value for cell in detail_sheet[2]]

    assert "semantic_group" in headers
    assert "来源幂等键" not in headers
    assert "原始时间" not in headers
    assert "遗留系统类型" not in headers
    assert values[1] == "inspection"


def test_statistics_export_admin_includes_audit_columns(client, seeded_project_schema):
    login(client, "admin1", "adminpass")

    response = client.get("/api/statistics/export?project=inspection")
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.data))
    detail_sheet = workbook["故障明细"]
    headers = [cell.value for cell in detail_sheet[1]]
    header_index = {header: index for index, header in enumerate(headers)}
    values = [cell.value for cell in detail_sheet[2]]

    assert "来源幂等键" in headers
    assert "原始时间" in headers
    assert "原始时区" in headers
    assert "遗留系统类型" in headers
    assert values[header_index["原始时间"]] == "2026-04-02 08:00:00"
    assert values[header_index["原始时区"]] == "Asia/Shanghai"


def test_guest_access_is_limited_to_home_and_statistics(client, seeded_project_schema):
    assert client.get("/").status_code == 200
    assert client.get("/statistics").status_code == 200

    assert client.get("/design/style2").status_code == 302
    assert client.get("/design/style2/statistics").status_code == 302

    assert client.get("/api/stats").status_code == 401
    assert client.get("/api/projects").status_code == 401
    assert client.get("/api/statistics/export").status_code == 401

    assert client.get("/faults").status_code == 302
    assert client.get("/photos").status_code == 302
    assert client.get("/map").status_code == 302
    assert client.get("/admin").status_code == 302

    assert client.get("/api/faults").status_code == 401
    assert client.get("/api/stations").status_code == 401
    assert client.get("/api/photos").status_code == 401


def test_photos_endpoints_respect_project_scope(client, seeded_project_schema, tmp_path, monkeypatch):
    login(client, "operator1", "operatorpass")

    photo_root = tmp_path / "photo-root"
    photo_root.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr("config.Config.PHOTO_ROOT_PATH", str(photo_root))

    allowed_photo = photo_root / "allowed.jpg"
    allowed_photo.write_bytes(b"\xff\xd8\xff\xd9")
    hidden_photo = photo_root / "hidden.jpg"
    hidden_photo.write_bytes(b"\xff\xd8\xff\xd9")

    conn = sqlite3.connect(app.config["DATABASE_PATH"])
    conn.execute(
        """
        INSERT INTO photos
            (rel_path, abs_path, filename, ext, station_id, match_status, match_method, project_id, project_hint)
        VALUES
            (?, ?, 'allowed.jpg', '.jpg', 1, 'matched', 'manual', 2, 'inspection'),
            (?, ?, 'hidden.jpg', '.jpg', 2, 'matched', 'manual', 1, 'unified')
        """,
        (str(allowed_photo.name), str(allowed_photo), str(hidden_photo.name), str(hidden_photo)),
    )
    allowed_id = conn.execute(
        "SELECT id FROM photos WHERE filename = 'allowed.jpg'"
    ).fetchone()[0]
    hidden_id = conn.execute(
        "SELECT id FROM photos WHERE filename = 'hidden.jpg'"
    ).fetchone()[0]
    conn.commit()
    conn.close()

    photos_response = client.get("/api/photos")
    assert photos_response.status_code == 200
    photos_data = photos_response.get_json()
    assert photos_data["total"] == 1
    assert photos_data["photos"][0]["filename"] == "allowed.jpg"

    groups_response = client.get("/api/photos/groups")
    assert groups_response.status_code == 200
    groups_data = groups_response.get_json()
    assert groups_data["group_count"] == 1
    assert groups_data["groups"][0]["station_id"] == 1

    allowed_file = client.get(f"/photos/file/{allowed_id}")
    assert allowed_file.status_code == 200

    denied_file = client.get(f"/photos/file/{hidden_id}")
    assert denied_file.status_code == 403
    assert denied_file.get_json()["code"] == "PROJECT_ACCESS_DENIED"


def test_photo_groups_fault_filter_respects_project_scope(client, seeded_project_schema, project_test_db, tmp_path, monkeypatch):
    login(client, "operator1", "operatorpass")

    photo_root = tmp_path / "photo-root"
    photo_root.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr("config.Config.PHOTO_ROOT_PATH", str(photo_root))

    fault_photo = photo_root / "fault-only.jpg"
    fault_photo.write_bytes(b"\xff\xd8\xff\xd9")
    no_fault_photo = photo_root / "no-fault.jpg"
    no_fault_photo.write_bytes(b"\xff\xd8\xff\xd9")

    conn = sqlite3.connect(project_test_db)
    conn.execute(
        """
        INSERT INTO photos
            (rel_path, abs_path, filename, ext, station_id, match_status, match_method, project_id, project_hint)
        VALUES
            (?, ?, 'fault-only.jpg', '.jpg', 1, 'matched', 'manual', 2, 'inspection'),
            (?, ?, 'no-fault.jpg', '.jpg', 2, 'matched', 'manual', 2, 'inspection')
        """,
        (str(fault_photo.name), str(fault_photo), str(no_fault_photo.name), str(no_fault_photo)),
    )
    conn.execute(
        """
        INSERT INTO fault_reports
            (station_id, camera_id, fault_type, reporter_name, status, project_id, fault_type_label_snapshot, source_type)
        VALUES
            (1, NULL, 'Blur', 'Carol', 'open', 2, 'Blur', 'manual'),
            (2, NULL, 'Blur', 'Carol', 'open', 1, 'Blur', 'manual')
        """
    )
    conn.commit()
    conn.close()

    response = client.get("/api/photos/groups?has_fault=1")
    assert response.status_code == 200
    payload = response.get_json()

    assert payload["group_count"] == 1
    assert payload["groups"][0]["station_id"] == 1
    assert payload["groups"][0]["station_name"] == "Station A"


def test_fault_detail_includes_same_slot_history(client, seeded_project_schema, project_test_db):
    conn = sqlite3.connect(project_test_db)
    conn.execute(
        """
        INSERT INTO fault_reports
            (id, station_id, camera_id, fault_type, reporter_name, status, project_id, camera_slot_id, fault_type_label_snapshot, source_type)
        VALUES
            (3, 1, 1, 'Blur', 'Carol', 'closed', 2, 1, 'Blur', 'manual')
        """
    )
    conn.commit()
    conn.close()

    login(client, "operator1", "operatorpass")

    response = client.get("/api/faults/2/detail")
    assert response.status_code == 200
    data = response.get_json()

    assert data["fault"]["camera_slot_id"] == 1
    assert data["fault"]["slot_history_count"] == 1
    assert data["fault"]["slot_history"][0]["id"] == 3
    assert data["fault"]["slot_history"][0]["fault_label"] == "Blur"


def test_fault_detail_includes_assignment_metadata(client, seeded_project_schema, project_test_db):
    conn = sqlite3.connect(project_test_db)
    conn.execute(
        """
        UPDATE fault_reports
        SET assigned_to = 2,
            handling_started_at = '2026-04-02T08:30:00'
        WHERE id = 2
        """
    )
    conn.commit()
    conn.close()

    login(client, "operator1", "operatorpass")

    response = client.get("/api/faults/2/detail")
    assert response.status_code == 200
    data = response.get_json()

    assert data["fault"]["assigned_to"] == 2
    assert data["fault"]["assigned_to_username"] == "operator1"
    assert data["fault"]["handling_started_at"] == "2026-04-02T08:30:00"


def test_fault_detail_includes_import_audit_metadata(client, seeded_project_schema):
    login(client, "operator1", "operatorpass")

    response = client.get("/api/faults/2/detail")
    assert response.status_code == 200
    data = response.get_json()

    assert data["fault"]["camera_slot_id"] == 1
    assert data["fault"]["project_device_code"] == "INSPECT-CAM-001"
    assert data["fault"]["source_time_raw"] == "2026-04-02 08:00:00"
    assert data["fault"]["source_timezone"] == "Asia/Shanghai"


def test_fault_list_and_detail_fall_back_to_camera_location_text(client, seeded_project_schema, project_test_db):
    conn = sqlite3.connect(project_test_db)
    conn.execute("ALTER TABLE fault_reports ADD COLUMN camera_location_text TEXT")
    conn.execute(
        """
        UPDATE fault_reports
        SET camera_id = NULL,
            camera_slot_id = NULL,
            camera_location_text = '2号主变西北侧球机'
        WHERE id = 2
        """
    )
    conn.commit()
    conn.close()

    login(client, "operator1", "operatorpass")

    list_response = client.get("/api/faults?project=inspection")
    assert list_response.status_code == 200
    list_payload = list_response.get_json()
    fault = next(item for item in list_payload["faults"] if item["id"] == 2)
    assert fault["camera_location"] == "2号主变西北侧球机"

    detail_response = client.get("/api/faults/2/detail")
    assert detail_response.status_code == 200
    detail_payload = detail_response.get_json()
    assert detail_payload["fault"]["camera_location"] == "2号主变西北侧球机"
