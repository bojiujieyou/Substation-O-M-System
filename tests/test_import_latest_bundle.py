import csv
import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook

from app import build_station_slots_payload
from import_latest_bundle import import_inventory_bundle
from init_db import init_db


@pytest.fixture
def latest_bundle_db(tmp_path):
    return str(tmp_path / "test_latest_bundle.db")


@pytest.fixture
def seeded_latest_bundle_schema(latest_bundle_db):
    import config as config_module

    original_path = config_module.Config.DATABASE_PATH
    config_module.Config.DATABASE_PATH = latest_bundle_db
    init_db(force=True)
    config_module.Config.DATABASE_PATH = original_path

    conn = sqlite3.connect(latest_bundle_db)
    conn.executescript(
        """
        CREATE TABLE projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            short_name TEXT NOT NULL,
            color TEXT NOT NULL DEFAULT '#1a73e8',
            fault_type_version_id INTEGER,
            sort_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE camera_slots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slot_code TEXT NOT NULL,
            station_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            location_desc TEXT NOT NULL DEFAULT '',
            area TEXT NOT NULL DEFAULT '',
            channel_number INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(station_id, project_id, slot_code)
        );

        DROP TABLE cameras;
        CREATE TABLE cameras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slot_id INTEGER,
            station_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            project_camera_code TEXT,
            camera_index TEXT,
            area TEXT,
            location_desc TEXT,
            ip_address TEXT,
            channel_port INTEGER,
            channel_number INTEGER,
            status TEXT DEFAULT 'active',
            replaced_by_camera_id INTEGER,
            retired_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE UNIQUE INDEX idx_cameras_one_active_per_slot
        ON cameras(slot_id)
        WHERE status = 'active';
        """
    )
    conn.execute(
        """
        INSERT INTO projects (id, code, name, short_name, color, sort_order, is_active)
        VALUES (1, 'unified', 'Unified', 'UNI', '#1a73e8', 1, 1)
        """
    )
    conn.commit()
    conn.close()
    yield latest_bundle_db


def build_inventory_workbook(path: Path, station_name: str, cameras: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = station_name
    ws["A2"] = "110kV变电站"
    ws["A17"] = "通道"
    ws["B17"] = "位置"
    ws["C17"] = "区域"
    ws["D17"] = "IP地址"
    ws["E17"] = "槽位编码"
    ws["F17"] = "设备编号"

    start_row = 18
    for offset, camera in enumerate(cameras):
        row = start_row + offset
        ws[f"A{row}"] = f"通道{camera['channel']}"
        ws[f"B{row}"] = camera["location"]
        ws[f"C{row}"] = camera.get("area", "")
        ws[f"D{row}"] = camera["ip"]
        ws[f"E{row}"] = camera.get("slot_code", "")
        ws[f"F{row}"] = camera.get("project_camera_code", "")

    wb.save(path)


def write_gbk_csv(path: Path, rows: list[list[str]]):
    with path.open("w", encoding="gbk", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def test_import_inventory_bundle_uses_device_excel_camera_ips_and_keeps_recorder_mapping(
    seeded_latest_bundle_schema,
    tmp_path,
):
    station_csv = tmp_path / "变电站.csv"
    camera_csv = tmp_path / "摄像头.csv"
    device_root = tmp_path / "设备资料"
    county_dir = device_root / "CountyA"
    county_dir.mkdir(parents=True)
    build_inventory_workbook(
        county_dir / "Station A.xlsx",
        "110kVStation A",
        [
            {
                "channel": 1,
                "location": "Gate Cam #1",
                "ip": "192.168.5.10",
            }
        ],
    )

    write_gbk_csv(
        station_csv,
        [
            ["设备名称", "区域", "设备类型", "厂家", "IP地址", "端口", "用户名", "密码", "描述"],
            ["110kVStation A1", "Grid/CountyA/Station A", "录像机", "", "10.10.10.10", "8000", "", "", "NVR-1"],
        ],
    )
    write_gbk_csv(
        camera_csv,
        [
            ["监控点名称", "所在区域", "监控点类型", "所属设备", "IP地址及端口号", "所属设备编号", "通道号", "通道类型"],
            ["Gate Cam #1", "Grid/CountyA/Station A", "球机", "110kVStation A1", "10.10.10.10:8000", "REC-A-1", "1", "数字通道"],
        ],
    )

    report = import_inventory_bundle(
        station_csv=str(station_csv),
        camera_csv=str(camera_csv),
        database_path=seeded_latest_bundle_schema,
        project_code="unified",
        camera_ip_source_root=str(device_root),
    )

    assert report["camera_ip_station_seen"] == 1
    assert report["cameras_added"] == 1
    assert report["recorders_added"] == 1

    conn = sqlite3.connect(seeded_latest_bundle_schema)
    conn.row_factory = sqlite3.Row
    try:
        station = conn.execute("SELECT id, name FROM stations WHERE name = '110kVStation A'").fetchone()
        camera = conn.execute(
            """
            SELECT ip_address, channel_port, recorder_name, recorder_ip_address, recorder_port, project_camera_code, status
            FROM cameras
            WHERE station_id = ?
            """,
            (station["id"],),
        ).fetchone()
        recorder = conn.execute(
            """
            SELECT recorder_name, ip_address, port, status
            FROM station_recorders
            WHERE station_id = ?
            """,
            (station["id"],),
        ).fetchone()
        slots = build_station_slots_payload(
            conn,
            station["id"],
            {
                "enabled": True,
                "project_ids": [1],
                "requested_project": None,
                "projects": [],
            },
        )
    finally:
        conn.close()

    assert camera["ip_address"] == "192.168.5.10"
    assert camera["channel_port"] == 8000
    assert camera["recorder_name"] == "110kVStation A1"
    assert camera["recorder_ip_address"] == "10.10.10.10"
    assert camera["recorder_port"] == 8000
    assert camera["project_camera_code"] == "REC-A-1"
    assert camera["status"] == "active"

    assert recorder["recorder_name"] == "110kVStation A1"
    assert recorder["ip_address"] == "10.10.10.10"
    assert recorder["port"] == 8000
    assert recorder["status"] == "active"

    assert len(slots) == 1
    assert slots[0]["current_camera"]["ip_address"] == "192.168.5.10"
    assert slots[0]["recorder"]["recorder_name"] == "110kVStation A1"
    assert slots[0]["recorder"]["ip_address"] == "10.10.10.10"
