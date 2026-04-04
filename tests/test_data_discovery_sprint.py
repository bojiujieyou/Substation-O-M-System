import csv
import sqlite3
from pathlib import Path

from openpyxl import Workbook

from data_discovery_sprint import run_discovery


def build_inventory_workbook(path: Path, station_name: str, cameras: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = station_name
    ws["A17"] = "通道"
    ws["B17"] = "位置"
    ws["C17"] = "区域"
    ws["D17"] = "IP地址"
    ws["E17"] = "槽位编码"
    ws["F17"] = "设备编号"

    start_row = 18
    for offset, camera in enumerate(cameras):
        row = start_row + offset
        ws[f"A{row}"] = f"通道{camera['channel']}"
        ws[f"B{row}"] = camera["location"]
        ws[f"C{row}"] = camera.get("area", "")
        ws[f"D{row}"] = camera["ip"]
        ws[f"E{row}"] = camera.get("slot_code", "")
        ws[f"F{row}"] = camera.get("project_camera_code", "")

    wb.save(path)


def seed_discovery_db(db_path: Path) -> None:
    conn = sqlite3.connect(db_path)
    try:
        conn.executescript(
            """
            CREATE TABLE projects (
                id INTEGER PRIMARY KEY,
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                short_name TEXT NOT NULL,
                fault_type_version_id INTEGER
            );

            CREATE TABLE project_fault_types (
                id INTEGER PRIMARY KEY,
                version_id INTEGER NOT NULL,
                type_code TEXT NOT NULL,
                type_label TEXT NOT NULL,
                semantic_group TEXT,
                sort_order INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 1
            );

            CREATE TABLE stations (
                id INTEGER PRIMARY KEY,
                name TEXT NOT NULL,
                county TEXT
            );

            CREATE TABLE station_aliases (
                id INTEGER PRIMARY KEY,
                station_id INTEGER NOT NULL,
                alias TEXT NOT NULL,
                source TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE station_external_names (
                id INTEGER PRIMARY KEY,
                station_id INTEGER NOT NULL,
                source_system TEXT NOT NULL,
                external_name TEXT NOT NULL,
                normalized_name TEXT
            );

            CREATE TABLE camera_slots (
                id INTEGER PRIMARY KEY,
                slot_code TEXT NOT NULL,
                station_id INTEGER NOT NULL,
                project_id INTEGER NOT NULL,
                location_desc TEXT,
                area TEXT,
                channel_number INTEGER
            );

            CREATE TABLE cameras (
                id INTEGER PRIMARY KEY,
                slot_id INTEGER,
                station_id INTEGER NOT NULL,
                project_id INTEGER NOT NULL,
                project_camera_code TEXT,
                status TEXT DEFAULT 'active'
            );
            """
        )

        conn.execute(
            """
            INSERT INTO projects (id, code, name, short_name, fault_type_version_id)
            VALUES (1, 'unified', '统一平台', '统一', 100)
            """
        )
        conn.executemany(
            """
            INSERT INTO project_fault_types (id, version_id, type_code, type_label, semantic_group, sort_order, is_active)
            VALUES (?, ?, ?, ?, ?, ?, 1)
            """,
            [
                (1, 100, "NO_IMAGE", "无图像", "NO_IMAGE", 1),
                (2, 100, "BLUR", "模糊", "BLUR", 2),
            ],
        )
        conn.executemany(
            "INSERT INTO stations (id, name, county) VALUES (?, ?, ?)",
            [
                (1, "Alpha变电站", "测试县"),
                (2, "Beta变电站", "测试县"),
            ],
        )
        conn.execute(
            "INSERT INTO station_aliases (station_id, alias, source) VALUES (1, 'Alpha', 'manual')"
        )
        conn.execute(
            """
            INSERT INTO station_external_names (station_id, source_system, external_name, normalized_name)
            VALUES (1, 'import_excel', '外部Alpha', '外部alpha')
            """
        )
        conn.executemany(
            """
            INSERT INTO camera_slots (id, slot_code, station_id, project_id, location_desc, area, channel_number)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (1, "SLOT_A1", 1, 1, "North 1", "Area A", 1),
                (2, "SLOT_A2", 1, 1, "South 1", "Area A", 2),
            ],
        )
        conn.executemany(
            """
            INSERT INTO cameras (id, slot_id, station_id, project_id, project_camera_code, status)
            VALUES (?, ?, ?, ?, ?, 'active')
            """,
            [
                (1, 1, 1, 1, "CAM-1"),
                (2, 2, 1, 1, "CAM-2"),
            ],
        )
        conn.commit()
    finally:
        conn.close()


def test_run_discovery_generates_threshold_report(tmp_path):
    db_path = tmp_path / "discovery.db"
    seed_discovery_db(db_path)

    device_dir = tmp_path / "device_samples"
    device_dir.mkdir()
    device_file = device_dir / "alpha.xlsx"
    build_inventory_workbook(
        device_file,
        "Alpha变电站",
        [
            {"channel": 1, "location": "North 1", "area": "Area A", "ip": "10.0.0.1"},
            {"channel": 1, "location": "North-1", "area": "Area B", "ip": "10.0.0.2"},
            {"channel": 2, "location": "South 1", "area": "Area A", "ip": "10.0.0.3", "slot_code": "SLOT_FIXED"},
        ],
    )

    fault_file = tmp_path / "faults.csv"
    with fault_file.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "external_id",
                "station_name",
                "fault_type_label",
                "occurred_at",
                "source_timezone",
                "description",
            ],
        )
        writer.writeheader()
        writer.writerow(
            {
                "external_id": "A1",
                "station_name": "外部Alpha",
                "fault_type_label": "无图像",
                "occurred_at": "2026-04-01 08:00",
                "source_timezone": "Asia/Shanghai",
                "description": "camera offline",
            }
        )
        writer.writerow(
            {
                "external_id": "A2",
                "station_name": "Unknown Station",
                "fault_type_label": "未知类型",
                "occurred_at": "bad-time",
                "source_timezone": "Asia/Shanghai",
                "description": "bad row",
            }
        )
        writer.writerow(
            {
                "external_id": "",
                "station_name": "Alpha",
                "fault_type_label": "历史模糊",
                "occurred_at": "2026/04/01 09:30",
                "source_timezone": "Asia/Shanghai",
                "description": "legacy blur",
            }
        )

    type_mapping_file = tmp_path / "type_mapping.csv"
    with type_mapping_file.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["source_value", "target_code"])
        writer.writeheader()
        writer.writerow({"source_value": "历史模糊", "target_code": "BLUR"})

    report_path = tmp_path / "report.json"
    summary_path = tmp_path / "report.md"
    report = run_discovery(
        project_code="unified",
        source_type="import_excel",
        device_sources=[device_dir],
        fault_sources=[fault_file],
        database=db_path,
        type_mapping_path=type_mapping_file,
        report_path=report_path,
        summary_path=summary_path,
    )

    assert report_path.exists()
    assert summary_path.exists()

    device_report = report["device_inventory"]
    assert device_report["camera_count"] == 3
    assert device_report["generated_slot_code_count"] == 2
    assert device_report["conflicting_row_count"] == 0
    assert device_report["slot_code_threshold"]["status"] == "pass"

    fault_report = report["fault_history"]
    assert fault_report["row_count"] == 3
    assert fault_report["fault_type_mapping_breakdown"]["catalog_label"] == 1
    assert fault_report["fault_type_mapping_breakdown"]["mapping_file"] == 1
    assert fault_report["fault_type_mapping_breakdown"]["unmapped"] == 1
    assert fault_report["station_match_breakdown"]["station_aliases"] == 1
    assert fault_report["station_match_breakdown"]["unresolved"] == 1
    assert sum(fault_report["station_match_breakdown"].values()) == 3
    assert fault_report["timestamp_threshold"]["status"] == "block"
    assert fault_report["source_record_key_breakdown"]["canonical_fallback"] == 1

    assert report["performance"]["status"] in {"ready", "review_cache"}
    assert report["release_decision"]["overall"] == "block"

    summary = summary_path.read_text(encoding="utf-8")
    assert "Overall Decision: block" in summary
