#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Import camera IPs and platform recorder IPs from station workbooks."""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from import_latest_bundle import load_station_rows, resolve_existing_station
from import_review_support import get_project_row
from init_db import get_db_path
from parse_excel import parse_station_excel
from utils import backup_sqlite_database, create_db_connection


DEFAULT_PORT = 8000
PLATFORM_LABELS = ("市公司平台", "省公司平台")
RETIRABLE_RECORDER_SOURCES = ("inventory_csv", "device_workbook_platform")
IPV4_RE = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}$")


def clean(value):
    return str(value or "").replace("\t", "").strip()


def normalize_text(value):
    text = clean(value).lower()
    text = re.sub(r"\s+", "", text)
    return text


def base_station_name(value):
    text = clean(value)
    text = re.sub(r"^\d+\s*[kK][vV]\s*", "", text)
    return text.strip()


def normalize_location(value):
    text = normalize_text(value)
    for token in ("室外", "室内", "摄像机", "摄像头", "球机", "枪机", "球", "枪", "半球"):
        text = text.replace(token, "")
    text = re.sub(r"[-_/\\|,.，。:：;；（）()#]+", "", text)
    return text


def iter_workbook_files(source_root):
    root = Path(source_root)
    if not root.exists():
        return
    for pattern in ("*.xlsx", "*.xls"):
        for path in sorted(root.rglob(pattern)):
            name = path.name
            if name.startswith("~$") or "冲突副本" in name:
                continue
            yield path


def load_rows(filepath):
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def load_station_rows_for_match(conn):
    rows = load_station_rows(conn)
    if rows and isinstance(rows[0], tuple):
        query_rows = conn.execute(
            """
            SELECT id, name, county, voltage_level, location, ip_range, nvr_ip, nvr_port
            FROM stations
            ORDER BY id
            """
        ).fetchall()
        return [
            {
                "id": row[0],
                "name": row[1],
                "county": row[2],
                "voltage_level": row[3],
                "location": row[4],
                "ip_range": row[5],
                "nvr_ip": row[6],
                "nvr_port": row[7],
            }
            for row in query_rows
        ]
    return rows


def parse_platform_network(raw_value):
    parts = [clean(part) for part in clean(raw_value).split("/") if clean(part)]
    if not parts:
        return {"raw": "", "ip": "", "mask": "", "gateway": ""}
    if not IPV4_RE.match(parts[0]):
        return {"raw": "", "ip": "", "mask": "", "gateway": ""}
    remaining = parts[1:]
    mask = ""
    gateway = ""
    for part in remaining:
        if not mask and IPV4_RE.match(part) and part.startswith("255."):
            mask = part
            continue
        if not gateway and IPV4_RE.match(part):
            gateway = part
    if not mask:
        for part in remaining:
            if IPV4_RE.match(part) and part != gateway:
                mask = part
                break
    return {
        "raw": clean(raw_value),
        "ip": parts[0],
        "mask": mask,
        "gateway": gateway,
    }


def parse_platform_rows(rows):
    platforms = {}
    for row in rows[:20]:
        cells = list(row or [])
        second = clean(cells[1]) if len(cells) > 1 else ""
        third = clean(cells[2]) if len(cells) > 2 else ""
        if second not in PLATFORM_LABELS:
            continue

        bucket = platforms.setdefault(
            second,
            {
                "platform_name": second,
                "raw_network": "",
                "ip": "",
                "mask": "",
                "gateway": "",
                "cabinet_location": "",
            },
        )
        if third:
            parsed = parse_platform_network(third)
            if parsed["ip"]:
                bucket.update(
                    {
                        "raw_network": parsed["raw"],
                        "ip": parsed["ip"],
                        "mask": parsed["mask"],
                        "gateway": parsed["gateway"],
                    }
                )
            elif not bucket["cabinet_location"]:
                bucket["cabinet_location"] = third
    return platforms


def build_recorder_description(platform, workbook_name):
    parts = [f"来源工作簿: {workbook_name}"]
    if platform.get("raw_network"):
        parts.append(f"网络配置: {platform['raw_network']}")
    if platform.get("mask"):
        parts.append(f"子网掩码: {platform['mask']}")
    if platform.get("gateway"):
        parts.append(f"网关: {platform['gateway']}")
    if platform.get("cabinet_location"):
        parts.append(f"机柜位置: {platform['cabinet_location']}")
    return " | ".join(parts)


def choose_primary_platform(platforms):
    for label in PLATFORM_LABELS:
        platform = platforms.get(label)
        if platform and platform.get("ip"):
            return platform
    for platform in platforms.values():
        if platform.get("ip"):
            return platform
    return None


def workbook_name_matches_station(workbook_path, station_name):
    workbook_label = normalize_text(base_station_name(Path(workbook_path).stem))
    station_label = normalize_text(base_station_name(station_name))
    if not workbook_label or not station_label:
        return True
    return workbook_label in station_label or station_label in workbook_label


def is_local_recorder_name(recorder_name, station_name):
    recorder_label = normalize_text(recorder_name)
    station_label = normalize_text(base_station_name(station_name))
    if not recorder_label or not station_label:
        return False
    return bool(re.fullmatch(rf"{re.escape(station_label)}\d+", recorder_label))


def build_camera_maps(rows):
    by_channel = defaultdict(list)
    by_location = defaultdict(list)
    for row in rows:
        if row["channel_number"] not in (None, ""):
            by_channel[int(row["channel_number"])].append(row)
        location_key = normalize_location(row["location_desc"])
        if location_key:
            by_location[location_key].append(row)
    return by_channel, by_location


def infer_channel_offset(existing_rows, workbook_cameras):
    existing_channels = sorted(
        int(row["channel_number"])
        for row in existing_rows
        if row["channel_number"] not in (None, "")
    )
    workbook_channels = sorted(
        int(camera["channel_number"])
        for camera in workbook_cameras
        if camera.get("channel_number") not in (None, "")
    )
    if not existing_channels or len(existing_channels) != len(workbook_channels):
        return None
    offsets = {existing - workbook for existing, workbook in zip(existing_channels, workbook_channels)}
    if len(offsets) == 1:
        return offsets.pop()
    return None


def resolve_camera_match(camera_row, by_channel, by_location, used_ids, channel_offset=None):
    location_key = normalize_location(camera_row.get("location_desc") or camera_row.get("location"))
    channel_number = camera_row.get("channel_number")

    if location_key:
        location_matches = [row for row in by_location.get(location_key, []) if row["id"] not in used_ids]
        if len(location_matches) == 1:
            return location_matches[0]

        fuzzy_matches = {}
        for candidate_key, candidate_rows in by_location.items():
            if not candidate_key:
                continue
            if location_key in candidate_key or candidate_key in location_key:
                for row in candidate_rows:
                    if row["id"] not in used_ids:
                        fuzzy_matches[row["id"]] = row
        if len(fuzzy_matches) == 1:
            return next(iter(fuzzy_matches.values()))

    if channel_offset is not None and channel_number not in (None, ""):
        offset_matches = [
            row
            for row in by_channel.get(int(channel_number) + int(channel_offset), [])
            if row["id"] not in used_ids
        ]
        if len(offset_matches) == 1:
            return offset_matches[0]

    if channel_number not in (None, ""):
        channel_matches = [row for row in by_channel.get(int(channel_number), []) if row["id"] not in used_ids]
        if len(channel_matches) == 1:
            return channel_matches[0]

    return None


def should_replace_project_camera_code(camera_row):
    code = normalize_text(camera_row["project_camera_code"])
    location = normalize_text(camera_row["location_desc"])
    return not code or code == location


def update_camera_row(conn, *, existing_row, workbook_camera, primary_platform, station_name):
    channel_number = workbook_camera.get("channel_number")
    channel_text = str(channel_number) if channel_number not in (None, "") else existing_row["camera_index"]
    project_camera_code = existing_row["project_camera_code"]
    if channel_text and should_replace_project_camera_code(existing_row):
        project_camera_code = channel_text

    recorder_name = existing_row["recorder_name"]
    recorder_ip_address = existing_row["recorder_ip_address"]
    recorder_port = existing_row["recorder_port"]
    preserve_local_recorder = is_local_recorder_name(existing_row["recorder_name"], station_name)
    if primary_platform and primary_platform.get("ip") and not preserve_local_recorder:
        recorder_name = primary_platform["platform_name"]
        recorder_ip_address = primary_platform["ip"]
        recorder_port = DEFAULT_PORT

    conn.execute(
        """
        UPDATE cameras
        SET ip_address = ?,
            channel_number = ?,
            camera_index = ?,
            project_camera_code = ?,
            recorder_name = ?,
            recorder_ip_address = ?,
            recorder_port = ?
        WHERE id = ?
        """,
        (
            clean(workbook_camera.get("ip_address")) or existing_row["ip_address"],
            channel_number,
            channel_text,
            project_camera_code,
            recorder_name,
            recorder_ip_address,
            recorder_port,
            existing_row["id"],
        ),
    )
    if existing_row["slot_id"]:
        conn.execute(
            """
            UPDATE camera_slots
            SET channel_number = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (channel_number, existing_row["slot_id"]),
        )


def update_station_primary_platform(conn, *, station_id, primary_platform):
    if not primary_platform or not primary_platform.get("ip"):
        return
    conn.execute(
        """
        UPDATE stations
        SET nvr_ip = ?, nvr_port = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (primary_platform["ip"], DEFAULT_PORT, station_id),
    )


def sync_platform_recorders(conn, *, station_id, project_id, platforms, workbook_name):
    metrics = {
        "recorders_added": 0,
        "recorders_updated": 0,
        "recorders_retired": 0,
    }
    desired = []
    for label in PLATFORM_LABELS:
        platform = platforms.get(label)
        if not platform or not platform.get("ip"):
            continue
        desired.append(
            {
                "recorder_name": platform["platform_name"],
                "ip_address": platform["ip"],
                "port": DEFAULT_PORT,
                "description": build_recorder_description(platform, workbook_name),
                "source_key": f"{platform['platform_name']}|{platform['ip']}|{DEFAULT_PORT}",
            }
        )

    if not desired:
        return metrics

    existing_rows = conn.execute(
        """
        SELECT id, recorder_name, ip_address, port, source_type
        FROM station_recorders
        WHERE station_id = ? AND project_id = ? AND status = 'active'
          AND source_type IN (?, ?)
        """,
        (station_id, project_id, *RETIRABLE_RECORDER_SOURCES),
    ).fetchall()
    existing_by_key = {
        (clean(row["recorder_name"]), clean(row["ip_address"]), row["port"]): row
        for row in existing_rows
    }
    seen_ids = set()

    for recorder in desired:
        key = (clean(recorder["recorder_name"]), clean(recorder["ip_address"]), recorder["port"])
        existing = existing_by_key.get(key)
        if existing:
            conn.execute(
                """
                UPDATE station_recorders
                SET description = ?,
                    source_type = 'device_workbook_platform',
                    source_key = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (recorder["description"], recorder["source_key"], existing["id"]),
            )
            seen_ids.add(existing["id"])
            metrics["recorders_updated"] += 1
            continue

        cursor = conn.execute(
            """
            INSERT INTO station_recorders (
                station_id, project_id, recorder_name, ip_address, port,
                description, source_type, source_key, status
            )
            VALUES (?, ?, ?, ?, ?, ?, 'device_workbook_platform', ?, 'active')
            """,
            (
                station_id,
                project_id,
                recorder["recorder_name"],
                recorder["ip_address"],
                recorder["port"],
                recorder["description"],
                recorder["source_key"],
            ),
        )
        seen_ids.add(cursor.lastrowid)
        metrics["recorders_added"] += 1

    retire_ids = [row["id"] for row in existing_rows if row["id"] not in seen_ids]
    if retire_ids:
        placeholders = ", ".join(["?"] * len(retire_ids))
        conn.execute(
            f"""
            UPDATE station_recorders
            SET status = 'retired',
                retired_at = CURRENT_TIMESTAMP,
                updated_at = CURRENT_TIMESTAMP
            WHERE id IN ({placeholders})
            """,
            retire_ids,
        )
        metrics["recorders_retired"] += len(retire_ids)

    primary = desired[0]
    update_station_primary_platform(
        conn,
        station_id=station_id,
        primary_platform={"ip": primary["ip_address"]},
    )
    return metrics


def import_device_workbooks(*, source_root, database_path, project_code, dry_run=False):
    conn = create_db_connection(database_path)
    conn.row_factory = sqlite3.Row
    project = get_project_row(conn, project_code)
    if not project:
        raise RuntimeError(f"Project not found: {project_code}")

    station_rows = load_station_rows_for_match(conn)
    report = {
        "database": str(database_path),
        "source_root": str(source_root),
        "project_code": project_code,
        "dry_run": dry_run,
        "workbooks_seen": 0,
        "stations_matched": 0,
        "stations_skipped": 0,
        "cameras_updated": 0,
        "cameras_skipped": 0,
        "slot_channels_updated": 0,
        "recorders_added": 0,
        "recorders_updated": 0,
        "recorders_retired": 0,
        "matched_stations": [],
        "skipped_stations": [],
    }

    try:
        for workbook_path in iter_workbook_files(source_root):
            report["workbooks_seen"] += 1
            workbook_data = parse_station_excel(str(workbook_path))
            station_info = workbook_data["station"]
            if not workbook_name_matches_station(workbook_path, station_info["name"]):
                report["stations_skipped"] += 1
                report["skipped_stations"].append(
                    {
                        "workbook": workbook_path.name,
                        "station_name": station_info["name"],
                        "county": station_info.get("county", ""),
                        "reason": "workbook_station_name_mismatch",
                    }
                )
                continue
            matched_station = resolve_existing_station(
                station_rows,
                station_label=station_info["name"],
                voltage_level=station_info.get("voltage_level", ""),
                county=station_info.get("county", ""),
            )
            if not matched_station:
                report["stations_skipped"] += 1
                report["skipped_stations"].append(
                    {
                        "workbook": workbook_path.name,
                        "station_name": station_info["name"],
                        "county": station_info.get("county", ""),
                        "reason": "station_not_matched",
                    }
                )
                continue

            report["stations_matched"] += 1
            platform_rows = parse_platform_rows(load_rows(str(workbook_path)))
            primary_platform = choose_primary_platform(platform_rows)

            existing_cameras = conn.execute(
                """
                SELECT id, slot_id, project_camera_code, camera_index, location_desc,
                       ip_address, channel_number, recorder_name, recorder_ip_address, recorder_port
                FROM cameras
                WHERE station_id = ? AND status = 'active'
                ORDER BY id
                """,
                (matched_station["id"],),
            ).fetchall()
            station_uses_local_recorders = any(
                is_local_recorder_name(row["recorder_name"], matched_station["name"])
                for row in existing_cameras
            )
            if station_uses_local_recorders:
                update_station_primary_platform(
                    conn,
                    station_id=matched_station["id"],
                    primary_platform=primary_platform,
                )
            else:
                recorder_metrics = sync_platform_recorders(
                    conn,
                    station_id=matched_station["id"],
                    project_id=project["id"],
                    platforms=platform_rows,
                    workbook_name=workbook_path.name,
                )
                for key, value in recorder_metrics.items():
                    report[key] += value
            by_channel, by_location = build_camera_maps(existing_cameras)
            channel_offset = infer_channel_offset(existing_cameras, workbook_data["cameras"])
            used_ids = set()
            station_camera_updates = 0
            station_camera_skips = 0

            for camera_row in workbook_data["cameras"]:
                matched_camera = resolve_camera_match(
                    camera_row,
                    by_channel,
                    by_location,
                    used_ids,
                    channel_offset=channel_offset,
                )
                if not matched_camera:
                    station_camera_skips += 1
                    report["cameras_skipped"] += 1
                    continue
                used_ids.add(matched_camera["id"])
                update_camera_row(
                    conn,
                    existing_row=matched_camera,
                    workbook_camera=camera_row,
                    primary_platform=primary_platform,
                    station_name=matched_station["name"],
                )
                station_camera_updates += 1
                report["cameras_updated"] += 1
                if matched_camera["slot_id"]:
                    report["slot_channels_updated"] += 1

            report["matched_stations"].append(
                {
                    "workbook": workbook_path.name,
                    "station_name": matched_station["name"],
                    "station_id": matched_station["id"],
                    "camera_updates": station_camera_updates,
                    "camera_skips": station_camera_skips,
                    "platforms": {
                        label: {
                            "ip": platform_rows.get(label, {}).get("ip", ""),
                            "mask": platform_rows.get(label, {}).get("mask", ""),
                            "gateway": platform_rows.get(label, {}).get("gateway", ""),
                        }
                        for label in PLATFORM_LABELS
                    },
                }
            )

        if dry_run:
            conn.rollback()
        else:
            conn.commit()
        return report
    finally:
        conn.close()


def main():
    parser = argparse.ArgumentParser(description="Import camera and platform recorder IPs from workbooks")
    parser.add_argument("--source-root", required=True, help="Workbook root folder")
    parser.add_argument("--database", default=get_db_path(), help="SQLite database path")
    parser.add_argument("--project-code", default="unified", help="Project code")
    parser.add_argument("--report", help="Optional JSON report path")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    parser.add_argument("--skip-backup", action="store_true", help="Skip automatic backup")
    args = parser.parse_args()

    backup_path = None
    if not args.dry_run and not args.skip_backup:
        backup_path = backup_sqlite_database(args.database, label="import_device_workbook_ips")

    report = import_device_workbooks(
        source_root=args.source_root,
        database_path=args.database,
        project_code=args.project_code,
        dry_run=args.dry_run,
    )
    if backup_path:
        report["backup_path"] = str(backup_path)

    if args.report:
        report_path = Path(args.report)
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
