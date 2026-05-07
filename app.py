# app.py — Flask应用入口
import os
import math
import logging
import json
import re
import secrets
import hmac
from io import BytesIO
from datetime import datetime
from pathlib import Path
from functools import wraps
from flask import Flask, request, jsonify, g, current_app, session, redirect, render_template, send_file
from ai_fault_analysis import ensure_ai_runtime_schema, normalize_camera_hint
from config import Config, validate_runtime_config
from admin import (
    admin_bp,
    require_admin,
    upload_excel_scoped,
    delete_camera_scoped,
    add_camera_scoped,
    ensure_camera_recorder_metadata_columns,
)
from admin_fault_types import admin_fault_types_bp
from admin_notifications import admin_notifications_bp
from admin_projects import admin_projects_bp
from admin_review import admin_review_bp
from admin_user_access import admin_user_access_bp
from auth import auth_bp
from import_review_support import normalize_station_name
from notification_runtime import dispatch_notification_event
from photo_indexer import IMAGE_EXTENSIONS
from photo_thumbnails import build_thumbnail_payload, ensure_photo_thumbnail_columns
from project_access import (
    can_user_write_project,
    can_user_access_project,
    get_default_project_code,
    get_project_by_code,
    get_visible_projects,
    projects_enabled,
    table_exists,
)
from utils import get_db, close_db, init_app, validate_sql_identifier, validate_sql_type

app = Flask(__name__)
app.config.from_object(Config)
validate_runtime_config()
app.config['SECRET_KEY'] = Config.SECRET_KEY
app.config['TEMPLATES_AUTO_RELOAD'] = True
# CSRF 保护：双重提交 Cookie 模式
# session 中存储 csrf_token，前端 fetch 从 cookie 读入 X-CSRF-Token header
# 所有 POST/PUT/DELETE/PATCH 请求（除登录和静态资源外）均校验
app.config['WTF_CSRF_ENABLED'] = False

# 注册蓝图
DEFAULT_PENDING_FAULT_TYPE = '待现场确认'
FAULT_TYPE_MULTI_LABEL_SEPARATOR = ' | '
FAULT_TYPE_MULTI_CODE_SEPARATOR = ','
ROOT_CAUSE_LABELS = {
    'camera': '摄像头本体',
    'switch': '交换机及端口',
    'power': '供电（含交换机电源/集中电源/POE）',
    'network': '链路/网络',
    'platform': '平台/配置',
    'infrastructure_other': '其他基础设施',
    'unconfirmed': '未确认根因',
}
VALID_OWNER_TYPES = tuple(k for k in ROOT_CAUSE_LABELS.keys() if k != 'unconfirmed')
BASIC_SECURITY_HEADERS = {
    'X-Frame-Options': 'DENY',
    'X-Content-Type-Options': 'nosniff',
    'Referrer-Policy': 'strict-origin-when-cross-origin',
    'Permissions-Policy': 'camera=(), geolocation=(), microphone=()',
    'Content-Security-Policy': (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com https://cdn.jsdelivr.net https://unpkg.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://unpkg.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: https://*.is.autonavi.com https://*.tile.openstreetmap.org https://unpkg.com; "
        "connect-src 'self'; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'"
    ),
}


def normalize_fault_type_values(raw_value, *, is_code=False):
    if raw_value is None:
        return []

    if isinstance(raw_value, (list, tuple, set)):
        source_parts = list(raw_value)
    else:
        text = str(raw_value).strip()
        if not text:
            return []
        separator = FAULT_TYPE_MULTI_CODE_SEPARATOR if is_code else FAULT_TYPE_MULTI_LABEL_SEPARATOR
        if separator in text:
            source_parts = [part.strip() for part in text.split(separator)]
        elif is_code and ',' in text:
            source_parts = [part.strip() for part in text.split(',')]
        else:
            source_parts = [text]

    normalized = []
    seen = set()
    for part in source_parts:
        value = str(part or '').strip()
        if not value or value in seen:
            continue
        normalized.append(value)
        seen.add(value)
    return normalized


def join_fault_type_values(values, *, is_code=False):
    parts = normalize_fault_type_values(values, is_code=is_code)
    separator = FAULT_TYPE_MULTI_CODE_SEPARATOR if is_code else FAULT_TYPE_MULTI_LABEL_SEPARATOR
    return separator.join(parts)


def expand_fault_type_distribution(rows):
    aggregated = {}
    for row in rows:
        count = int(row['count'] or 0)
        if count <= 0:
            continue

        raw_group = str(row['semantic_group'] or '').strip()
        raw_label = str(row['fault_label'] or '').strip() or '未分类'
        groups = normalize_fault_type_values(raw_group, is_code=True) if raw_group else []
        labels = normalize_fault_type_values(raw_label) if raw_label else []

        if not groups and not labels:
            groups = ['UNCLASSIFIED']
            labels = ['未分类']

        if len(labels) > 1:
            if len(groups) == len(labels):
                pairs = list(zip(groups, labels))
            else:
                pairs = [(label, label) for label in labels]
        else:
            semantic_group = groups[0] if groups else raw_group or raw_label or 'UNCLASSIFIED'
            fault_label = labels[0] if labels else raw_label or '未分类'
            pairs = [(semantic_group, fault_label)]

        for semantic_group, fault_label in pairs:
            key = (semantic_group, fault_label)
            aggregated[key] = aggregated.get(key, 0) + count

    expanded = [
        {'semantic_group': semantic_group, 'fault_label': fault_label, 'count': count}
        for (semantic_group, fault_label), count in aggregated.items()
    ]
    expanded.sort(key=lambda item: (-item['count'], item['fault_label']))
    return expanded


CAMERA_REPLACEMENT_TYPE_KEYWORDS = ("摄像", "球机", "枪机", "半球", "云台")

app.register_blueprint(admin_bp)
app.register_blueprint(admin_fault_types_bp)
app.register_blueprint(admin_notifications_bp)
app.register_blueprint(admin_projects_bp)
app.register_blueprint(admin_review_bp)
app.register_blueprint(admin_user_access_bp)
app.register_blueprint(auth_bp)

app.view_functions['admin.upload_excel'] = require_admin(upload_excel_scoped)
app.view_functions['admin.delete_camera'] = require_admin(delete_camera_scoped)
app.view_functions['admin.add_camera'] = require_admin(add_camera_scoped)

# 注册utils的teardown
init_app(app)

# 全局错误处理
@app.errorhandler(404)
def not_found(e):
    if request.path.startswith("/api/"):
        return api_error("资源不存在", 404)
    return render_template("404.html") if False else ("页面不存在", 404)


@app.errorhandler(500)
def internal_error(e):
    logger.exception("未处理的异常: %s", e)
    if request.path.startswith("/api/"):
        return api_error("服务器内部错误", 500)
    return ("服务器内部错误", 500)


@app.errorhandler(Exception)
def handle_unhandled(e):
    logger.exception("未捕获的异常: %s", e)
    if request.path.startswith("/api/"):
        return api_error("服务器内部错误", 500)
    return ("服务器内部错误", 500)

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('station_monitor')


def _build_map_tile_providers():
    providers = []

    primary = {
        'name': 'primary',
        'url': Config.MAP_TILE_URL,
        'attribution': Config.MAP_TILE_ATTRIBUTION,
        'subdomains': Config.MAP_TILE_SUBDOMAINS,
        'maxZoom': Config.MAP_TILE_MAX_ZOOM,
    }
    if primary['url']:
        providers.append(primary)

    fallback_url = (Config.MAP_TILE_FALLBACK_URL or '').strip()
    if fallback_url and fallback_url != primary['url']:
        providers.append(
            {
                'name': 'fallback',
                'url': fallback_url,
                'attribution': Config.MAP_TILE_FALLBACK_ATTRIBUTION,
                'subdomains': Config.MAP_TILE_FALLBACK_SUBDOMAINS,
                'maxZoom': Config.MAP_TILE_FALLBACK_MAX_ZOOM,
            }
        )

    return providers


@app.context_processor
def inject_map_tile_config():
    return {
        'map_tile_providers': _build_map_tile_providers(),
    }


def _safe_dispatch_fault_notification(db, fault_id: int, event_type: str):
    try:
        return dispatch_notification_event(db, fault_id, event_type, logger=logger)
    except Exception:
        logger.exception(
            "Failed to evaluate notification event: fault=%s event=%s",
            fault_id,
            event_type,
        )
        return None


def _sync_closed_fault_to_worklog(fault_id: int):
    """闭环后自动写入工作记录.xlsx，失败不影响闭环本身。"""
    try:
        from worklog_sync import sync_fault_to_worklog
        from utils import create_db_connection
        db = create_db_connection(
            Config.DATABASE_PATH,
            database_url=Config.DATABASE_URL,
            row_factory=True,
            enable_wal=True,
        )
        sync_fault_to_worklog(db, fault_id)
        db.close()
    except Exception:
        logger.exception("worklog_sync: failed for fault=%s", fault_id)


# ============================================================
# Token认证中间件（决策#1, #2）
# ============================================================

def require_token(f):
    """Token认证装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        # 从请求头获取token
        auth_header = request.headers.get('Authorization', '')

        if not auth_header.startswith('Bearer '):
            return jsonify({'error': '未提供认证令牌'}), 401

        token = auth_header[7:]  # 去掉 "Bearer " 前缀

        if not token:
            return jsonify({'error': '令牌无效'}), 401

        # 使用常量时间比较防止时序攻击
        if not Config.API_TOKEN or not hmac.compare_digest(token, Config.API_TOKEN):
            return jsonify({'error': '令牌无效或已过期'}), 403

        return f(*args, **kwargs)
    return decorated

# ============================================================
# 通用API响应格式（决策#5）
# ============================================================

def api_error(message, status_code=400):
    """统一错误响应格式"""
    return jsonify({'error': message}), status_code

def api_success(data, status_code=200):
    """统一成功响应格式"""
    return jsonify(data), status_code


PUBLIC_GUEST_ENDPOINTS = {
    'index',
    'statistics',
    'login_page',
}


@app.before_request
def ensure_authenticated_session_is_permanent():
    if session.get('user_id') and not session.permanent:
        session.permanent = True


@app.before_request
def restrict_guest_access():
    """未登录仅开放首页与统计报表，其余页面和数据接口需先登录。"""
    if session.get('user_id'):
        return None

    endpoint = request.endpoint or ''
    if endpoint == 'static' or endpoint.startswith('auth.'):
        return None

    if endpoint in PUBLIC_GUEST_ENDPOINTS:
        return None

    if request.path.startswith('/api/'):
        return api_error('请先登录', 401)

    return redirect('/login')


@app.before_request
def ensure_csrf_token():
    """确保已登录用户的 session 中存在 CSRF token"""
    if session.get('user_id') and 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)


# CSRF 双重提交 Cookie 保护
CSRF_SAFE_METHODS = {'GET', 'HEAD', 'OPTIONS'}


@app.before_request
def enforce_csrf():
    """对状态变更请求校验 CSRF token（双重提交 Cookie 模式）"""
    if request.method in CSRF_SAFE_METHODS:
        return None
    # 静态资源和认证登录页豁免
    if request.endpoint in ('static', 'auth.login'):
        return None
    # 健康检查豁免
    if request.endpoint == 'health':
        return None

    session_token = session.get('csrf_token')
    header_token = request.headers.get('X-CSRF-Token')

    if not session_token or not header_token:
        return api_error('缺少 CSRF token', 403)

    if not hmac.compare_digest(session_token, header_token):
        return api_error('CSRF token 不匹配', 403)

    return None


@app.after_request
def apply_basic_security_headers(response):
    for header_name, header_value in BASIC_SECURITY_HEADERS.items():
        response.headers.setdefault(header_name, header_value)
    return response


def normalize_photo_row(row):
    photo = dict(row)
    thumbnail_blob = photo.pop('thumbnail_data', None)
    photo['has_thumbnail'] = bool(thumbnail_blob)
    for key, value in list(photo.items()):
        if isinstance(value, memoryview):
            photo[key] = value.tobytes().decode('utf-8', errors='replace')
        elif isinstance(value, bytes):
            photo[key] = value.decode('utf-8', errors='replace')
    photo['is_image'] = (photo.get('ext', '').lower() in IMAGE_EXTENSIONS)
    return photo


def _photo_thumbnail_fields(photo_columns):
    available = set(photo_columns or [])
    fields = []
    for field_name in (
        "thumbnail_data",
        "thumbnail_content_type",
        "thumbnail_width",
        "thumbnail_height",
        "thumbnail_source_mtime",
        "thumbnail_generated_at",
    ):
        if field_name in available:
            fields.append(field_name)
    return fields


def _fetch_photo_asset_row(db, photo_id, *, include_project=False):
    photo_columns = ensure_photo_thumbnail_columns(db)
    select_fields = ["id", "abs_path", "ext", "file_mtime"]
    if include_project and "project_id" in photo_columns:
        select_fields.append("project_id")
    select_fields.extend(_photo_thumbnail_fields(photo_columns))
    row = db.execute(
        f"SELECT {', '.join(select_fields)} FROM photos WHERE id = ?",
        (photo_id,),
    ).fetchone()
    return row, photo_columns


def _persist_thumbnail_for_photo_id(db, photo_id, payload, file_mtime):
    db.execute(
        """
        UPDATE photos
        SET thumbnail_data = ?,
            thumbnail_content_type = ?,
            thumbnail_width = ?,
            thumbnail_height = ?,
            thumbnail_source_mtime = ?,
            thumbnail_generated_at = ?
        WHERE id = ?
        """,
        (
            payload["thumbnail_data"],
            payload["thumbnail_content_type"],
            payload["thumbnail_width"],
            payload["thumbnail_height"],
            file_mtime,
            payload["thumbnail_generated_at"],
            photo_id,
        ),
    )
    db.commit()


def _send_photo_thumbnail(db, row, photo_columns):
    thumbnail_data = row["thumbnail_data"] if "thumbnail_data" in photo_columns else None
    thumbnail_type = row["thumbnail_content_type"] if "thumbnail_content_type" in photo_columns else None
    thumbnail_source_mtime = row["thumbnail_source_mtime"] if "thumbnail_source_mtime" in photo_columns else None
    file_mtime = row["file_mtime"] if "file_mtime" in row.keys() else None

    if thumbnail_data and thumbnail_type and thumbnail_source_mtime == file_mtime:
        return send_file(BytesIO(thumbnail_data), mimetype=thumbnail_type, conditional=True)

    root = get_photo_root()
    file_path = Path(row['abs_path']).resolve()
    if not is_path_under_root(file_path, root):
        logger.warning(f"Blocked photo traversal attempt: photo_id={row['id']}, path={file_path}")
        return api_error('非法路径访问', 403)

    if file_path.exists() and file_path.is_file():
        payload = build_thumbnail_payload(file_path)
        if payload:
            _persist_thumbnail_for_photo_id(db, row['id'], payload, file_mtime)
            return send_file(
                BytesIO(payload["thumbnail_data"]),
                mimetype=payload["thumbnail_content_type"],
                conditional=True,
            )
        return send_file(str(file_path), conditional=True)

    if thumbnail_data and thumbnail_type:
        return send_file(BytesIO(thumbnail_data), mimetype=thumbnail_type, conditional=True)

    return api_error('照片文件不存在', 404)


def get_photo_root():
    return Path(Config.PHOTO_ROOT_PATH).resolve()


def is_path_under_root(file_path, root_path):
    try:
        file_path.resolve().relative_to(root_path.resolve())
        return True
    except ValueError:
        return False


def get_table_columns(db, table_name):
    """返回表字段集合（委托给 utils.py 统一实现，带安全校验）。"""
    from utils import get_table_columns as _get_table_columns
    return _get_table_columns(db, table_name)


def ensure_fault_report_multi_camera_schema(db):
    if not table_exists(db, "fault_reports"):
        return set()
    columns = get_table_columns(db, "fault_reports")
    missing = False
    if "fault_group_key" not in columns:
        db.execute("ALTER TABLE fault_reports ADD COLUMN fault_group_key TEXT")
        db.execute("CREATE INDEX IF NOT EXISTS idx_fault_group_key ON fault_reports(fault_group_key)")
        missing = True
    root_cause_columns = {
        "fault_owner_type": "TEXT",
        "is_batch_impact": "INTEGER",
        "root_cause_type": "TEXT",
        "impact_camera_count": "INTEGER",
        "fault_owner_confirmed_by": "INTEGER",
        "fault_owner_confirmed_at": "TEXT",
    }
    for col_name, col_sql in root_cause_columns.items():
        if col_name not in columns:
            validate_sql_identifier(col_name, kind="column")
            validate_sql_type(col_sql)
            db.execute(f"ALTER TABLE fault_reports ADD COLUMN {col_name} {col_sql}")
            missing = True
    if missing:
        db.commit()
        columns = get_table_columns(db, "fault_reports")

    # 复合索引：project_id + created_at（仅当列存在时创建）
    if 'project_id' in columns:
        db.execute("CREATE INDEX IF NOT EXISTS idx_fault_project_time ON fault_reports(project_id, created_at)")

    if 'camera_slot_id' in columns and 'impact_camera_count' in columns:
        needs_fix = db.execute(
            "SELECT COUNT(*) FROM fault_reports WHERE impact_camera_count > 1 AND camera_slot_id IS NOT NULL"
        ).fetchone()[0]
        if needs_fix > 0:
            db.execute(
                "UPDATE fault_reports SET camera_slot_id = NULL WHERE impact_camera_count > 1 AND camera_slot_id IS NOT NULL"
            )
            db.commit()

    return columns


def ensure_fault_report_camera_detail_schema(db):
    if table_exists(db, "fault_report_cameras"):
        columns = get_table_columns(db, "fault_report_cameras")
        missing = False
        required_columns = {
            "camera_slot_id": "INTEGER",
            "project_id": "INTEGER",
            "project_device_code": "TEXT",
            "camera_label": "TEXT",
            "recovery_state": "TEXT DEFAULT 'pending'",
            "affects_statistics": "INTEGER DEFAULT 1",
            "detail_fault_reason": "TEXT",
            "detail_resolution": "TEXT",
            "detail_note": "TEXT",
            "created_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
            "updated_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        }
        for col_name, col_sql in required_columns.items():
            if col_name not in columns:
                validate_sql_identifier(col_name, kind="column")
                validate_sql_type(col_sql)
                db.execute(f"ALTER TABLE fault_report_cameras ADD COLUMN {col_name} {col_sql}")
                missing = True
        db.execute("CREATE INDEX IF NOT EXISTS idx_fault_report_cameras_fault ON fault_report_cameras(fault_report_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_fault_report_cameras_camera ON fault_report_cameras(camera_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_fault_report_cameras_recovery ON fault_report_cameras(recovery_state)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_fault_report_cameras_affects_stats ON fault_report_cameras(affects_statistics)")
        if missing:
            db.commit()
            columns = get_table_columns(db, "fault_report_cameras")
        return columns

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS fault_report_cameras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fault_report_id INTEGER NOT NULL,
            camera_id INTEGER NOT NULL,
            camera_slot_id INTEGER,
            project_id INTEGER,
            project_device_code TEXT,
            camera_label TEXT,
            recovery_state TEXT DEFAULT 'pending',
            affects_statistics INTEGER DEFAULT 1,
            detail_fault_reason TEXT,
            detail_resolution TEXT,
            detail_note TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(fault_report_id, camera_id),
            FOREIGN KEY (fault_report_id) REFERENCES fault_reports(id) ON DELETE CASCADE,
            FOREIGN KEY (camera_id) REFERENCES cameras(id)
        )
        """
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_fault_report_cameras_fault ON fault_report_cameras(fault_report_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_fault_report_cameras_camera ON fault_report_cameras(camera_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_fault_report_cameras_recovery ON fault_report_cameras(recovery_state)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_fault_report_cameras_affects_stats ON fault_report_cameras(affects_statistics)")
    db.commit()
    return get_table_columns(db, "fault_report_cameras")


def build_fault_camera_label(camera_row):
    if not camera_row:
        return ""
    for field_name in ("location_desc", "area", "camera_index", "project_camera_code"):
        if field_name in camera_row.keys():
            value = str(camera_row[field_name] or "").strip()
            if value:
                return value
    return str(camera_row["id"])


def fetch_fault_camera_details(db, fault_report_id):
    if not table_exists(db, "fault_report_cameras"):
        return []
    detail_columns = get_table_columns(db, "fault_report_cameras")
    select_fields = [
        "d.id",
        "d.fault_report_id",
        "d.camera_id",
        "d.recovery_state",
    ]
    optional_fields = [
        "camera_slot_id",
        "project_id",
        "project_device_code",
        "camera_label",
        "affects_statistics",
        "detail_fault_reason",
        "detail_resolution",
        "detail_note",
    ]
    for field_name in optional_fields:
        if field_name in detail_columns:
            select_fields.append(f"d.{field_name}")
    select_fields.extend([
        "c.camera_index AS camera_index",
        "c.location_desc AS camera_location_desc",
        "c.area AS camera_area",
    ])
    rows = db.execute(
        f"""
        SELECT {', '.join(select_fields)}
        FROM fault_report_cameras d
        LEFT JOIN cameras c ON d.camera_id = c.id
        WHERE d.fault_report_id = ?
        ORDER BY d.id ASC
        """,
        (fault_report_id,),
    ).fetchall()
    return [dict(row) for row in rows]


def fetch_fault_camera_details_map(db, fault_report_ids):
    if not fault_report_ids or not table_exists(db, "fault_report_cameras"):
        return {}
    detail_columns = get_table_columns(db, "fault_report_cameras")
    select_fields = [
        "d.id",
        "d.fault_report_id",
        "d.camera_id",
        "d.recovery_state",
    ]
    optional_fields = [
        "camera_slot_id",
        "project_id",
        "project_device_code",
        "camera_label",
        "affects_statistics",
        "detail_fault_reason",
        "detail_resolution",
        "detail_note",
    ]
    for field_name in optional_fields:
        if field_name in detail_columns:
            select_fields.append(f"d.{field_name}")
    select_fields.extend([
        "c.camera_index AS camera_index",
        "c.location_desc AS camera_location_desc",
        "c.area AS camera_area",
    ])
    placeholders = ", ".join(["?"] * len(fault_report_ids))
    rows = db.execute(
        f"""
        SELECT {', '.join(select_fields)}
        FROM fault_report_cameras d
        LEFT JOIN cameras c ON d.camera_id = c.id
        WHERE d.fault_report_id IN ({placeholders})
        ORDER BY d.fault_report_id ASC, d.id ASC
        """,
        list(fault_report_ids),
    ).fetchall()
    detail_map = {int(fault_id): [] for fault_id in fault_report_ids}
    for row in rows:
        item = dict(row)
        detail_map.setdefault(int(item['fault_report_id']), []).append(item)
    return detail_map


CAMERA_RECOVERY_STATE_LABELS = {
    'pending': '待确认',
    'resolved': '已修复',
    'self_recovered': '自恢复',
}


def _fault_camera_detail_label(detail):
    for key in ('camera_location_desc', 'camera_label', 'project_device_code', 'camera_index', 'camera_area'):
        value = str(detail.get(key) or '').strip()
        if value:
            return value
    camera_id = detail.get('camera_id')
    return f"摄像头#{camera_id}" if camera_id not in (None, '') else '未命名摄像头'



def summarize_fault_camera_details(camera_details):
    normalized_details = [item for item in (camera_details or []) if isinstance(item, dict)]
    if not normalized_details:
        return {
            'camera_labels': [],
            'camera_locations_text': '',
            'camera_recovery_text': '',
            'resolved_count': 0,
            'self_recovered_count': 0,
        }

    labels = []
    seen_labels = set()
    recovery_segments = []
    resolved_count = 0
    self_recovered_count = 0
    for detail in normalized_details:
        label = _fault_camera_detail_label(detail)
        if label not in seen_labels:
            labels.append(label)
            seen_labels.add(label)
        recovery_state = str(detail.get('recovery_state') or '').strip() or 'pending'
        if recovery_state == 'resolved':
            resolved_count += 1
        elif recovery_state == 'self_recovered':
            self_recovered_count += 1
        recovery_segments.append(f"{label}（{CAMERA_RECOVERY_STATE_LABELS.get(recovery_state, recovery_state)}）")

    return {
        'camera_labels': labels,
        'camera_locations_text': '、'.join(labels),
        'camera_recovery_text': '；'.join(recovery_segments),
        'resolved_count': resolved_count,
        'self_recovered_count': self_recovered_count,
    }



def attach_fault_camera_detail_summary(payload):
    if not isinstance(payload, dict):
        return payload
    summary = summarize_fault_camera_details(payload.get('camera_details') or [])
    payload['camera_locations_text'] = summary['camera_locations_text']
    payload['camera_recovery_text'] = summary['camera_recovery_text']
    payload['resolved_camera_count'] = summary['resolved_count']
    payload['self_recovered_camera_count'] = summary['self_recovered_count']
    impact_camera_count = payload.get('impact_camera_count')
    if summary['camera_locations_text'] and impact_camera_count not in (None, '', 0, 1, '0', '1'):
        payload['camera_display_text'] = summary['camera_locations_text']
    else:
        payload['camera_display_text'] = str(
            payload.get('camera_location')
            or payload.get('camera_location_text')
            or payload.get('camera_area')
            or summary['camera_locations_text']
            or ''
        ).strip()
    return payload



VALID_CAMERA_RECOVERY_STATES = ('pending', 'resolved', 'self_recovered')


def normalize_fault_camera_detail_updates(raw_items):
    if raw_items in (None, ""):
        return []
    if not isinstance(raw_items, list):
        raise ValueError('camera_details_invalid')

    normalized = []
    seen_camera_ids = set()
    for raw_item in raw_items:
        if not isinstance(raw_item, dict):
            raise ValueError('camera_details_invalid')
        try:
            camera_id = int(raw_item.get('camera_id'))
        except (TypeError, ValueError):
            raise ValueError('camera_detail_camera_id_invalid')
        if camera_id in seen_camera_ids:
            raise ValueError('camera_detail_camera_id_duplicated')
        seen_camera_ids.add(camera_id)

        recovery_state = str(raw_item.get('recovery_state') or '').strip() or 'resolved'
        if recovery_state not in VALID_CAMERA_RECOVERY_STATES:
            raise ValueError('camera_detail_recovery_state_invalid')

        normalized.append({
            'camera_id': camera_id,
            'recovery_state': recovery_state,
            'detail_fault_reason': str(raw_item.get('detail_fault_reason') or '').strip() or None,
            'detail_resolution': str(raw_item.get('detail_resolution') or '').strip() or None,
            'detail_note': str(raw_item.get('detail_note') or '').strip() or None,
        })
    return normalized


def apply_fault_camera_detail_closure(db, fault_id, *, fault_owner_type=None, batch_impact_value=None, handler_note=None, camera_detail_updates=None):
    if not table_exists(db, "fault_report_cameras"):
        return

    detail_columns = ensure_fault_report_camera_detail_schema(db)
    existing_rows = db.execute(
        "SELECT id, camera_id FROM fault_report_cameras WHERE fault_report_id = ? ORDER BY id ASC",
        (fault_id,),
    ).fetchall()
    if not existing_rows:
        return

    existing_camera_ids = [int(row['camera_id']) for row in existing_rows]
    detail_map = {int(row['camera_id']): int(row['id']) for row in existing_rows}
    normalized_updates = normalize_fault_camera_detail_updates(camera_detail_updates)

    if normalized_updates:
        for item in normalized_updates:
            if item['recovery_state'] == 'self_recovered' and not item.get('detail_fault_reason'):
                raise ValueError('self_recovered_requires_reason')

    if normalized_updates:
        provided_camera_ids = [item['camera_id'] for item in normalized_updates]
        if sorted(provided_camera_ids) != sorted(existing_camera_ids):
            raise ValueError('camera_detail_camera_scope_mismatch')
    elif fault_owner_type == 'camera':
        raise ValueError('camera_details_required_for_camera_owner')
    else:
        fallback_state = 'resolved'
        if batch_impact_value == 0 and len(existing_camera_ids) == 1:
            fallback_state = 'resolved'
        owner_label = fault_owner_type if fault_owner_type else '未分类'
        batch_label = '共因' if batch_impact_value == 1 else '单因'
        detail_resolution = handler_note or f'{batch_label}故障，归属: {owner_label}'
        normalized_updates = [
            {
                'camera_id': camera_id,
                'recovery_state': fallback_state,
                'detail_fault_reason': None,
                'detail_resolution': detail_resolution,
                'detail_note': None,
            }
            for camera_id in existing_camera_ids
        ]

    updatable_optional_fields = [
        field_name for field_name in ('detail_fault_reason', 'detail_resolution', 'detail_note')
        if field_name in detail_columns
    ]
    has_affects_stats = 'affects_statistics' in detail_columns
    update_fields = ["recovery_state = ?", "updated_at = CURRENT_TIMESTAMP"]
    if has_affects_stats:
        update_fields.append("affects_statistics = ?")
    update_fields.extend(f"{field_name} = ?" for field_name in updatable_optional_fields)
    update_sql = f"UPDATE fault_report_cameras SET {', '.join(update_fields)} WHERE id = ?"

    for item in normalized_updates:
        row_id = detail_map.get(item['camera_id'])
        if row_id is None:
            raise ValueError('camera_detail_camera_scope_mismatch')
        update_params = [item['recovery_state']]
        if has_affects_stats:
            update_params.append(0 if item['recovery_state'] == 'self_recovered' else 1)
        for field_name in updatable_optional_fields:
            update_params.append(item.get(field_name))
        update_params.append(row_id)
        db.execute(update_sql, update_params)

    pending_check = db.execute(
        "SELECT COUNT(*) FROM fault_report_cameras WHERE fault_report_id = ? AND recovery_state = 'pending'",
        (fault_id,),
    ).fetchone()[0]
    if pending_check > 0:
        raise ValueError('camera_detail_incomplete_closure')





def normalize_camera_ids(value):
    if value in (None, ""):
        return []
    raw_items = value if isinstance(value, list) else [value]
    normalized = []
    seen = set()
    for item in raw_items:
        if item in (None, ""):
            continue
        try:
            camera_id = int(item)
        except (TypeError, ValueError):
            raise ValueError("camera_id_invalid")
        if camera_id in seen:
            continue
        seen.add(camera_id)
        normalized.append(camera_id)
    return normalized


def fetch_camera_rows_by_ids(db, camera_ids, camera_columns):
    if not camera_ids:
        return []
    select_fields = ["id", "station_id"]
    for field_name in ("project_id", "slot_id", "project_camera_code", "camera_index", "location_desc", "area"):
        if field_name in camera_columns:
            select_fields.append(field_name)
    placeholders = ", ".join(["?"] * len(camera_ids))
    rows = db.execute(
        f"SELECT {', '.join(select_fields)} FROM cameras WHERE id IN ({placeholders})",
        camera_ids,
    ).fetchall()
    row_map = {int(row["id"]): row for row in rows}
    ordered_rows = []
    for camera_id in camera_ids:
        row = row_map.get(camera_id)
        if not row:
            return None
        ordered_rows.append(row)
    return ordered_rows


def resolve_fault_type_update_payload(db, fault_row, fault_report_columns, fault_type_value, fault_type_code):
    resolved_types = normalize_fault_type_values(fault_type_value)
    resolved_codes = normalize_fault_type_values(fault_type_code, is_code=True)
    resolved_type = join_fault_type_values(resolved_types) or None
    resolved_code = join_fault_type_values(resolved_codes, is_code=True) or None
    resolved_version_id = None
    project_id = fault_row['project_id'] if 'project_id' in fault_report_columns and 'project_id' in fault_row.keys() else None

    if resolved_codes and project_id:
        project_row = db.execute(
            "SELECT fault_type_version_id FROM projects WHERE id = ?",
            (project_id,),
        ).fetchone()
        if project_row and project_row['fault_type_version_id']:
            fault_type_rows = db.execute(
                """
                SELECT type_code, type_label
                FROM project_fault_types
                WHERE version_id = ?
                  AND is_active = 1
                """,
                (project_row['fault_type_version_id'],),
            ).fetchall()
            fault_type_map = {row['type_code']: row['type_label'] for row in fault_type_rows}
            fault_type_label_map = {row['type_label']: row['type_code'] for row in fault_type_rows}
            canonical_codes = []
            for raw_code in resolved_codes:
                if raw_code in fault_type_map:
                    canonical_codes.append(raw_code)
                    continue
                matched_code = fault_type_label_map.get(raw_code)
                if matched_code:
                    canonical_codes.append(matched_code)
                    continue
                raise ValueError('故障类型不存在或未发布')
            resolved_types = [fault_type_map[code] for code in canonical_codes]
            resolved_type = join_fault_type_values(resolved_types)
            resolved_code = join_fault_type_values(canonical_codes, is_code=True)
            resolved_version_id = project_row['fault_type_version_id']
    elif resolved_types and project_id:
        project_row = db.execute(
            "SELECT fault_type_version_id FROM projects WHERE id = ?",
            (project_id,),
        ).fetchone()
        if project_row and project_row['fault_type_version_id']:
            fault_type_rows = db.execute(
                """
                SELECT type_code, type_label
                FROM project_fault_types
                WHERE version_id = ?
                  AND is_active = 1
                """,
                (project_row['fault_type_version_id'],),
            ).fetchall()
            label_to_code_map = {row['type_label']: row['type_code'] for row in fault_type_rows}
            canonical_codes = [label_to_code_map.get(label) for label in resolved_types]
            if canonical_codes and all(canonical_codes):
                resolved_code = join_fault_type_values(canonical_codes, is_code=True)
                resolved_version_id = project_row['fault_type_version_id']

    if not resolved_type:
        resolved_code = None
        resolved_version_id = None

    return {
        'fault_type': resolved_type,
        'fault_type_code': resolved_code,
        'fault_type_version_id': resolved_version_id,
    }


def ensure_fault_report_soft_delete_schema(db):
    if not table_exists(db, "fault_reports"):
        return set()
    columns = get_table_columns(db, "fault_reports")
    required_columns = {
        "deleted_at": "TIMESTAMP",
        "deleted_by": "INTEGER",
        "planned_handle_time": "TIMESTAMP",
    }
    missing = False
    for column_name, column_sql in required_columns.items():
        if column_name in columns:
            continue
        validate_sql_identifier(column_name, kind="column")
        validate_sql_type(column_sql)
        db.execute(f"ALTER TABLE fault_reports ADD COLUMN {column_name} {column_sql}")
        missing = True
    if missing:
        db.commit()
        columns = get_table_columns(db, "fault_reports")
    # 复合索引：deleted_at（用于软删除筛选）
    if 'deleted_at' in columns:
        db.execute("CREATE INDEX IF NOT EXISTS idx_fault_deleted_at ON fault_reports(deleted_at)")
    return columns


def build_fault_deleted_clause(fault_report_columns, *, alias="f", mode="active"):
    if "deleted_at" not in fault_report_columns:
        return " AND 1=0" if mode == "only" else ""
    column_name = f"{alias}.deleted_at" if alias else "deleted_at"
    if mode == "only":
        return f" AND {column_name} IS NOT NULL"
    if mode == "all":
        return ""
    return f" AND {column_name} IS NULL"


def get_fault_deleted_mode():
    raw_mode = str(request.args.get("deleted") or "").strip().lower()
    if raw_mode in {"only", "all"} and session.get("role") == "admin":
        return raw_mode
    return "active"


def _looks_like_fault_camera_location(value):
    text = str(value or "").strip()
    if not text or len(text) > 80:
        return False
    strong_markers = ("摄像机", "摄像头", "球机", "枪机", "半球", "云台")
    if any(marker in text for marker in strong_markers):
        return True
    has_channel_marker = "#" in text or bool(re.search(r"通道\s*\d+", text))
    location_markers = ("主变", "场地", "室", "门", "围墙", "东", "南", "西", "北")
    return has_channel_marker and any(marker in text for marker in location_markers)


def derive_fault_camera_location(description):
    text = str(description or "").strip()
    if not text:
        return ""
    text = text.splitlines()[0].strip()
    text = re.sub(r"\s*\|\s*地点[:：].*$", "", text).strip()
    text = re.sub(r"\s*\|\s*区县[:：].*$", "", text).strip()
    text = re.sub(r"\s+", " ", text)

    direct_match = re.search(
        r"(?P<location>[\u4e00-\u9fffA-Za-z0-9#\-/_.（）()]+?(?:摄像机|摄像头|球机|枪机|半球|云台|通道\s*\d+))",
        text,
    )
    if direct_match:
        candidate = direct_match.group("location").strip(" |,，;；:：")
        if _looks_like_fault_camera_location(candidate):
            return candidate

    normalized = normalize_camera_hint(text)
    normalized = re.sub(r"(排查|维修|检修|恢复|更换|重启|处理|排障).*$", "", normalized).strip(" |,，;；:：")
    if _looks_like_fault_camera_location(normalized):
        return normalized
    return ""


def enrich_fault_camera_location(payload):
    camera_location = str(payload.get("camera_location") or "").strip()
    if camera_location:
        return payload
    fallback = derive_fault_camera_location(
        payload.get("camera_location_text") or payload.get("description")
    )
    if fallback:
        payload["camera_location"] = fallback
        if not str(payload.get("camera_location_text") or "").strip():
            payload["camera_location_text"] = fallback
    return payload


def project_access_denied():
    return jsonify({'error': '无权访问该项目', 'code': 'PROJECT_ACCESS_DENIED'}), 403


def get_current_user_project_context(db):
    user_id = session.get('user_id')
    role = session.get('role')
    visible_projects = get_visible_projects(
        db,
        user_id=user_id,
        role=role,
        include_inactive=False,
    )
    return {
        'user_id': user_id,
        'role': role,
        'projects': visible_projects,
        'default_project_code': get_default_project_code(visible_projects),
    }


def ensure_project_read_access(db, project_code: str):
    user_id = session.get('user_id')
    role = session.get('role')
    if role is None:
        return True
    return can_user_access_project(
        db,
        user_id=user_id,
        role=role,
        project_code=project_code,
    )


def ensure_project_write_access(db, project_code: str):
    user_id = session.get('user_id')
    role = session.get('role')
    if role is None:
        return True
    return can_user_write_project(
        db,
        user_id=user_id,
        role=role,
        project_code=project_code,
    )


def build_project_scope(db, requested_project_code: str | None = None):
    requested_project_code = (requested_project_code or '').strip()
    if not projects_enabled(db):
        return {
            'enabled': False,
            'project_ids': None,
            'requested_project': None,
            'projects': [],
        }, None

    role = session.get('role')
    user_id = session.get('user_id')
    include_inactive = bool(requested_project_code and requested_project_code != 'all')

    if role is None:
        visible_projects = get_visible_projects(
            db,
            user_id=None,
            role='admin',
            include_inactive=include_inactive,
        )
        for project in visible_projects:
            project['can_write'] = False
    else:
        visible_projects = get_visible_projects(
            db,
            user_id=user_id,
            role=role,
            include_inactive=include_inactive,
        )

    if requested_project_code and requested_project_code != 'all':
        requested_project = get_project_by_code(
            db,
            requested_project_code,
            include_inactive=True,
        )
        if not requested_project:
            return None, api_error('项目不存在', 404)
        if role is None:
            if not requested_project.get('is_active'):
                return None, api_error('项目不存在', 404)
        elif not any(project['code'] == requested_project_code for project in visible_projects):
            return None, project_access_denied()
        return {
            'enabled': True,
            'project_ids': [requested_project['id']],
            'requested_project': requested_project,
            'projects': visible_projects,
        }, None

    active_projects = [project for project in visible_projects if project.get('is_active')]
    return {
        'enabled': True,
        'project_ids': [project['id'] for project in active_projects],
        'requested_project': None,
        'projects': active_projects,
    }, None


def extend_params(params, extra):
    if extra:
        params.extend(extra)
    return params


FAULT_STATUS_SORT_SQL = """
CASE COALESCE(f.status, 'open')
    WHEN 'open' THEN 0
    WHEN 'handling' THEN 1
    WHEN 'closed' THEN 2
    ELSE 3
END
"""

OVERDUE_FAULT_THRESHOLD_DAYS = 7

RESPONSE_BUCKET_DEFINITIONS = [
    {'label': '2小时内', 'min_seconds': 0, 'max_seconds': 2 * 3600},
    {'label': '2-8小时', 'min_seconds': 2 * 3600, 'max_seconds': 8 * 3600},
    {'label': '8-24小时', 'min_seconds': 8 * 3600, 'max_seconds': 24 * 3600},
    {'label': '24小时以上', 'min_seconds': 24 * 3600, 'max_seconds': None},
]

CLOSE_BUCKET_DEFINITIONS = [
    {'label': '当天闭环', 'min_seconds': 0, 'max_seconds': 24 * 3600},
    {'label': '1-3天', 'min_seconds': 24 * 3600, 'max_seconds': 3 * 24 * 3600},
    {'label': '3-7天', 'min_seconds': 3 * 24 * 3600, 'max_seconds': 7 * 24 * 3600},
    {'label': '7天以上', 'min_seconds': 7 * 24 * 3600, 'max_seconds': None},
]


def build_project_in_clause(alias: str, project_ids: list[int] | None):
    if project_ids is None:
        return "", []
    if not project_ids:
        return " AND 1 = 0", []
    placeholders = ", ".join(["?"] * len(project_ids))
    return f" AND {alias}.project_id IN ({placeholders})", list(project_ids)


def _column_expr(columns, alias: str, name: str, fallback: str = "NULL"):
    if name in columns:
        return f"{alias}.{name}"
    return fallback


def _count_with_alias(db, table_name: str, alias: str, project_ids: list[int] | None, active_only: bool = False):
    columns = get_table_columns(db, table_name)
    query = f"SELECT COUNT(*) as count FROM {table_name} {alias} WHERE 1=1"
    params = []
    if 'project_id' in columns:
        project_sql, project_params = build_project_in_clause(alias, project_ids)
        query += project_sql
        params.extend(project_params)
    if active_only and 'status' in columns:
        query += f" AND {alias}.status = 'active'"
    return db.execute(query, params).fetchone()['count']


def _bucketize_duration_rows(rows, definitions):
    bucket_counts = []
    for definition in definitions:
        bucket_counts.append({
            'label': definition['label'],
            'count': 0,
        })

    for row in rows or []:
        try:
            duration_seconds = float(row['duration_seconds'])
        except (TypeError, ValueError, KeyError):
            continue
        if duration_seconds < 0:
            continue

        for index, definition in enumerate(definitions):
            min_seconds = definition['min_seconds']
            max_seconds = definition['max_seconds']
            if duration_seconds < min_seconds:
                continue
            if max_seconds is not None and duration_seconds >= max_seconds:
                continue
            bucket_counts[index]['count'] += 1
            break

    return bucket_counts


def _build_statistics_payload(db, year: int | None, requested_project_code: str | None):
    project_scope, error = build_project_scope(db, requested_project_code)
    if error:
        return None, error

    fault_report_columns = get_table_columns(db, "fault_reports")
    camera_columns = get_table_columns(db, "cameras")
    project_fault_type_columns = get_table_columns(db, "project_fault_types") if table_exists(db, "project_fault_types") else set()
    has_project_fault_types = {"version_id", "type_code", "semantic_group", "type_label"}.issubset(project_fault_type_columns)

    fault_scope_sql = ""
    fault_scope_params = []
    if project_scope['enabled'] and 'project_id' in fault_report_columns:
        fault_scope_sql, fault_scope_params = build_project_in_clause("f", project_scope['project_ids'])

    camera_scope_sql = ""
    camera_scope_params = []
    if project_scope['enabled'] and 'project_id' in camera_columns:
        camera_scope_sql, camera_scope_params = build_project_in_clause("c", project_scope['project_ids'])

    station_count = db.execute("SELECT COUNT(*) as count FROM stations").fetchone()['count']
    if project_scope['enabled'] and (('project_id' in fault_report_columns) or ('project_id' in camera_columns)):
        station_sources = []
        station_params = []
        if 'project_id' in camera_columns:
            active_sql = " AND c.status = 'active'" if 'status' in camera_columns else ""
            station_sources.append(
                f"SELECT c.station_id FROM cameras c WHERE 1=1{camera_scope_sql}{active_sql}"
            )
            station_params.extend(camera_scope_params)
        if 'project_id' in fault_report_columns:
            station_sources.append(
                f"SELECT f.station_id FROM fault_reports f WHERE 1=1{fault_scope_sql}"
            )
            station_params.extend(fault_scope_params)
        if station_sources:
            station_count = db.execute(
                "SELECT COUNT(DISTINCT station_id) as count FROM ("
                + " UNION ALL ".join(station_sources)
                + ") scoped_stations",
                station_params,
            ).fetchone()['count']

    camera_count = _count_with_alias(
        db,
        "cameras",
        "c",
        project_scope['project_ids'] if project_scope['enabled'] else None,
        active_only=True,
    )

    selected_fault_where = [f"1=1{fault_scope_sql}"]
    selected_fault_params = list(fault_scope_params)
    if 'deleted_at' in fault_report_columns:
        selected_fault_where.append("f.deleted_at IS NULL")
    if year:
        selected_fault_where.append("strftime('%Y', f.created_at) = ?")
        selected_fault_params.append(str(year))
    selected_fault_where_sql = " AND ".join(selected_fault_where)

    fault_count = db.execute(
        f"SELECT COUNT(*) as count FROM fault_reports f WHERE {selected_fault_where_sql}",
        selected_fault_params,
    ).fetchone()['count']

    fault_group_key_expr = "CAST(f.id AS TEXT)"
    if 'fault_group_key' in fault_report_columns:
        fault_group_expr = _column_expr(fault_report_columns, "f", "fault_group_key")
        fault_group_key_expr = f"COALESCE(NULLIF(TRIM({fault_group_expr}), ''), CAST(f.id AS TEXT))"


    fault_this_month = db.execute(
        f"""
        SELECT COUNT(*) as count
        FROM fault_reports f
        WHERE 1=1{fault_scope_sql}
          AND strftime('%Y-%m', f.created_at) = strftime('%Y-%m', 'now')
        """,
        fault_scope_params,
    ).fetchone()['count']
    fault_this_year = db.execute(
        f"""
        SELECT COUNT(*) as count
        FROM fault_reports f
        WHERE 1=1{fault_scope_sql}
          AND strftime('%Y', f.created_at) = strftime('%Y', 'now')
        """,
        fault_scope_params,
    ).fetchone()['count']

    target_year = year or datetime.now().year
    monthly_data = {f"{target_year}-{month:02d}": 0 for month in range(1, 13)}
    monthly_event_data = {f"{target_year}-{month:02d}": 0 for month in range(1, 13)}
    month_expr = "strftime('%Y-%m', f.created_at)"
    county_expr = "COALESCE(NULLIF(TRIM(s.county), ''), '未知')"
    voltage_level_expr = "COALESCE(NULLIF(TRIM(s.voltage_level), ''), '其他')"
    monthly_rows = db.execute(
        f"""
        SELECT {month_expr} as month, COUNT(*) as cnt
        FROM fault_reports f
        WHERE 1=1{fault_scope_sql}
          AND strftime('%Y', f.created_at) = ?
        GROUP BY {month_expr}
        """,
        fault_scope_params + [str(target_year)],
    ).fetchall()

    for row in monthly_rows:
        monthly_data[row['month']] = row['cnt']

    monthly_event_rows = db.execute(
        f"""
        SELECT event_month AS month, COUNT(*) AS cnt
        FROM (
            SELECT
                {month_expr} AS event_month,
                {fault_group_key_expr} AS fault_event_key
            FROM fault_reports f
            WHERE 1=1{fault_scope_sql}
              AND strftime('%Y', f.created_at) = ?
            GROUP BY {month_expr}, {fault_group_key_expr}
        ) deduped
        GROUP BY event_month
        """,
        fault_scope_params + [str(target_year)],
    ).fetchall()

    for row in monthly_event_rows:
        monthly_event_data[row['month']] = row['cnt']



    available_year_rows = db.execute(
        f"""
        SELECT DISTINCT strftime('%Y', f.created_at) as year
        FROM fault_reports f
        WHERE 1=1{fault_scope_sql}
          AND f.created_at IS NOT NULL
        ORDER BY strftime('%Y', f.created_at) DESC
        """,
        fault_scope_params,
    ).fetchall()

    available_years = [row['year'] for row in available_year_rows if row['year']]

    semantic_group_expr = "NULL"
    semantic_label_expr = "NULL"
    fault_type_join = ""
    if has_project_fault_types and {'fault_type_version_id', 'fault_type_code'}.issubset(fault_report_columns):
        fault_type_join = (
            " LEFT JOIN project_fault_types pft"
            " ON pft.version_id = f.fault_type_version_id"
            " AND pft.type_code = f.fault_type_code"
        )
        semantic_group_expr = "pft.semantic_group"
        semantic_label_expr = "pft.type_label"

    fault_type_code_expr = _column_expr(fault_report_columns, "f", "fault_type_code")
    fault_type_snapshot_expr = _column_expr(fault_report_columns, "f", "fault_type_label_snapshot")
    legacy_fault_type_expr = _column_expr(fault_report_columns, "f", "fault_type")

    semantic_key_expr = (
        f"COALESCE({semantic_group_expr}, NULLIF(TRIM({fault_type_code_expr}), ''), "
        f"NULLIF(TRIM({fault_type_snapshot_expr}), ''), NULLIF(TRIM({legacy_fault_type_expr}), ''), 'UNCLASSIFIED')"
    )
    semantic_label_select_expr = (
        f"COALESCE(NULLIF(TRIM({fault_type_snapshot_expr}), ''), NULLIF(TRIM({semantic_label_expr}), ''), "
        f"NULLIF(TRIM({legacy_fault_type_expr}), ''), NULLIF(TRIM({fault_type_code_expr}), ''), '未分类')"
    )

    fault_type_rows = db.execute(
        f"""
        SELECT
            normalized.semantic_group,
            normalized.fault_label,
            COUNT(*) AS count
        FROM (
            SELECT
                {semantic_key_expr} AS semantic_group,
                {semantic_label_select_expr} AS fault_label
            FROM fault_reports f
            {fault_type_join}
            WHERE {selected_fault_where_sql}
        ) AS normalized
        GROUP BY normalized.semantic_group, normalized.fault_label
        ORDER BY count DESC, normalized.fault_label ASC
        """,
        selected_fault_params,
    ).fetchall()

    county_rows = db.execute(
        f"""
        SELECT
            records.county,
            records.count,
            events.event_count
        FROM (
            SELECT
                {county_expr} AS county,
                COUNT(*) AS count
            FROM fault_reports f
            JOIN stations s ON f.station_id = s.id
            WHERE {selected_fault_where_sql}
            GROUP BY {county_expr}
        ) records
        JOIN (
            SELECT
                deduped.county,
                COUNT(*) AS event_count
            FROM (
                SELECT
                    {county_expr} AS county,
                    {fault_group_key_expr} AS fault_event_key
                FROM fault_reports f
                JOIN stations s ON f.station_id = s.id
                WHERE {selected_fault_where_sql}
                GROUP BY {county_expr}, {fault_group_key_expr}
            ) deduped
            GROUP BY deduped.county
        ) events ON events.county = records.county
        ORDER BY records.count DESC, events.event_count DESC, records.county ASC
        """,
        [*selected_fault_params, *selected_fault_params],
    ).fetchall()


    voltage_rows = db.execute(
        f"""
        SELECT {voltage_level_expr} AS voltage_level, COUNT(*) AS count
        FROM fault_reports f
        JOIN stations s ON f.station_id = s.id
        WHERE {selected_fault_where_sql}
        GROUP BY {voltage_level_expr}
        ORDER BY count DESC, voltage_level ASC
        """,
        selected_fault_params,
    ).fetchall()



    detail_table_exists = table_exists(db, "fault_report_cameras")
    camera_ranking_rows = []
    if detail_table_exists:
        detail_columns = get_table_columns(db, "fault_report_cameras")
        detail_project_scope_sql = ""
        detail_project_scope_params = []
        if project_scope['enabled'] and 'project_id' in detail_columns:
            detail_project_scope_sql, detail_project_scope_params = build_project_in_clause("d", project_scope['project_ids'])
        detail_where = [f"1=1{detail_project_scope_sql}"]
        detail_params = list(detail_project_scope_params)
        if year:
            detail_where.append("strftime('%Y', f.created_at) = ?")
            detail_params.append(str(year))
        if 'affects_statistics' in detail_columns:
            detail_where.append("COALESCE(d.affects_statistics, 1) = 1")
        detail_where_sql = " AND ".join(detail_where)
        camera_ranking_rows = db.execute(
            f"""
            SELECT
                records.camera_id,
                records.camera_location,
                records.station_name,
                records.fault_count,
                events.fault_event_count
            FROM (
                SELECT
                    d.camera_id,
                    COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(c.area), ''), NULLIF(TRIM(d.camera_label), ''), NULLIF(TRIM(d.project_device_code), ''), '未命名设备') AS camera_location,
                    s.name AS station_name,
                    COUNT(*) AS fault_count
                FROM fault_report_cameras d
                JOIN fault_reports f ON f.id = d.fault_report_id
                JOIN stations s ON f.station_id = s.id
                LEFT JOIN cameras c ON c.id = d.camera_id
                WHERE {detail_where_sql}
                GROUP BY d.camera_id, COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(c.area), ''), NULLIF(TRIM(d.camera_label), ''), NULLIF(TRIM(d.project_device_code), ''), '未命名设备'), s.name
            ) records
            JOIN (
                SELECT
                    d.camera_id,
                    COUNT(DISTINCT {fault_group_key_expr}) AS fault_event_count
                FROM fault_report_cameras d
                JOIN fault_reports f ON f.id = d.fault_report_id
                WHERE {detail_where_sql}
                GROUP BY d.camera_id
            ) events ON events.camera_id = records.camera_id
            ORDER BY records.fault_count DESC, events.fault_event_count DESC, records.camera_id ASC
            LIMIT 5
            """,
            [*detail_params, *detail_params],
        ).fetchall()
    if not camera_ranking_rows:
        camera_location_expr = "COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(c.area), ''), '未命名设备')"
        camera_owner_filter = ""
        if 'fault_owner_type' in fault_report_columns and 'root_cause_type' in fault_report_columns:
            root_cause_expr_cam = _column_expr(fault_report_columns, "f", "root_cause_type")
            fault_owner_expr_cam = _column_expr(fault_report_columns, "f", "fault_owner_type")
            camera_owner_filter = (
                f" AND (COALESCE(NULLIF(TRIM({root_cause_expr_cam}), ''), "
                f"NULLIF(TRIM({fault_owner_expr_cam}), ''), 'camera') = 'camera')"
            )
        camera_ranking_rows = db.execute(
            f"""
            SELECT
                records.camera_id,
                records.camera_location,
                records.station_name,
                records.fault_count,
                events.fault_event_count
            FROM (
                SELECT
                    f.camera_id,
                    {camera_location_expr} AS camera_location,
                    s.name AS station_name,
                    COUNT(*) AS fault_count
                FROM fault_reports f
                JOIN stations s ON f.station_id = s.id
                LEFT JOIN cameras c ON f.camera_id = c.id
                WHERE {selected_fault_where_sql}
                  AND f.camera_id IS NOT NULL
                  {camera_owner_filter}
                GROUP BY f.camera_id, {camera_location_expr}, s.name
            ) records

            JOIN (
                SELECT
                    f.camera_id,
                    COUNT(DISTINCT {fault_group_key_expr}) AS fault_event_count
                FROM fault_reports f
                WHERE {selected_fault_where_sql}
                  AND f.camera_id IS NOT NULL
                  {camera_owner_filter}
                GROUP BY f.camera_id
            ) events ON events.camera_id = records.camera_id
            ORDER BY records.fault_count DESC, events.fault_event_count DESC, records.camera_id ASC
            LIMIT 5
            """,
            [*selected_fault_params, *selected_fault_params],
        ).fetchall()



    station_ranking_rows = db.execute(
        f"""
        SELECT
            records.station_id,
            records.station_name,
            records.county,
            records.fault_count,
            records.unresolved_count,
            events.fault_event_count
        FROM (
            SELECT
                f.station_id,
                s.name AS station_name,
                {county_expr} AS county,
                COUNT(*) AS fault_count,
                SUM(CASE WHEN COALESCE(f.status, 'open') != 'closed' THEN 1 ELSE 0 END) AS unresolved_count
            FROM fault_reports f
            JOIN stations s ON f.station_id = s.id
            WHERE {selected_fault_where_sql}
            GROUP BY f.station_id, s.name, {county_expr}
        ) records

        JOIN (
            SELECT
                f.station_id,
                COUNT(DISTINCT {fault_group_key_expr}) AS fault_event_count
            FROM fault_reports f
            WHERE {selected_fault_where_sql}
            GROUP BY f.station_id
        ) events ON events.station_id = records.station_id
        ORDER BY records.fault_count DESC, events.fault_event_count DESC, records.unresolved_count DESC, records.station_name ASC
        LIMIT 5
        """,
        [*selected_fault_params, *selected_fault_params],
    ).fetchall()



    # Deduplicated fault event count（按 fault_group_key 收敛为独立事件；无分组键时回退到记录本身）
    fault_event_count = fault_count
    if 'fault_group_key' in fault_report_columns:
        fault_group_expr = _column_expr(fault_report_columns, "f", "fault_group_key")
        if Config.DATABASE_BACKEND == "postgresql":
            fault_event_count_row = db.execute(
                f"""
                SELECT COUNT(*) AS event_count
                FROM (
                    SELECT DISTINCT ON (COALESCE(NULLIF(TRIM({fault_group_expr}), ''), CAST(f.id AS TEXT)))
                        1 AS event_marker
                    FROM fault_reports f
                    WHERE {selected_fault_where_sql}
                ) deduped
                """,
                selected_fault_params,
            ).fetchone()
        else:
            fault_event_count_row = db.execute(
                f"""
                SELECT COUNT(*) AS event_count
                FROM (
                    SELECT MIN(f.id) AS _rep_id
                    FROM fault_reports f
                    WHERE {selected_fault_where_sql}
                    GROUP BY COALESCE(NULLIF(TRIM({fault_group_expr}), ''), f.id)
                ) deduped
                """,
                selected_fault_params,
            ).fetchone()
        fault_event_count = fault_event_count_row['event_count'] or 0

    # Root cause statistics (deduplicated by fault_group_key)

    root_cause_distribution = []
    batch_impact_stats = {'batch_fault_count': 0, 'avg_impact_cameras': None, 'total_impact_cameras': 0}
    if 'fault_owner_type' in fault_report_columns:
        root_cause_expr = _column_expr(fault_report_columns, "f", "root_cause_type")
        fault_owner_expr = _column_expr(fault_report_columns, "f", "fault_owner_type")
        fault_group_expr = _column_expr(fault_report_columns, "f", "fault_group_key")
        is_batch_expr = _column_expr(fault_report_columns, "f", "is_batch_impact", "0")
        impact_count_expr = _column_expr(fault_report_columns, "f", "impact_camera_count", "0")

        cause_select = f"COALESCE(NULLIF(TRIM({root_cause_expr}), ''), NULLIF(TRIM({fault_owner_expr}), ''), 'unconfirmed')"

        if Config.DATABASE_BACKEND == "postgresql":
            root_cause_rows = db.execute(
                f"""
                SELECT deduped.cause_type, COUNT(*) AS count
                FROM (
                    SELECT DISTINCT ON (COALESCE(NULLIF(TRIM({fault_group_expr}), ''), CAST(f.id AS TEXT)))
                        {cause_select} AS cause_type
                    FROM fault_reports f
                    WHERE {selected_fault_where_sql}
                ) deduped
                GROUP BY deduped.cause_type
                ORDER BY count DESC
                """,
                selected_fault_params,
            ).fetchall()


        else:
            root_cause_rows = db.execute(
                f"""
                SELECT cause_type, COUNT(*) AS count
                FROM (
                    SELECT MIN(f.id) as _rep_id,
                           {cause_select} AS cause_type
                    FROM fault_reports f
                    WHERE {selected_fault_where_sql}
                    GROUP BY COALESCE(NULLIF(TRIM({fault_group_expr}), ''), f.id)
                ) deduped
                GROUP BY cause_type
                ORDER BY count DESC
                """,
                selected_fault_params,
            ).fetchall()

        root_cause_distribution = [
            {
                'cause_type': row['cause_type'],
                'cause_label': ROOT_CAUSE_LABELS.get(row['cause_type'], row['cause_type']),
                'count': row['count'],
            }
            for row in root_cause_rows
        ]

        # Batch impact metrics（按 fault_group_key 去重，避免一组多条记录重复累加 impact_camera_count）
        if Config.DATABASE_BACKEND == "postgresql":
            batch_impact_row = db.execute(
                f"""
                SELECT
                    COUNT(*) AS batch_fault_count,
                    AVG(impact_camera_count_value) AS avg_impact_cameras,
                    SUM(impact_camera_count_value) AS total_impact_cameras
                FROM (
                    SELECT DISTINCT ON (COALESCE(NULLIF(TRIM({fault_group_expr}), ''), CAST(f.id AS TEXT)))
                        CAST(COALESCE({impact_count_expr}, 0) AS REAL) AS impact_camera_count_value
                    FROM fault_reports f
                    WHERE {selected_fault_where_sql}
                      AND COALESCE({is_batch_expr}, 0) = 1
                ) deduped
                """,
                selected_fault_params,
            ).fetchone()
        else:
            batch_impact_row = db.execute(
                f"""
                SELECT
                    COUNT(*) AS batch_fault_count,
                    AVG(impact_camera_count_value) AS avg_impact_cameras,
                    SUM(impact_camera_count_value) AS total_impact_cameras
                FROM (
                    SELECT
                        MIN(f.id) AS _rep_id,
                        CAST(COALESCE({impact_count_expr}, 0) AS REAL) AS impact_camera_count_value
                    FROM fault_reports f
                    WHERE {selected_fault_where_sql}
                      AND COALESCE({is_batch_expr}, 0) = 1
                    GROUP BY COALESCE(NULLIF(TRIM({fault_group_expr}), ''), f.id)
                ) deduped
                """,
                selected_fault_params,
            ).fetchone()

        if batch_impact_row:
            batch_impact_stats = {
                'batch_fault_count': batch_impact_row['batch_fault_count'] or 0,
                'avg_impact_cameras': round(batch_impact_row['avg_impact_cameras'], 1) if batch_impact_row['avg_impact_cameras'] is not None else None,
                'total_impact_cameras': batch_impact_row['total_impact_cameras'] or 0,
            }

    status_expr = _column_expr(fault_report_columns, "f", "status", "'open'")
    handling_started_expr = _column_expr(fault_report_columns, "f", "handling_started_at")
    closed_at_expr = _column_expr(fault_report_columns, "f", "closed_at")
    equipment_type_expr = _column_expr(fault_report_columns, "f", "equipment_type", "''")
    equipment_quantity_expr = _column_expr(fault_report_columns, "f", "equipment_quantity", "0")

    kpi_row = db.execute(
        f"""
        SELECT
            SUM(CASE WHEN {status_expr} = 'open' THEN 1 ELSE 0 END) AS open_count,
            SUM(CASE WHEN {status_expr} = 'handling' THEN 1 ELSE 0 END) AS handling_count,
            SUM(CASE WHEN {status_expr} = 'closed' THEN 1 ELSE 0 END) AS closed_count,
            COUNT(CASE WHEN {handling_started_expr} IS NOT NULL THEN 1 END) AS response_sample_count,
            COUNT(CASE WHEN {closed_at_expr} IS NOT NULL THEN 1 END) AS close_sample_count,
            SUM(
                CASE
                    WHEN COALESCE({status_expr}, 'open') != 'closed'
                         AND (julianday('now') - julianday(f.created_at)) >= ?
                    THEN 1
                    ELSE 0
                END
            ) AS overdue_unresolved_count,
            AVG(CASE
                WHEN {handling_started_expr} IS NOT NULL
                THEN (julianday({handling_started_expr}) - julianday(f.created_at)) * 86400.0
            END) AS avg_response_seconds,
            AVG(CASE
                WHEN {closed_at_expr} IS NOT NULL
                THEN (julianday({closed_at_expr}) - julianday(f.created_at)) * 86400.0
            END) AS avg_close_seconds
        FROM fault_reports f
        WHERE {selected_fault_where_sql}
        """,
        [OVERDUE_FAULT_THRESHOLD_DAYS, *selected_fault_params],
    ).fetchone()

    response_duration_rows = db.execute(
        f"""
        SELECT
            (julianday({handling_started_expr}) - julianday(f.created_at)) * 86400.0 AS duration_seconds
        FROM fault_reports f
        WHERE {selected_fault_where_sql}
          AND {handling_started_expr} IS NOT NULL
        """,
        selected_fault_params,
    ).fetchall()

    close_duration_rows = db.execute(
        f"""
        SELECT
            (julianday({closed_at_expr}) - julianday(f.created_at)) * 86400.0 AS duration_seconds
        FROM fault_reports f
        WHERE {selected_fault_where_sql}
          AND {closed_at_expr} IS NOT NULL
        """,
        selected_fault_params,
    ).fetchall()

    response_buckets = _bucketize_duration_rows(response_duration_rows, RESPONSE_BUCKET_DEFINITIONS)
    close_buckets = _bucketize_duration_rows(close_duration_rows, CLOSE_BUCKET_DEFINITIONS)

    camera_replacement_where = [f"1=1{fault_scope_sql}"]
    camera_replacement_params = list(fault_scope_params)
    camera_replacement_where.append(f"{status_expr} = 'closed'")
    camera_replacement_where.append(f"{closed_at_expr} IS NOT NULL")
    camera_replacement_where.append(f"NULLIF(TRIM({equipment_type_expr}), '') IS NOT NULL")
    if year:
        camera_replacement_where.append(f"strftime('%Y', {closed_at_expr}) = ?")
        camera_replacement_params.append(str(year))
    camera_replacement_keyword_sql = " OR ".join(
        f"{equipment_type_expr} LIKE ?" for _ in CAMERA_REPLACEMENT_TYPE_KEYWORDS
    )
    camera_replacement_where.append(f"({camera_replacement_keyword_sql})")
    camera_replacement_params.extend([f"%{keyword}%" for keyword in CAMERA_REPLACEMENT_TYPE_KEYWORDS])
    camera_replacement_where_sql = " AND ".join(camera_replacement_where)

    camera_replacement_row = db.execute(
        f"""
        SELECT
            COUNT(*) AS replacement_record_count,
            SUM(
                CASE
                    WHEN COALESCE({equipment_quantity_expr}, 0) > 0 THEN {equipment_quantity_expr}
                    ELSE 1
                END
            ) AS replacement_camera_count,
            COUNT(
                CASE
                    WHEN COALESCE({equipment_quantity_expr}, 0) <= 0 THEN 1
                END
            ) AS inferred_quantity_record_count
        FROM fault_reports f
        WHERE {camera_replacement_where_sql}
        """,
        camera_replacement_params,
    ).fetchone()

    fault_station_count_row = db.execute(
        f"""
        SELECT COUNT(DISTINCT f.station_id) AS fault_station_count
        FROM fault_reports f
        WHERE {selected_fault_where_sql}
        """,
        selected_fault_params,
    ).fetchone()

    covered_fault_station_count = 0
    uncovered_station_rows = []
    if table_exists(db, "photos"):
        photo_columns = get_table_columns(db, "photos")
        photo_scope_sql = ""
        photo_scope_params = []
        if project_scope['enabled'] and 'project_id' in photo_columns:
            photo_scope_sql, photo_scope_params = build_project_in_clause("p", project_scope['project_ids'])

        covered_fault_station_row = db.execute(
            f"""
            SELECT COUNT(DISTINCT p.station_id) AS covered_station_count
            FROM photos p
            WHERE p.match_status = 'matched'
              AND p.station_id IS NOT NULL
              {photo_scope_sql}
              AND p.station_id IN (
                  SELECT DISTINCT f.station_id
                  FROM fault_reports f
                  WHERE {selected_fault_where_sql}
              )
            """,
            [*photo_scope_params, *selected_fault_params],
        ).fetchone()
        covered_fault_station_count = covered_fault_station_row['covered_station_count'] or 0

        uncovered_station_rows = db.execute(
            f"""
            SELECT
                f.station_id,
                s.name AS station_name,
                {county_expr} AS county,
                COUNT(*) AS fault_count,
                SUM(CASE WHEN COALESCE(f.status, 'open') != 'closed' THEN 1 ELSE 0 END) AS unresolved_count
            FROM fault_reports f
            JOIN stations s ON f.station_id = s.id
            WHERE {selected_fault_where_sql}
              AND NOT EXISTS (
                  SELECT 1
                  FROM photos p
                  WHERE p.match_status = 'matched'
                    AND p.station_id = f.station_id
                    {photo_scope_sql}
              )
            GROUP BY f.station_id, s.name, {county_expr}
            ORDER BY unresolved_count DESC, fault_count DESC, s.name ASC
            LIMIT 6
            """,
            [*selected_fault_params, *photo_scope_params],
        ).fetchall()
    else:
        uncovered_station_rows = db.execute(
            f"""
            SELECT
                f.station_id,
                s.name AS station_name,
                {county_expr} AS county,
                COUNT(*) AS fault_count,
                SUM(CASE WHEN COALESCE(f.status, 'open') != 'closed' THEN 1 ELSE 0 END) AS unresolved_count
            FROM fault_reports f
            JOIN stations s ON f.station_id = s.id
            WHERE {selected_fault_where_sql}
            GROUP BY f.station_id, s.name, {county_expr}
            ORDER BY unresolved_count DESC, fault_count DESC, s.name ASC
            LIMIT 6
            """,
            selected_fault_params,
        ).fetchall()


    fault_station_count = fault_station_count_row['fault_station_count'] or 0
    uncovered_station_count = max(fault_station_count - covered_fault_station_count, 0)
    unresolved_total = (kpi_row['open_count'] or 0) + (kpi_row['handling_count'] or 0)
    overdue_unresolved_count = kpi_row['overdue_unresolved_count'] or 0

    requested_project = project_scope.get('requested_project')

    return {
        'stations': station_count,
        'cameras': camera_count,
        'faults': fault_count,
        'fault_events': fault_event_count,
        'faults_this_month': fault_this_month,
        'faults_this_year': fault_this_year,

        'fault_rate': round((fault_count / camera_count) * 100, 2) if camera_count else 0,
        'target_year': target_year,
        'selected_year': year,
        'available_years': available_years,
        'project_scope': {
            'requested_project': requested_project['code'] if requested_project else 'all',
            'visible_projects': [project['code'] for project in project_scope['projects']],
        },
        'monthly_trend': [
            {
                'month': month,
                'count': count,
                'event_count': monthly_event_data.get(month, count),
            }
            for month, count in sorted(monthly_data.items())
        ],

        'fault_type_distribution': expand_fault_type_distribution(fault_type_rows),
        'county_distribution': [dict(row) for row in county_rows],
        'voltage_distribution': [dict(row) for row in voltage_rows],
        'camera_ranking': [dict(row) for row in camera_ranking_rows],
        'station_ranking': [dict(row) for row in station_ranking_rows],
        'root_cause_distribution': root_cause_distribution,
        'batch_impact': batch_impact_stats,
        'batch_impact_event_count': batch_impact_stats['batch_fault_count'],
        'response_buckets': response_buckets,

        'close_buckets': close_buckets,
        'photo_coverage': {
            'fault_station_count': fault_station_count,
            'covered_station_count': covered_fault_station_count,
            'uncovered_station_count': uncovered_station_count,
            'coverage_ratio': round((covered_fault_station_count / fault_station_count) * 100, 2) if fault_station_count else 0,
            'uncovered_stations': [dict(row) for row in uncovered_station_rows],
        },
        'kpi': {
            'open_count': kpi_row['open_count'] or 0,
            'handling_count': kpi_row['handling_count'] or 0,
            'closed_count': kpi_row['closed_count'] or 0,
            'response_sample_count': kpi_row['response_sample_count'] or 0,
            'close_sample_count': kpi_row['close_sample_count'] or 0,
            'overdue_unresolved_count': overdue_unresolved_count,
            'overdue_unresolved_ratio': round((overdue_unresolved_count / unresolved_total) * 100, 2) if unresolved_total else 0,
            'overdue_threshold_days': OVERDUE_FAULT_THRESHOLD_DAYS,
            'avg_response_seconds': round(kpi_row['avg_response_seconds'], 2) if kpi_row['avg_response_seconds'] is not None else None,
            'avg_close_seconds': round(kpi_row['avg_close_seconds'], 2) if kpi_row['avg_close_seconds'] is not None else None,
            'camera_replacement_count': camera_replacement_row['replacement_camera_count'] or 0,
            'camera_replacement_record_count': camera_replacement_row['replacement_record_count'] or 0,
            'camera_replacement_inferred_record_count': camera_replacement_row['inferred_quantity_record_count'] or 0,
        },
    }, None


def _build_statistics_detail_rows(db, project_scope: dict, year: int | None):
    fault_report_columns = get_table_columns(db, "fault_reports")
    has_projects = table_exists(db, "projects")
    has_project_fault_types = table_exists(db, "project_fault_types")
    has_users = table_exists(db, "users")

    fault_scope_sql = ""
    fault_scope_params = []
    if project_scope['enabled'] and 'project_id' in fault_report_columns:
        fault_scope_sql, fault_scope_params = build_project_in_clause("f", project_scope['project_ids'])

    detail_where = [f"1=1{fault_scope_sql}"]
    detail_params = list(fault_scope_params)
    if year:
        detail_where.append("strftime('%Y', f.created_at) = ?")
        detail_params.append(str(year))
    detail_where_sql = " AND ".join(detail_where)

    project_join = " LEFT JOIN projects p ON p.id = f.project_id" if has_projects and 'project_id' in fault_report_columns else ""
    fault_type_join = ""
    if has_project_fault_types and {'fault_type_version_id', 'fault_type_code'}.issubset(fault_report_columns):
        fault_type_join = (
            " LEFT JOIN project_fault_types pft"
            " ON pft.version_id = f.fault_type_version_id"
            " AND pft.type_code = f.fault_type_code"
        )
    assigned_user_join = " LEFT JOIN users u ON u.id = f.assigned_to" if has_users and 'assigned_to' in fault_report_columns else ""

    fault_type_code_expr = _column_expr(fault_report_columns, "f", "fault_type_code")
    fault_type_snapshot_expr = _column_expr(fault_report_columns, "f", "fault_type_label_snapshot")
    legacy_fault_type_expr = _column_expr(fault_report_columns, "f", "fault_type")

    semantic_group_expr = "pft.semantic_group" if fault_type_join else "NULL"
    semantic_label_expr = "pft.type_label" if fault_type_join else "NULL"
    fault_label_expr = (
        f"COALESCE(NULLIF(TRIM({fault_type_snapshot_expr}), ''), NULLIF(TRIM({semantic_label_expr}), ''), "
        f"NULLIF(TRIM({legacy_fault_type_expr}), ''), NULLIF(TRIM({fault_type_code_expr}), ''), '未分类')"
    )
    semantic_key_expr = (
        f"COALESCE({semantic_group_expr}, NULLIF(TRIM({fault_type_code_expr}), ''), "
        f"NULLIF(TRIM({fault_type_snapshot_expr}), ''), NULLIF(TRIM({legacy_fault_type_expr}), ''), 'UNCLASSIFIED')"
    )

    select_fields = [
        "f.id",
        "s.name AS station_name",
        "s.voltage_level",
        "s.county",
        "COALESCE(p.code, '') AS project_code" if project_join else "'' AS project_code",
        "COALESCE(p.name, '') AS project_name" if project_join else "'' AS project_name",
        "COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(c.area), ''), '') AS camera_location",
        "COALESCE(NULLIF(TRIM(c.area), ''), '') AS camera_area",
        _column_expr(fault_report_columns, "f", "camera_slot_id") + " AS camera_slot_id",
        _column_expr(fault_report_columns, "f", "project_device_code") + " AS project_device_code",
        f"{semantic_key_expr} AS semantic_group",
        f"{fault_type_code_expr} AS fault_type_code",
        f"{fault_label_expr} AS fault_label",
        _column_expr(fault_report_columns, "f", "description") + " AS description",
        _column_expr(fault_report_columns, "f", "status", "'open'") + " AS status",
        _column_expr(fault_report_columns, "f", "reporter_name") + " AS reporter_name",
        _column_expr(fault_report_columns, "f", "reporter_contact") + " AS reporter_contact",
        _column_expr(fault_report_columns, "f", "created_at") + " AS created_at",
        _column_expr(fault_report_columns, "f", "handling_started_at") + " AS handling_started_at",
        _column_expr(fault_report_columns, "f", "closed_at") + " AS closed_at",
        _column_expr(fault_report_columns, "f", "planned_handle_time") + " AS planned_handle_time",
        _column_expr(fault_report_columns, "f", "source_type", "''") + " AS source_type",
        _column_expr(fault_report_columns, "f", "source_batch_id") + " AS source_batch_id",
        _column_expr(fault_report_columns, "f", "source_record_key") + " AS source_record_key",
        _column_expr(fault_report_columns, "f", "source_time_raw") + " AS source_time_raw",
        _column_expr(fault_report_columns, "f", "source_timezone") + " AS source_timezone",
        _column_expr(fault_report_columns, "f", "system_type") + " AS system_type",
        _column_expr(fault_report_columns, "f", "handler_name") + " AS handler_name",
        _column_expr(fault_report_columns, "f", "handler_note") + " AS handler_note",
        _column_expr(fault_report_columns, "f", "assigned_to") + " AS assigned_to",
        "COALESCE(u.username, '') AS assigned_to_username" if assigned_user_join else "'' AS assigned_to_username",
        _column_expr(fault_report_columns, "f", "tags_json") + " AS tags_json",
        _column_expr(fault_report_columns, "f", "fault_owner_type") + " AS fault_owner_type",
        _column_expr(fault_report_columns, "f", "root_cause_type") + " AS root_cause_type",
        _column_expr(fault_report_columns, "f", "is_batch_impact", "0") + " AS is_batch_impact",
        _column_expr(fault_report_columns, "f", "impact_camera_count") + " AS impact_camera_count",
    ]

    rows = db.execute(
        f"""
        SELECT {', '.join(select_fields)}
        FROM fault_reports f
        JOIN stations s ON s.id = f.station_id
        LEFT JOIN cameras c ON c.id = f.camera_id
        {project_join}
        {fault_type_join}
        {assigned_user_join}
        WHERE {detail_where_sql}
        ORDER BY f.created_at DESC, f.id DESC
        """,
        detail_params,
    ).fetchall()
    result = [dict(row) for row in rows]
    if not result:
        return result

    detail_map = fetch_fault_camera_details_map(db, [int(row['id']) for row in result])
    for row in result:
        details = detail_map.get(int(row['id']), [])
        row['camera_details'] = details
        attach_fault_camera_detail_summary(row)
        row['is_aggregated'] = '是' if len(details) > 1 else '否'
        row['statistics_camera_count'] = sum(1 for d in details if d.get('affects_statistics', 1) != 0)
    return result



def parse_tags_json(raw_value):
    if not raw_value:
        return []
    if isinstance(raw_value, list):
        candidates = raw_value
    else:
        try:
            parsed = json.loads(raw_value)
        except (TypeError, ValueError):
            parsed = str(raw_value).split(",")
        candidates = parsed if isinstance(parsed, list) else str(raw_value).split(",")

    seen = set()
    tags = []
    for item in candidates:
        tag = str(item or "").strip()
        if not tag:
            continue
        lowered = tag.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        tags.append(tag)
    return tags


def normalize_tags_payload(raw_value):
    if isinstance(raw_value, str):
        return parse_tags_json([part.strip() for part in raw_value.split(",")])
    return parse_tags_json(raw_value)


def build_station_recorders_payload(db, station_id, project_scope):
    if not table_exists(db, "station_recorders"):
        return []

    recorder_columns = get_table_columns(db, "station_recorders")
    if "project_id" not in recorder_columns:
        return []

    query = """
        SELECT
            sr.id,
            sr.station_id,
            sr.project_id,
            sr.recorder_name,
            sr.ip_address,
            sr.port,
            sr.description,
            sr.source_type,
            sr.source_key,
            sr.status,
            p.code AS project_code,
            p.name AS project_name,
            p.short_name AS project_short_name,
            p.color AS project_color,
            p.sort_order AS project_sort_order
        FROM station_recorders sr
        LEFT JOIN projects p ON p.id = sr.project_id
        WHERE sr.station_id = ?
          AND sr.status = 'active'
    """
    params = [station_id]
    if project_scope['enabled']:
        project_sql, project_params = build_project_in_clause("sr", project_scope['project_ids'])
        query += project_sql
        params.extend(project_params)
    query += """
        ORDER BY
            COALESCE(p.sort_order, 0),
            COALESCE(sr.recorder_name, ''),
            COALESCE(sr.ip_address, ''),
            COALESCE(sr.port, 0),
            sr.id
    """
    return [dict(row) for row in db.execute(query, params).fetchall()]


def _normalize_station_slot_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _is_legacy_station_slot_code(slot_code):
    return _normalize_station_slot_text(slot_code).startswith("legacy_")


def _canonicalize_station_slot_location(value, channel_number=None):
    text = _normalize_station_slot_text(value)
    if not text:
        return ""

    text = re.sub(r"(?<=\d)\s*kv", "", text, flags=re.IGNORECASE)
    if channel_number not in (None, ""):
        escaped_channel = re.escape(str(channel_number).strip())
        text = re.sub(
            rf"[-_\s#（）()]*?(?:通道\s*)?{escaped_channel}\s*#?\s*[-_\s#（）()]*?(?:球机|球|枪机|枪|半球|摄像机|摄像头)$",
            "",
            text,
            flags=re.IGNORECASE,
        )
    text = re.sub(
        r"[-_\s#（）()]*?(?:通道\s*)?\d+\s*#?\s*[-_\s#（）()]*?(?:球机|球|枪机|枪|半球|摄像机|摄像头)$",
        "",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"[-_#（）()\s]+$", "", text)
    return text.strip()


def _build_station_slot_semantic_key(payload, channel_number=None):
    if not payload:
        return ""

    resolved_channel = payload.get("channel_number", channel_number)
    location_value = (
        payload.get("location_desc")
        or payload.get("source_slot_location_desc")
        or payload.get("area")
        or ""
    )
    canonical_location = _canonicalize_station_slot_location(location_value, resolved_channel)
    if not canonical_location or resolved_channel in (None, ""):
        return ""

    station_key = payload.get("station_id")
    project_key = payload.get("project_id")
    return "|".join(
        [
            str(project_key or ""),
            str(station_key or ""),
            canonical_location,
            str(resolved_channel),
        ]
    )


def _station_slot_display_priority(label, channel_number):
    normalized = _normalize_station_slot_text(label)
    if not normalized:
        return (0, 0)
    canonical = _canonicalize_station_slot_location(label, channel_number)
    suffix_removed = canonical and canonical != normalized
    return (1 if not suffix_removed else 0, -len(str(label)))


def _pick_station_slot_display_location(*slots):
    best_label = ""
    best_score = None
    for slot in slots:
        if not slot:
            continue
        label = str(slot.get("location_desc") or "").strip()
        if not label:
            continue
        score = _station_slot_display_priority(label, slot.get("channel_number"))
        if best_score is None or score > best_score:
            best_label = label
            best_score = score
    return best_label


def _station_slot_preference(slot):
    slot_code = str(slot.get("slot_code") or "")
    return (
        1 if slot.get("current_camera") else 0,
        1 if not _is_legacy_station_slot_code(slot_code) else 0,
        1 if _station_slot_display_priority(slot.get("location_desc"), slot.get("channel_number"))[0] else 0,
        int(slot.get("fault_count") or 0),
        int(slot.get("history_camera_count") or 0),
        -len(slot_code),
    )


def _merge_station_recent_faults(primary_faults, secondary_faults):
    merged = {}
    for fault in list(primary_faults or []) + list(secondary_faults or []):
        if not fault or not fault.get("id"):
            continue
        merged[fault["id"]] = fault
    return sorted(
        merged.values(),
        key=lambda item: (
            str(item.get("created_at") or ""),
            int(item.get("id") or 0),
        ),
        reverse=True,
    )[:3]


def _station_history_camera_priority(camera):
    status = str(camera.get("status") or "").strip().lower()
    return (
        1 if camera.get("replaced_by_camera_id") else 0,
        1 if status == "replaced" else 0,
        1 if status == "retired" else 0,
        1 if camera.get("ip_address") else 0,
        -int(camera.get("id") or 0),
    )


def _filter_station_duplicate_history_cameras(history_cameras, current_camera=None):
    deduped_by_id = []
    seen_ids = set()
    for camera in history_cameras or []:
        if not camera:
            continue
        camera_id = camera.get("id")
        if camera_id and camera_id in seen_ids:
            continue
        if camera_id:
            seen_ids.add(camera_id)
        deduped_by_id.append(camera)

    current_key = _build_station_slot_semantic_key(current_camera) if current_camera else ""
    buckets = {}
    passthrough = []
    for index, camera in enumerate(deduped_by_id):
        semantic_key = _build_station_slot_semantic_key(camera)
        if not semantic_key:
            passthrough.append((index, camera))
            continue
        buckets.setdefault(semantic_key, []).append((index, camera))

    filtered = [camera for _, camera in passthrough]
    for semantic_key, bucket in buckets.items():
        entries = [camera for _, camera in bucket]
        if len(entries) == 1:
            filtered.append(entries[0])
            continue

        non_legacy_entries = [
            camera for camera in entries if not _is_legacy_station_slot_code(camera.get("source_slot_code"))
        ]

        if semantic_key == current_key:
            replacement_linked = [
                camera
                for camera in entries
                if camera.get("replaced_by_camera_id")
            ]
            if replacement_linked:
                filtered.extend(replacement_linked)
                continue

            replaced_status_entries = [
                camera
                for camera in entries
                if str(camera.get("status") or "").strip().lower() == "replaced"
            ]
            if replaced_status_entries:
                filtered.extend(replaced_status_entries)
                continue

            if non_legacy_entries and len(non_legacy_entries) != len(entries):
                filtered.extend(non_legacy_entries)
                continue

            filtered.append(max(entries, key=_station_history_camera_priority))
            continue

        if non_legacy_entries and len(non_legacy_entries) != len(entries):
            filtered.extend(non_legacy_entries)
            continue

        filtered.extend(entries)

    return sorted(
        filtered,
        key=lambda item: (
            str(item.get("retired_at") or item.get("created_at") or ""),
            int(item.get("id") or 0),
        ),
        reverse=True,
    )


def _merge_station_duplicate_slots(primary, secondary):
    preferred = primary if _station_slot_preference(primary) >= _station_slot_preference(secondary) else secondary
    fallback = secondary if preferred is primary else primary
    merged_current_camera = preferred.get("current_camera") or fallback.get("current_camera")
    merged_history = _filter_station_duplicate_history_cameras(
        list(preferred.get("history_cameras") or []) + list(fallback.get("history_cameras") or []),
        current_camera=merged_current_camera,
    )
    return {
        **fallback,
        **preferred,
        "location_desc": _pick_station_slot_display_location(primary, secondary),
        "area": preferred.get("area") or fallback.get("area") or "",
        "current_camera": merged_current_camera,
        "history_cameras": merged_history,
        "history_camera_count": len(merged_history),
        "fault_count": max(int(primary.get("fault_count") or 0), int(secondary.get("fault_count") or 0)),
        "recent_faults": _merge_station_recent_faults(primary.get("recent_faults"), secondary.get("recent_faults")),
        "recorder": preferred.get("recorder") or fallback.get("recorder"),
    }


def dedupe_station_slots_payload(slots):
    deduped = []
    key_index_map = {}
    for slot in slots or []:
        semantic_key = _build_station_slot_semantic_key(slot)
        if not semantic_key:
            deduped.append(slot)
            continue

        if semantic_key not in key_index_map:
            key_index_map[semantic_key] = len(deduped)
            deduped.append(slot)
            continue

        existing_index = key_index_map[semantic_key]
        deduped[existing_index] = _merge_station_duplicate_slots(deduped[existing_index], slot)
    return deduped


def build_station_slots_payload(db, station_id, project_scope):
    ensure_camera_recorder_metadata_columns(db)
    camera_columns = get_table_columns(db, "cameras")
    fault_report_columns = ensure_fault_report_soft_delete_schema(db)
    station_recorders = build_station_recorders_payload(db, station_id, project_scope)

    if not (
        table_exists(db, "camera_slots")
        and 'slot_id' in camera_columns
        and 'project_id' in camera_columns
    ):
        return []

    slot_query = """
        SELECT
            s.id AS slot_id,
            s.slot_code,
            s.station_id,
            s.project_id,
            s.location_desc,
            s.area,
            s.channel_number,
            p.code AS project_code,
            p.name AS project_name,
            p.short_name AS project_short_name,
            p.color AS project_color,
            p.sort_order AS project_sort_order,
            c.id AS current_camera_id,
            c.project_camera_code AS current_project_camera_code,
            c.camera_index AS current_camera_index,
            c.ip_address AS current_ip_address,
            c.channel_port AS current_channel_port,
            c.channel_number AS current_channel_number,
            c.recorder_name AS current_recorder_name,
            c.recorder_ip_address AS current_recorder_ip_address,
            c.recorder_port AS current_recorder_port,
            c.area AS current_area,
            c.location_desc AS current_location_desc,
            c.created_at AS current_created_at
        FROM camera_slots s
        LEFT JOIN projects p ON p.id = s.project_id
        LEFT JOIN cameras c
            ON c.slot_id = s.id
           AND c.status = 'active'
        WHERE s.station_id = ?
    """
    slot_params = [station_id]
    if project_scope['enabled']:
        project_sql, project_params = build_project_in_clause("s", project_scope['project_ids'])
        slot_query += project_sql
        slot_params.extend(project_params)
    slot_query += """
        ORDER BY
            COALESCE(p.sort_order, 0),
            COALESCE(s.area, ''),
            COALESCE(s.location_desc, ''),
            COALESCE(s.channel_number, 0),
            s.id
    """
    slot_rows = db.execute(slot_query, slot_params).fetchall()
    if not slot_rows:
        return []

    slot_ids = [row['slot_id'] for row in slot_rows]
    placeholders = ", ".join(["?"] * len(slot_ids))

    history_query = f"""
        SELECT
            c.id,
            c.slot_id,
            c.station_id,
            c.project_id,
            c.project_camera_code,
            c.camera_index,
            c.area,
            c.location_desc,
            c.ip_address,
            c.channel_port,
            c.channel_number,
            c.recorder_name,
            c.recorder_ip_address,
            c.recorder_port,
            c.status,
            c.replaced_by_camera_id,
            c.retired_at,
            c.created_at,
            s.slot_code AS source_slot_code,
            s.location_desc AS source_slot_location_desc
        FROM cameras c
        LEFT JOIN camera_slots s ON s.id = c.slot_id
        WHERE c.slot_id IN ({placeholders})
          AND c.status != 'active'
        ORDER BY COALESCE(c.retired_at, c.created_at) DESC, c.id DESC
    """
    history_rows = db.execute(history_query, slot_ids).fetchall()
    history_map = {slot_id: [] for slot_id in slot_ids}
    for row in history_rows:
        history_map.setdefault(row['slot_id'], []).append(dict(row))

    recorders_by_project = {}
    recorder_by_exact = {}
    recorder_by_ip = {}
    recorder_by_name = {}
    for recorder in station_recorders:
        project_id = recorder.get('project_id')
        recorders_by_project.setdefault(project_id, []).append(recorder)
        recorder_name_key = normalize_station_name(recorder.get('recorder_name'))
        if recorder_name_key:
            recorder_by_name.setdefault((project_id, recorder_name_key), []).append(recorder)

        ip_key = str(recorder.get('ip_address') or '').strip()
        if not ip_key:
            continue

        port_value = recorder.get('port')
        if port_value is not None:
            recorder_by_exact[(project_id, ip_key, port_value)] = recorder

        ip_bucket_key = (project_id, ip_key)
        recorder_by_ip.setdefault(ip_bucket_key, []).append(recorder)

    def match_slot_recorder(project_id, current_camera, history_cameras):
        candidates = []
        if current_camera:
            candidates.append(current_camera)
        candidates.extend(history_cameras or [])

        for camera in candidates:
            recorder_name_key = normalize_station_name(camera.get('recorder_name'))
            if recorder_name_key:
                named_recorders = recorder_by_name.get((project_id, recorder_name_key), [])
                if len(named_recorders) == 1:
                    return named_recorders[0]

                recorder_ip_key = str(camera.get('recorder_ip_address') or '').strip()
                recorder_port = camera.get('recorder_port')
                for recorder in named_recorders:
                    if recorder_ip_key and recorder_ip_key != str(recorder.get('ip_address') or '').strip():
                        continue
                    if recorder_port is not None and recorder_port != recorder.get('port'):
                        continue
                    return recorder

            ip_key = str(camera.get('ip_address') or '').strip()
            if not ip_key:
                continue

            port_value = camera.get('channel_port')
            if port_value is not None:
                recorder = recorder_by_exact.get((project_id, ip_key, port_value))
                if recorder:
                    return recorder

            recorder_bucket = recorder_by_ip.get((project_id, ip_key), [])
            if len(recorder_bucket) == 1:
                return recorder_bucket[0]

        project_recorders = recorders_by_project.get(project_id, [])
        if len(project_recorders) == 1:
            return project_recorders[0]
        return None

    fault_count_map = {slot_id: 0 for slot_id in slot_ids}
    recent_faults_map = {slot_id: [] for slot_id in slot_ids}
    if 'camera_slot_id' in fault_report_columns:
        fault_query = f"""
            SELECT
                f.camera_slot_id AS slot_id,
                COUNT(*) AS fault_count,
                MAX(f.created_at) AS last_fault_at
            FROM fault_reports f
            WHERE f.station_id = ?
              AND f.camera_slot_id IN ({placeholders})
              AND f.deleted_at IS NULL
        """
        fault_params = [station_id, *slot_ids]
        if project_scope['enabled'] and 'project_id' in fault_report_columns:
            project_sql, project_params = build_project_in_clause("f", project_scope['project_ids'])
            fault_query += project_sql
            fault_params.extend(project_params)
        fault_query += " GROUP BY f.camera_slot_id"
        for row in db.execute(fault_query, fault_params).fetchall():
            fault_count_map[row['slot_id']] = row['fault_count']

        recent_fault_query = f"""
            SELECT
                f.id,
                f.camera_slot_id AS slot_id,
                COALESCE(f.fault_type_label_snapshot, f.fault_type) AS fault_label,
                COALESCE(f.description, '') AS description,
                f.status,
                f.created_at,
                f.closed_at,
                COALESCE(f.handler_note, '') AS handler_note
            FROM fault_reports f
            WHERE f.station_id = ?
              AND f.camera_slot_id IN ({placeholders})
              AND f.deleted_at IS NULL
        """
        recent_fault_params = [station_id, *slot_ids]
        if project_scope['enabled'] and 'project_id' in fault_report_columns:
            project_sql, project_params = build_project_in_clause("f", project_scope['project_ids'])
            recent_fault_query += project_sql
            recent_fault_params.extend(project_params)
        recent_fault_query += " ORDER BY f.created_at DESC, f.id DESC"
        for row in db.execute(recent_fault_query, recent_fault_params).fetchall():
            bucket = recent_faults_map.setdefault(row['slot_id'], [])
            if len(bucket) < 3:
                bucket.append(dict(row))

    slots = []
    for row in slot_rows:
        row_dict = dict(row)
        current_camera = None
        if row['current_camera_id']:
            current_camera = {
                'id': row['current_camera_id'],
                'station_id': row['station_id'],
                'project_id': row['project_id'],
                'project_camera_code': row['current_project_camera_code'],
                'camera_index': row['current_camera_index'],
                'ip_address': row['current_ip_address'],
                'channel_port': row['current_channel_port'],
                'channel_number': row['current_channel_number'],
                'recorder_name': row['current_recorder_name'],
                'recorder_ip_address': row['current_recorder_ip_address'],
                'recorder_port': row['current_recorder_port'],
                'area': row['current_area'],
                'location_desc': row['current_location_desc'],
                'created_at': row['current_created_at'],
                'status': 'active',
                'source_slot_code': row['slot_code'],
                'source_slot_location_desc': row['location_desc'],
            }

        matched_recorder = match_slot_recorder(
            row['project_id'],
            current_camera,
            history_map.get(row['slot_id'], []),
        )

        slots.append({
            'slot_id': row['slot_id'],
            'slot_code': row['slot_code'],
            'station_id': row['station_id'],
            'project_id': row['project_id'],
            'project_code': row['project_code'],
            'project_name': row['project_name'],
            'project_short_name': row['project_short_name'],
            'project_color': row['project_color'],
            'project_sort_order': row['project_sort_order'],
            'location_desc': row['location_desc'],
            'area': row['area'],
            'channel_number': row['channel_number'],
            'current_camera': current_camera,
            'history_cameras': history_map.get(row['slot_id'], []),
            'history_camera_count': len(history_map.get(row['slot_id'], [])),
            'fault_count': fault_count_map.get(row['slot_id'], 0),
            'recent_faults': recent_faults_map.get(row['slot_id'], []),
            'recorder': matched_recorder,
        })
    return dedupe_station_slots_payload(slots)

# ============================================================
# API: 项目列表
# ============================================================

@app.route('/api/projects', methods=['GET'])
def get_projects_api():
    """获取项目列表与当前用户可见范围"""
    db = get_db()
    context = get_current_user_project_context(db)
    if context['user_id']:
        projects = context['projects']
    else:
        projects = get_visible_projects(db, user_id=None, role='admin', include_inactive=False)
        for project in projects:
            project['can_write'] = False

    return api_success({
        'projects': projects,
        'default_project_code': get_default_project_code(projects),
        'multi_project_enabled': projects_enabled(db),
    })


# ============================================================
# API: 项目故障类型
# ============================================================

@app.route('/api/projects/<string:project_code>/fault-types', methods=['GET'])
def get_project_fault_types(project_code):
    """获取项目当前已发布故障类型"""
    db = get_db()
    project = get_project_by_code(db, project_code, include_inactive=True)
    if not project:
        return api_error('项目不存在', 404)
    if not ensure_project_read_access(db, project_code):
        return project_access_denied()

    if not table_exists(db, 'project_fault_types') or not table_exists(db, 'project_fault_type_versions'):
        return api_success({
            'project': project,
            'fault_types': [],
        })

    if not project.get('fault_type_version_id'):
        return api_success({
            'project': project,
            'fault_types': [],
        })

    rows = db.execute(
        """
        SELECT type_code, type_label, semantic_group, sort_order, is_active
        FROM project_fault_types
        WHERE version_id = ?
        ORDER BY sort_order, id
        """,
        (project['fault_type_version_id'],),
    ).fetchall()

    return api_success({
        'project': project,
        'fault_types': [dict(row) for row in rows],
    })


# ============================================================
# API: 统计概览
# ============================================================

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """首页统计数据"""
    db = get_db()
    year = request.args.get('year', type=int)

    # 变电站数量（不受年份影响）
    station_count = db.execute("SELECT COUNT(*) as count FROM stations").fetchone()['count']

    # 摄像头数量（不受年份影响）
    camera_count = db.execute("SELECT COUNT(*) as count FROM cameras").fetchone()['count']

    # 故障报修数量（支持年份筛选）
    if year:
        fault_count = db.execute(
            "SELECT COUNT(*) as count FROM fault_reports WHERE strftime('%Y', created_at) = ?",
            (str(year),)
        ).fetchone()['count']
    else:
        fault_count = db.execute("SELECT COUNT(*) as count FROM fault_reports").fetchone()['count']

    # 本月故障数量
    fault_this_month = db.execute("""
        SELECT COUNT(*) as count FROM fault_reports
        WHERE strftime('%Y-%m', created_at) = strftime('%Y-%m', 'now')
    """).fetchone()['count']

    # 本年故障数量
    fault_this_year = db.execute("""
        SELECT COUNT(*) as count FROM fault_reports
        WHERE strftime('%Y', created_at) = strftime('%Y', 'now')
    """).fetchone()['count']

    return api_success({
        'stations': station_count,
        'cameras': camera_count,
        'faults': fault_count,
        'faults_this_month': fault_this_month,
        'faults_this_year': fault_this_year
    })

# ============================================================
# API: 统计导出Excel
# ============================================================

@app.route('/api/statistics/export', methods=['GET'])
def export_statistics():
    """导出统计报表为Excel"""
    year = request.args.get('year', type=int)
    db = get_db()

    # 概览数据
    station_count = db.execute("SELECT COUNT(*) as count FROM stations").fetchone()['count']
    camera_count = db.execute("SELECT COUNT(*) as count FROM cameras").fetchone()['count']

    target_year = year or datetime.now().year
    monthly_data = {f"{target_year}-{m:02d}": 0 for m in range(1, 13)}
    monthly_event_data = {f"{target_year}-{m:02d}": 0 for m in range(1, 13)}

    # 故障记录明细
    query = """
        SELECT f.id, s.name as station_name, s.voltage_level, s.county,
               c.area as camera_area, c.location_desc as camera_location,
               f.fault_type, f.description, f.status,
               f.reporter_name, f.reporter_contact,
               f.created_at, f.closed_at, f.handler_name, f.handler_note,
               f.fault_group_key, f.fault_owner_type, f.root_cause_type,
               f.is_batch_impact, f.impact_camera_count
        FROM fault_reports f
        JOIN stations s ON f.station_id = s.id
        LEFT JOIN cameras c ON f.camera_id = c.id
    """
    params = []
    if year:
        query += " WHERE strftime('%Y', f.created_at) = ?"
        params.append(str(year))
    else:
        query += " WHERE strftime('%Y', f.created_at) = ?"
        params.append(str(target_year))
    query += " ORDER BY f.created_at DESC"
    faults = db.execute(query, params).fetchall()
    fault_rows = [dict(row) for row in faults]
    detail_map = fetch_fault_camera_details_map(db, [int(row['id']) for row in fault_rows]) if fault_rows else []
    for fault_row in fault_rows:
        fault_row['camera_details'] = detail_map.get(int(fault_row['id']), []) if detail_map else []
        attach_fault_camera_detail_summary(fault_row)

    fault_count = len(fault_rows)

    # 记录口径 / 事件口径汇总
    fault_event_keys = set()
    county_record_data = {}
    county_event_data = {}
    voltage_record_data = {}
    voltage_event_data = {}
    monthly_event_seen = set()


    for f in fault_rows:

        created_at = (f['created_at'] or '').strip()
        month_key = created_at[:7] if len(created_at) >= 7 else ''
        fault_event_key = (f['fault_group_key'] or '').strip() or str(f['id'])
        county = (f['county'] or '').strip() or '未知'
        voltage_level = (f['voltage_level'] or '').strip() or '其他'

        if month_key in monthly_data:
            monthly_data[month_key] += 1
        voltage_record_data[voltage_level] = voltage_record_data.get(voltage_level, 0) + 1
        county_record_data[county] = county_record_data.get(county, 0) + 1

        if fault_event_key not in fault_event_keys:
            fault_event_keys.add(fault_event_key)
            county_event_data[county] = county_event_data.get(county, 0) + 1
            voltage_event_data[voltage_level] = voltage_event_data.get(voltage_level, 0) + 1


        if month_key and (month_key, fault_event_key) not in monthly_event_seen:
            monthly_event_seen.add((month_key, fault_event_key))
            if month_key in monthly_event_data:
                monthly_event_data[month_key] += 1

    fault_event_count = len(fault_event_keys)
    county_rows = [
        {
            'county': county,
            'count': count,
            'event_count': county_event_data.get(county, 0),
        }
        for county, count in sorted(
            county_record_data.items(),
            key=lambda item: (-item[1], -county_event_data.get(item[0], 0), item[0])
        )
    ]
    voltage_rows = [
        {
            'voltage_level': voltage_level,
            'count': count,
            'event_count': voltage_event_data.get(voltage_level, 0),
        }
        for voltage_level, count in sorted(
            voltage_record_data.items(),
            key=lambda item: (-item[1], -voltage_event_data.get(item[0], 0), item[0])
        )
    ]



    # 生成Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb = Workbook()

        # 样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2f67f6', end_color='2f67f6', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: 概览
        ws1 = wb.active
        ws1.title = '概览'
        overview = [
            ('变电站总数', station_count),
            ('摄像头总数', camera_count),
            ('故障记录数', fault_count),
            ('独立故障事件数', fault_event_count),
            ('故障率', f"{(fault_count/camera_count*100):.2f}%" if camera_count > 0 else '0%'),
        ]
        ws1.append(['指标', '数值'])

        for k, v in overview:
            ws1.append([k, v])
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 15

        # Sheet 2: 月度趋势
        ws2 = wb.create_sheet('月度趋势')
        ws2.append(['月份', '故障记录数', '独立事件数'])
        for month, cnt in sorted(monthly_data.items()):
            ws2.append([month, cnt, monthly_event_data.get(month, 0)])
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15


        # Sheet 3: 故障明细
        ws3 = wb.create_sheet('故障明细')
        headers = ['ID', '变电站', '电压等级', '县区', '摄像头位置', '摄像头明细', '逐路恢复摘要', '自恢复路数', '故障类型',
                   '描述', '状态', '报修人', '联系方式', '报修时间', '关闭时间', '处理人', '处理备注',
                   '故障归属', '根因确认', '共因标记', '影响摄像头数', '是否聚合单', '计入统计摄像头数']
        ws3.append(headers)
        for f in fault_rows:
            fault_owner_label = ROOT_CAUSE_LABELS.get((f.get('fault_owner_type') or ''), f.get('fault_owner_type') or '')
            root_cause_label = ROOT_CAUSE_LABELS.get((f.get('root_cause_type') or ''), f.get('root_cause_type') or '')
            camera_details = f.get('camera_details') or []
            stats_camera_count = sum(1 for d in camera_details if d.get('affects_statistics', 1) != 0)
            is_aggregated = '是' if len(camera_details) > 1 else '否'
            ws3.append([
                f['id'], f['station_name'] or '', f['voltage_level'] or '', f['county'] or '',
                f.get('camera_display_text') or f.get('camera_location') or f.get('camera_area') or '',
                f.get('camera_locations_text') or '',
                f.get('camera_recovery_text') or '',
                f.get('self_recovered_camera_count', ''),
                f['fault_type'] or '',
                f['description'] or '', f['status'] or '', f['reporter_name'] or '',
                f['reporter_contact'] or '', f['created_at'] or '', f['closed_at'] or '',
                f['handler_name'] or '', f['handler_note'] or '',
                fault_owner_label, root_cause_label,
                f.get('is_batch_impact', ''),
                f.get('impact_camera_count', ''),
                is_aggregated,
                stats_camera_count,
            ])

        for col in ws3.columns:
            ws3.column_dimensions[col[0].column_letter].width = 15

        # Sheet 4: 县区统计
        ws4 = wb.create_sheet('县区统计')
        ws4.append(['县区', '故障记录数', '独立事件数'])
        for row in county_rows:
            ws4.append([row['county'], row['count'], row['event_count']])
        ws4.column_dimensions['A'].width = 15
        ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 15


        # Sheet 5: 电压等级统计
        ws5 = wb.create_sheet('电压等级统计')
        ws5.append(['电压等级', '故障记录数', '独立事件数'])
        for row in voltage_rows:
            ws5.append([row['voltage_level'], row['count'], row['event_count']])
        ws5.column_dimensions['A'].width = 15
        ws5.column_dimensions['B'].width = 15
        ws5.column_dimensions['C'].width = 15


        # 输出
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"统计报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.exception('export_statistics failed')
        return api_error(f'导出失败: {e}', 500)

# ============================================================
# API: 变电站列表
# ============================================================

@app.route('/api/stations', methods=['GET'])
def get_stations():
    """获取变电站列表"""
    db = get_db()

    # 支持按县区筛选
    county = request.args.get('county')
    query = "SELECT id, name, voltage_level, county, location, latitude, longitude FROM stations"
    params = []

    if county:
        query += " WHERE county = ?"
        params.append(county)

    query += " ORDER BY county, name"

    rows = db.execute(query, params).fetchall()

    return api_success({
        'stations': [dict(row) for row in rows],
        'total': len(rows)
    })

# ============================================================
# API: 变电站详情
# ============================================================

@app.route('/api/stations/<int:station_id>', methods=['GET'])
def get_station(station_id):
    """获取变电站详情（含JOIN）"""
    db = get_db()

    # 获取变电站信息
    station = db.execute("""
        SELECT s.*,
               (SELECT COUNT(*) FROM cameras WHERE station_id = s.id) as camera_count,
               (SELECT COUNT(*) FROM fault_reports WHERE station_id = s.id) as fault_count
        FROM stations s
        WHERE s.id = ?
    """, (station_id,)).fetchone()

    if not station:
        return api_error('变电站不存在', 404)

    # 获取摄像头列表
    cameras = db.execute("""
        SELECT * FROM cameras WHERE station_id = ?
        ORDER BY camera_index, channel_number
    """, (station_id,)).fetchall()

    return api_success({
        'station': dict(station),
        'cameras': [dict(c) for c in cameras]
    })


@app.route('/api/stations/<int:station_id>/slots', methods=['GET'])
def get_station_slots(station_id):
    """获取站点槽位列表"""
    db = get_db()
    station = db.execute(
        "SELECT id, name FROM stations WHERE id = ?",
        (station_id,),
    ).fetchone()
    if not station:
        return api_error('变电站不存在', 404)

    project_scope = {
        'enabled': False,
        'project_ids': None,
        'requested_project': None,
        'projects': [],
    }
    recorders = build_station_recorders_payload(
        db,
        station_id,
        project_scope,
    )
    slots = build_station_slots_payload(
        db,
        station_id,
        project_scope,
    )
    return api_success({
        'station': dict(station),
        'slots': slots,
        'recorders': recorders,
        'total': len(slots),
    })

# ============================================================
# API: 密码查询（需Token认证，决策#1）
# ============================================================

@app.route('/api/stations/<int:station_id>/password', methods=['GET'])
@require_token
def get_station_password(station_id):
    """获取变电站密码（需认证）"""
    db = get_db()

    station = db.execute("SELECT id, name, nvr_ip, nvr_port FROM stations WHERE id = ?",
                         (station_id,)).fetchone()

    if not station:
        return api_error('变电站不存在', 404)

    # 注意：实际密码存储需要单独字段，这里返回NVR连接信息作为示例
    # 完整实现需要在stations表增加password字段
    return api_success({
        'station_id': station['id'],
        'station_name': station['name'],
        'nvr_ip': station['nvr_ip'],
        'nvr_port': station['nvr_port'],
        'message': '密码字段需要扩展数据库'
    })

# ============================================================
# API: 摄像头列表
# ============================================================

@app.route('/api/cameras', methods=['GET'])
def get_cameras():
    """获取摄像头列表"""
    db = get_db()

    # 支持按变电站筛选
    station_id = request.args.get('station_id', type=int)

    query = """
        SELECT c.*, s.name as station_name, s.voltage_level
        FROM cameras c
        JOIN stations s ON c.station_id = s.id
    """
    params = []

    if station_id:
        query += " WHERE c.station_id = ?"
        params.append(station_id)

    query += " ORDER BY s.county, s.name, c.camera_index, c.channel_number"

    rows = db.execute(query, params).fetchall()

    return api_success({
        'cameras': [dict(row) for row in rows],
        'total': len(rows)
    })

# ============================================================
# API: 按IP查摄像头
# ============================================================

@app.route('/api/cameras/by-ip', methods=['GET'])
def get_camera_by_ip():
    """按IP地址查询摄像头"""
    ip = request.args.get('ip', '').strip()

    if not ip:
        return api_error('未提供IP地址')

    db = get_db()

    camera = db.execute("""
        SELECT c.*, s.name as station_name, s.voltage_level, s.county
        FROM cameras c
        JOIN stations s ON c.station_id = s.id
        WHERE c.ip_address = ?
    """, (ip,)).fetchone()

    if not camera:
        return api_error('该IP暂未录入系统，请选择变电站手动报修', 404)

    return api_success({'camera': dict(camera)})

# ============================================================
# API: 故障提交（决策#7：幂等键）
# ============================================================

@app.route('/api/faults', methods=['POST'])
def create_fault():
    """提交故障报修"""
    data = request.get_json()
    if not data:
        return api_error('请求体无效')

    logger.info(f"Fault report: station={data.get('station_id')}, type={data.get('fault_type')}, reporter={data.get('reporter_name')}")

    # 必填字段验证
    required = ['station_id', 'reporter_name']
    for field in required:
        if not data.get(field):
            return api_error(f'缺少必填字段: {field}')

    db = get_db()
    fault_report_columns = ensure_fault_report_multi_camera_schema(db)
    camera_columns = get_table_columns(db, "cameras")
    project = None
    project_id = None
    project_code = (data.get('project') or '').strip()
    try:
        selected_camera_ids = normalize_camera_ids(data.get('camera_ids'))
        if not selected_camera_ids and data.get('camera_id') not in (None, ""):
            selected_camera_ids = normalize_camera_ids([data.get('camera_id')])
    except ValueError:
        return api_error('摄像头参数无效', 400)
    fault_type_value = str(data.get('fault_type') or '').strip() or DEFAULT_PENDING_FAULT_TYPE
    fault_type_label_snapshot = fault_type_value
    fault_type_version_id = None
    camera_rows = []

    if selected_camera_ids:
        camera_rows = fetch_camera_rows_by_ids(db, selected_camera_ids, camera_columns)
        if not camera_rows:
            return api_error('摄像头不存在', 404)

    if project_code:
        project = get_project_by_code(db, project_code, include_inactive=False)
        if not project:
            return api_error('项目不存在', 404)
        if not ensure_project_read_access(db, project['code']):
            return project_access_denied()
        project_id = project['id']
    elif 'project_id' in fault_report_columns and projects_enabled(db):
        inferred_project_ids = {
            row['project_id']
            for row in camera_rows
            if 'project_id' in camera_columns and row['project_id'] is not None
        }
        if len(inferred_project_ids) > 1:
            return api_error('摄像头不属于同一项目', 400)
        if inferred_project_ids:
            project_id = next(iter(inferred_project_ids))
        if project_id is None:
            default_project_code = get_default_project_code(
                get_visible_projects(
                    db,
                    user_id=session.get('user_id'),
                    role=session.get('role') or 'admin',
                    include_inactive=False,
                )
            ) or 'unified'
            project = get_project_by_code(db, default_project_code, include_inactive=False)
            project_id = project['id'] if project else None

    if project is None and project_id is not None and projects_enabled(db):
        project_row = db.execute(
            "SELECT code FROM projects WHERE id = ?",
            (project_id,),
        ).fetchone()
        if project_row:
            project = get_project_by_code(db, project_row['code'], include_inactive=True)

    if projects_enabled(db) and project_id is not None:
        if not project or not project.get('is_active'):
            return api_error('椤圭洰涓嶅瓨鍦ㄦ垨宸插仠鐢?', 404)
        if not ensure_project_write_access(db, project['code']):
            return project_access_denied()

    # 验证变电站存在
    station = db.execute("SELECT id, name FROM stations WHERE id = ?",
                         (data['station_id'],)).fetchone()
    if not station:
        return api_error('变电站不存在', 404)

    for camera_row in camera_rows:
        if camera_row['station_id'] != data['station_id']:
            return api_error('摄像头与变电站不匹配', 400)
        camera_project_id = camera_row['project_id'] if 'project_id' in camera_columns else None
        if project_id is not None and 'project_id' in camera_columns and camera_project_id not in (None, project_id):
            return api_error('摄像头与项目不匹配', 400)

    if data.get('fault_type_code') and project and project.get('fault_type_version_id'):
        fault_type_row = db.execute(
            """
            SELECT type_label
            FROM project_fault_types
            WHERE version_id = ?
              AND type_code = ?
              AND is_active = 1
            """,
            (project['fault_type_version_id'], data.get('fault_type_code')),
        ).fetchone()
        if not fault_type_row:
            return api_error('故障类型不存在或未发布', 400)
        fault_type_value = fault_type_row['type_label']
        fault_type_label_snapshot = fault_type_row['type_label']
        fault_type_version_id = project['fault_type_version_id']

    # 计算幂等键（决策#7）
    # 幂等键 = 每个 camera_id + FLOOR(report_time / 300秒)
    report_time = data.get('report_time')
    import time
    current_time = int(report_time or time.time())
    window = math.floor(current_time / 300)

    if selected_camera_ids:
        idempotency_keys = [f"{camera_id}_{window}" for camera_id in selected_camera_ids]
    else:
        # 使用IP文本的哈希
        import hashlib
        ip_text = data.get('camera_ip_free_text', '')
        if ip_text:
            idempotency_keys = [hashlib.md5(ip_text.encode()).hexdigest()[:16]]
        else:
            idempotency_keys = [None]

    # 检查幂等冲突
    conflicting_camera_ids = []
    for index, idempotency_key in enumerate(idempotency_keys):
        if not idempotency_key:
            continue
        existing = db.execute(
            "SELECT id FROM fault_reports WHERE idempotency_key = ?",
            (idempotency_key,),
        ).fetchone()
        if existing:
            if selected_camera_ids:
                conflicting_camera_ids.append(selected_camera_ids[index])
            else:
                return api_error('该摄像头5分钟内有报修记录，请勿重复提交', 409)

    if conflicting_camera_ids:
        if len(selected_camera_ids) == 1:
            return api_error('该摄像头5分钟内有报修记录，请勿重复提交', 409)
        placeholders = ", ".join(["?"] * len(conflicting_camera_ids))
        label_rows = db.execute(
            f"""
            SELECT id, COALESCE(NULLIF(TRIM(location_desc), ''), NULLIF(TRIM(area), ''), camera_index, CAST(id AS TEXT)) AS label
            FROM cameras
            WHERE id IN ({placeholders})
            ORDER BY id
            """,
            conflicting_camera_ids,
        ).fetchall()
        labels = [row['label'] for row in label_rows]
        suffix = f"：{'、'.join(labels)}" if labels else ''
        return api_error(f'以下摄像头5分钟内已有报修记录，请勿重复提交{suffix}', 409)

    # 插入故障记录
    try:
        import hashlib
        fault_group_key = None
        if len(selected_camera_ids) > 1:
            seed = f"{data['station_id']}|{project_id or ''}|{','.join(str(item) for item in selected_camera_ids)}|{current_time}"
            fault_group_key = hashlib.md5(seed.encode('utf-8')).hexdigest()[:16]

        # 创建阶段忽略归因相关输入；归因只允许在闭环阶段确认。
        impact_camera_count = len(selected_camera_ids) if selected_camera_ids else None
        detail_table_enabled = len(selected_camera_ids) > 1
        if detail_table_enabled:
            ensure_fault_report_camera_detail_schema(db)

        created_fault_ids = []
        db.execute("BEGIN")

        if not camera_rows:
            camera_rows = [None]

        if detail_table_enabled:
            primary_camera_row = camera_rows[0]
            primary_camera_id = None  # 聚合工单不挂单一摄像头
            primary_camera_slot_id = None  # 聚合工单不挂单一slot，明细表是唯一摄像头来源
            primary_project_device_code = None
            if primary_camera_row and 'project_camera_code' in camera_columns and primary_camera_row['project_camera_code']:
                primary_project_device_code = primary_camera_row['project_camera_code']
            elif primary_camera_row and 'camera_index' in camera_columns:
                primary_project_device_code = primary_camera_row['camera_index']

            insert_columns = [
                'station_id', 'camera_id', 'fault_type', 'description',
                'reporter_name', 'reporter_contact', 'status', 'idempotency_key'
            ]
            insert_values = [
                data['station_id'],
                primary_camera_id,
                fault_type_value,
                data.get('description', ''),
                data['reporter_name'],
                data.get('reporter_contact', ''),
                'open',
                fault_group_key or idempotency_keys[0],
            ]

            if 'project_id' in fault_report_columns:
                insert_columns.append('project_id')
                insert_values.append(project_id)
            if 'camera_slot_id' in fault_report_columns:
                insert_columns.append('camera_slot_id')
                insert_values.append(primary_camera_slot_id)
            if 'fault_type_label_snapshot' in fault_report_columns:
                insert_columns.append('fault_type_label_snapshot')
                insert_values.append(fault_type_label_snapshot)
            if 'fault_type_code' in fault_report_columns and data.get('fault_type_code'):
                insert_columns.append('fault_type_code')
                insert_values.append(data.get('fault_type_code'))
            if 'fault_type_version_id' in fault_report_columns:
                insert_columns.append('fault_type_version_id')
                insert_values.append(fault_type_version_id)
            if 'project_device_code' in fault_report_columns:
                insert_columns.append('project_device_code')
                insert_values.append(primary_project_device_code)
            if 'fault_group_key' in fault_report_columns:
                insert_columns.append('fault_group_key')
                insert_values.append(fault_group_key)
            if 'root_cause_type' in fault_report_columns:
                insert_columns.append('root_cause_type')
                insert_values.append(None)
            if 'impact_camera_count' in fault_report_columns and impact_camera_count is not None:
                insert_columns.append('impact_camera_count')
                insert_values.append(impact_camera_count)
            if 'is_batch_impact' in fault_report_columns:
                insert_columns.append('is_batch_impact')
                insert_values.append(1)

            placeholders = ', '.join(['?'] * len(insert_columns))
            cursor = db.execute(
                f"""
                INSERT INTO fault_reports ({', '.join(insert_columns)})
                VALUES ({placeholders})
                """,
                insert_values,
            )
            primary_fault_id = cursor.lastrowid
            created_fault_ids.append(primary_fault_id)

            detail_columns = [
                'fault_report_id', 'camera_id', 'camera_slot_id', 'recovery_state',
                'camera_label', 'detail_fault_reason', 'detail_resolution', 'detail_note'
            ]
            detail_placeholders = ', '.join(['?'] * len(detail_columns))
            detail_optional_columns = ensure_fault_report_camera_detail_schema(db)
            for camera_row in camera_rows:
                current_camera_slot_id = camera_row['slot_id'] if 'slot_id' in camera_columns else None
                current_project_device_code = None
                if 'project_camera_code' in camera_columns and camera_row['project_camera_code']:
                    current_project_device_code = camera_row['project_camera_code']
                elif 'camera_index' in camera_columns:
                    current_project_device_code = camera_row['camera_index']
                detail_insert_columns = list(detail_columns)
                detail_insert_values = [
                    primary_fault_id,
                    camera_row['id'],
                    current_camera_slot_id,
                    'pending',
                    build_fault_camera_label(camera_row),
                    None,
                    None,
                    None,
                ]
                if 'project_id' in detail_optional_columns:
                    detail_insert_columns.append('project_id')
                    detail_insert_values.append(project_id)
                if 'project_device_code' in detail_optional_columns:
                    detail_insert_columns.append('project_device_code')
                    detail_insert_values.append(current_project_device_code)
                detail_placeholders = ', '.join(['?'] * len(detail_insert_columns))
                db.execute(
                    f"""
                    INSERT INTO fault_report_cameras ({', '.join(detail_insert_columns)})
                    VALUES ({detail_placeholders})
                    """,
                    detail_insert_values,
                )
        else:
            for index, camera_row in enumerate(camera_rows):
                current_camera_id = camera_row['id'] if camera_row else None
                current_camera_slot_id = camera_row['slot_id'] if camera_row and 'slot_id' in camera_columns else None
                current_project_device_code = None
                if camera_row and 'project_camera_code' in camera_columns and camera_row['project_camera_code']:
                    current_project_device_code = camera_row['project_camera_code']
                elif camera_row and 'camera_index' in camera_columns:
                    current_project_device_code = camera_row['camera_index']

                insert_columns = [
                    'station_id', 'camera_id', 'fault_type', 'description',
                    'reporter_name', 'reporter_contact', 'status', 'idempotency_key'
                ]
                insert_values = [
                    data['station_id'],
                    current_camera_id,
                    fault_type_value,
                    data.get('description', ''),
                    data['reporter_name'],
                    data.get('reporter_contact', ''),
                    'open',
                    idempotency_keys[index] if index < len(idempotency_keys) else None,
                ]

                if 'project_id' in fault_report_columns:
                    insert_columns.append('project_id')
                    insert_values.append(project_id)
                if 'camera_slot_id' in fault_report_columns:
                    insert_columns.append('camera_slot_id')
                    insert_values.append(current_camera_slot_id)
                if 'fault_type_label_snapshot' in fault_report_columns:
                    insert_columns.append('fault_type_label_snapshot')
                    insert_values.append(fault_type_label_snapshot)
                if 'fault_type_code' in fault_report_columns and data.get('fault_type_code'):
                    insert_columns.append('fault_type_code')
                    insert_values.append(data.get('fault_type_code'))
                if 'fault_type_version_id' in fault_report_columns:
                    insert_columns.append('fault_type_version_id')
                    insert_values.append(fault_type_version_id)
                if 'project_device_code' in fault_report_columns:
                    insert_columns.append('project_device_code')
                    insert_values.append(current_project_device_code)
                if 'fault_group_key' in fault_report_columns:
                    insert_columns.append('fault_group_key')
                    insert_values.append(fault_group_key)
                if 'root_cause_type' in fault_report_columns:
                    insert_columns.append('root_cause_type')
                    insert_values.append(None)
                if 'impact_camera_count' in fault_report_columns and impact_camera_count is not None:
                    insert_columns.append('impact_camera_count')
                    insert_values.append(impact_camera_count)

                placeholders = ', '.join(['?'] * len(insert_columns))
                cursor = db.execute(
                    f"""
                    INSERT INTO fault_reports ({', '.join(insert_columns)})
                    VALUES ({placeholders})
                    """,
                    insert_values,
                )
                created_fault_ids.append(cursor.lastrowid)

        db.commit()

        for fault_id in created_fault_ids:
            _safe_dispatch_fault_notification(db, fault_id, 'fault_created')

        return api_success({
            'fault_id': created_fault_ids[0],
            'fault_ids': created_fault_ids,
            'fault_count': len(created_fault_ids),
            'fault_group_key': fault_group_key,
            'is_aggregated': bool(detail_table_enabled),
            'message': '故障报修提交成功'
        }, 201)

    except Exception as e:
        db.rollback()
        return api_error(f'提交失败: {e}', 500)


# ============================================================
# API: 故障列表
# ============================================================

@app.route('/api/faults', methods=['GET'])
def get_faults():
    """获取故障记录列表（分页）"""
    db = get_db()

    # 支持筛选
    ensure_ai_runtime_schema(db)
    deleted_mode = get_fault_deleted_mode()
    status = request.args.get('status')
    station_id = request.args.get('station_id', type=int)
    year = request.args.get('year', type=int)
    project_code = request.args.get('project', '').strip()
    source_type = request.args.get('source_type', '').strip()

    # 分页参数
    page = max(request.args.get('page', default=1, type=int), 1)
    page_size = request.args.get('page_size', default=50, type=int)
    page_size = min(max(page_size, 1), 200)
    offset = (page - 1) * page_size

    # 构建WHERE条件（用于两个查询）
    where_clause = " WHERE 1=1"
    count_where = " WHERE 1=1"
    params = []
    fault_report_columns = get_table_columns(db, "fault_reports")

    if status:
        where_clause += " AND f.status = ?"
        count_where += " AND f.status = ?"
        params.append(status)

    if station_id:
        where_clause += " AND f.station_id = ?"
        count_where += " AND f.station_id = ?"
        params.append(station_id)

    if year:
        where_clause += " AND strftime('%Y', f.created_at) = ?"
        count_where += " AND strftime('%Y', f.created_at) = ?"
        params.append(str(year))

    if source_type:
        if 'source_type' not in fault_report_columns:
            return api_error('source_type filter unavailable', 409)
        where_clause += " AND f.source_type = ?"
        count_where += " AND f.source_type = ?"
        params.append(source_type)

    if project_code and project_code != 'all':
        if not projects_enabled(db) or 'project_id' not in fault_report_columns:
            return api_error('当前数据库尚未启用项目筛选', 409)
        project = get_project_by_code(db, project_code, include_inactive=True)
        if not project:
            return api_error('项目不存在', 404)
        if not ensure_project_read_access(db, project_code):
            return project_access_denied()
        where_clause += " AND f.project_id = ?"
        count_where += " AND f.project_id = ?"
        params.append(project['id'])

    deleted_clause = build_fault_deleted_clause(fault_report_columns, alias="f", mode=deleted_mode)
    where_clause += deleted_clause
    count_where += deleted_clause

    # 分离的COUNT查询（不JOIN cameras表，避免列名问题）
    count_query = f"""
        SELECT COUNT(*) as total
        FROM fault_reports f
        JOIN stations s ON f.station_id = s.id
        {count_where}
    """
    total_row = db.execute(count_query, params).fetchone()
    total = total_row['total'] if total_row else 0

    # 主查询
    query = f"""
        SELECT f.*, s.name as station_name, s.voltage_level,
               COALESCE(NULLIF(TRIM(c.area), ''), NULLIF(TRIM(cs.area), ''), '') as camera_area,
               COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(cs.location_desc), ''), NULLIF(TRIM(f.camera_location_text), ''), '') as camera_location
        FROM fault_reports f
        JOIN stations s ON f.station_id = s.id
        LEFT JOIN cameras c ON f.camera_id = c.id
        LEFT JOIN camera_slots cs ON f.camera_slot_id = cs.id
        {where_clause}
        ORDER BY {FAULT_STATUS_SORT_SQL}, f.created_at DESC, f.id DESC
        LIMIT ? OFFSET ?
    """
    params.extend([page_size, offset])

    rows = db.execute(query, params).fetchall()
    fault_items = []
    detail_table_exists = table_exists(db, "fault_report_cameras")
    detail_map = {}
    if detail_table_exists:
        ensure_fault_report_camera_detail_schema(db)
        detail_map = fetch_fault_camera_details_map(db, [int(row['id']) for row in rows])
    for row in rows:

        fault_item = enrich_fault_camera_location(dict(row))
        fault_item['camera_details'] = detail_map.get(int(fault_item['id']), []) if detail_table_exists else []
        attach_fault_camera_detail_summary(fault_item)
        fault_items.append(fault_item)


    return api_success({
        'faults': fault_items,
        'total': total,
        'page': page,
        'page_size': page_size,
        'deleted_mode': deleted_mode,
    })


# ============================================================
# API: 故障状态更新（决策#7：三状态机）
# ============================================================

@app.route('/api/faults/<int:fault_id>/status', methods=['PUT'])
def update_fault_status(fault_id):
    """更新故障状态"""
    data = request.get_json()
    if not data or 'status' not in data:
        return api_error('未提供状态')

    new_status = data['status']
    valid_statuses = ['open', 'handling', 'closed']

    if new_status not in valid_statuses:
        return api_error(f'无效状态，可选值: {valid_statuses}')

    # 状态转换验证（决策#7）
    db = get_db()
    fault_report_columns = ensure_fault_report_multi_camera_schema(db)
    fault_report_columns = ensure_fault_report_soft_delete_schema(db)
    deleted_clause = build_fault_deleted_clause(fault_report_columns, alias="", mode="active")


    select_fields = ["id", "status"]
    if 'project_id' in fault_report_columns:
        select_fields.append("project_id")
    if 'impact_camera_count' in fault_report_columns:
        select_fields.append("impact_camera_count")
    if 'fault_owner_type' in fault_report_columns:
        select_fields.append("fault_owner_type")
    if 'is_batch_impact' in fault_report_columns:
        select_fields.append("is_batch_impact")
    fault = db.execute(
        f"SELECT {', '.join(select_fields)} FROM fault_reports WHERE id = ?{deleted_clause}",
        (fault_id,),
    ).fetchone()

    if not fault:
        return api_error('故障记录不存在', 404)

    current_status = fault['status']

    # 状态转换规则
    valid_transitions = {
        'open': ['handling', 'closed'],
        'handling': ['closed'],
        'closed': []  # 已关闭的不能转换
    }

    if new_status not in valid_transitions.get(current_status, []):
        return api_error(f'不能从 {current_status} 转换为 {new_status}')

    # 任意状态进入 closed 都走同一套闭环校验，避免 open->closed 成为旁路。
    if new_status == 'closed':
        handler_name = str(data.get('handler_name') or '').strip()
        handler_note = str(data.get('handler_note') or '').strip()
        equipment_type = str(data.get('equipment_type') or '').strip()
        equipment_quantity = data.get('equipment_quantity', 0)

        if not handler_name or not handler_note:
            return api_error('关闭故障需要提供处理人姓名和处理备注')

        try:
            equipment_quantity = int(equipment_quantity or 0)
        except (TypeError, ValueError):
            return api_error('更换设备数量无效', 400)

        fault_owner_type = str(data.get('fault_owner_type') or '').strip()
        if fault_owner_type and fault_owner_type not in VALID_OWNER_TYPES:
            return api_error('故障归属类型无效', 400)

        root_cause_type = str(data.get('root_cause_type') or '').strip()
        if root_cause_type and root_cause_type not in VALID_OWNER_TYPES:
            return api_error('根因类型无效', 400)

        batch_impact_value = None
        if 'is_batch_impact' in data:
            batch_val = data.get('is_batch_impact')
            if batch_val is not None and str(batch_val).strip() != '':
                try:
                    batch_impact_value = int(batch_val)
                except (TypeError, ValueError):
                    return api_error('是否共因取值无效', 400)
                if batch_impact_value not in (0, 1):
                    return api_error('是否共因只能为 0 或 1', 400)

        impact_camera_count = fault['impact_camera_count'] if 'impact_camera_count' in fault.keys() else None
        if impact_camera_count is not None and int(impact_camera_count) > 1 and not fault_owner_type:
            return api_error('多摄像头故障闭环时必须填写故障归属', 400)

        if fault_owner_type == 'camera' and batch_impact_value == 1:
            return api_error('摄像头本体故障不能标记为共因故障，请拆分为多条单摄像头故障', 400)

        raw_camera_detail_updates = data.get('camera_details')
        try:
            normalized_camera_detail_updates = normalize_fault_camera_detail_updates(raw_camera_detail_updates)
        except ValueError as error:
            error_code = str(error)
            if error_code == 'camera_details_invalid':
                return api_error('camera_details 必须是数组', 400)
            if error_code == 'camera_detail_camera_id_invalid':
                return api_error('camera_details.camera_id 无效', 400)
            if error_code == 'camera_detail_camera_id_duplicated':
                return api_error('camera_details.camera_id 不能重复', 400)
            if error_code == 'camera_detail_recovery_state_invalid':
                return api_error('camera_details.recovery_state 无效', 400)
            return api_error('camera_details 参数无效', 400)

        all_self_recovered = (
            normalized_camera_detail_updates
            and all(item['recovery_state'] == 'self_recovered' for item in normalized_camera_detail_updates)
        )
        if all_self_recovered:
            equipment_type = ''
            equipment_quantity = 0

        update_fields = [
            "status = 'closed'",
            "handler_name = ?",
            "handler_note = ?",
            "closed_at = CURRENT_TIMESTAMP",
            "updated_at = CURRENT_TIMESTAMP",
        ]
        update_params = [handler_name, handler_note]

        if {"equipment_type", "equipment_quantity"}.issubset(fault_report_columns):
            update_fields.append("equipment_type = ?")
            update_fields.append("equipment_quantity = ?")
            update_params.extend([equipment_type, equipment_quantity])

        if 'fault_owner_type' in fault_report_columns:
            update_fields.append("fault_owner_type = ?")
            update_params.append(fault_owner_type or None)
            if fault_owner_type:
                if 'fault_owner_confirmed_by' in fault_report_columns:
                    update_fields.append("fault_owner_confirmed_by = ?")
                    update_params.append(session.get('user_id'))
                if 'fault_owner_confirmed_at' in fault_report_columns:
                    update_fields.append("fault_owner_confirmed_at = CURRENT_TIMESTAMP")

        if 'root_cause_type' in fault_report_columns:
            update_fields.append("root_cause_type = ?")
            update_params.append(root_cause_type or None)

        if 'is_batch_impact' in fault_report_columns and 'is_batch_impact' in data:
            update_fields.append("is_batch_impact = ?")
            update_params.append(batch_impact_value)

        if 'fault_type' in data or 'fault_type_code' in data:
            try:
                resolved_fault_type = resolve_fault_type_update_payload(
                    db,
                    fault,
                    fault_report_columns,
                    data.get('fault_type'),
                    data.get('fault_type_code'),
                )
            except ValueError as error:
                return api_error(str(error), 400)

            if 'fault_type' in fault_report_columns and resolved_fault_type['fault_type']:
                update_fields.append("fault_type = ?")
                update_params.append(resolved_fault_type['fault_type'])
            if 'fault_type_label_snapshot' in fault_report_columns and resolved_fault_type['fault_type']:
                update_fields.append("fault_type_label_snapshot = ?")
                update_params.append(resolved_fault_type['fault_type'])
            if 'fault_type_code' in fault_report_columns:
                update_fields.append("fault_type_code = ?")
                update_params.append(resolved_fault_type['fault_type_code'])
            if 'fault_type_version_id' in fault_report_columns:
                update_fields.append("fault_type_version_id = ?")
                update_params.append(resolved_fault_type['fault_type_version_id'])

        update_params.append(fault_id)
        try:
            db.execute(
                f"""
                UPDATE fault_reports
                SET {', '.join(update_fields)}
                WHERE id = ?
                """,
                update_params,
            )
            apply_fault_camera_detail_closure(
                db,
                fault_id,
                fault_owner_type=fault_owner_type or None,
                batch_impact_value=batch_impact_value,
                handler_note=handler_note,
                camera_detail_updates=normalized_camera_detail_updates,
            )
        except ValueError as error:
            error_code = str(error)
            if error_code == 'camera_details_required_for_camera_owner':
                db.rollback()
                return api_error('摄像头本体故障闭环时必须逐个填写 camera_details', 400)
            if error_code == 'camera_detail_camera_scope_mismatch':
                db.rollback()
                return api_error('camera_details 必须完整覆盖当前聚合工单的全部摄像头', 400)
            if error_code == 'self_recovered_requires_reason':
                db.rollback()
                return api_error('自恢复的摄像头必须填写故障原因（detail_fault_reason）', 400)
            if error_code == 'camera_detail_incomplete_closure':
                db.rollback()
                return api_error('闭环后仍有摄像头明细处于待处理状态，请检查 camera_details 是否完整', 400)
            raise

    else:
        if new_status == 'handling' and 'planned_handle_time' in data and 'planned_handle_time' in fault_report_columns:
            planned = data.get('planned_handle_time')
            planned_value = None
            if planned not in (None, ''):
                planned_text = str(planned).strip()
                parsed = None
                for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S'):
                    try:
                        parsed = datetime.strptime(planned_text, fmt)
                        break
                    except ValueError:
                        continue
                if parsed:
                    planned_value = parsed.strftime('%Y-%m-%d %H:%M:%S')
            db.execute(
                "UPDATE fault_reports SET status = ?, planned_handle_time = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (new_status, planned_value, fault_id),
            )
        else:
            db.execute("""
                UPDATE fault_reports
                SET status = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (new_status, fault_id))

    db.commit()
    logger.info(f"Fault status updated: id={fault_id}, {current_status} -> {new_status}")
    if new_status == 'closed':
        _safe_dispatch_fault_notification(db, fault_id, 'fault_closed')
        _sync_closed_fault_to_worklog(fault_id)

    return api_success({'message': f'状态已更新为 {new_status}'})

# ============================================================
# API: 故障记录删除（仅admin）
# ============================================================

@app.route('/api/faults/<int:fault_id>', methods=['DELETE'])
@require_admin
def delete_fault(fault_id):
    """删除故障记录（仅admin）"""
    db = get_db()
    fault_report_columns = ensure_fault_report_soft_delete_schema(db)

    fault = db.execute("SELECT id, deleted_at FROM fault_reports WHERE id = ?",
                        (fault_id,)).fetchone()

    if not fault:
        return api_error('故障记录不存在', 404)

    if fault['deleted_at']:
        return api_success({'message': '故障记录已在回收站', 'fault_id': fault_id})

    db.execute(
        """
        UPDATE fault_reports
        SET deleted_at = CURRENT_TIMESTAMP,
            updated_at = CURRENT_TIMESTAMP,
            deleted_by = ?
        WHERE id = ?
        """,
        (session.get('user_id'), fault_id),
    )
    db.commit()
    logger.info(f"Fault soft-deleted: id={fault_id}")

    return api_success({'message': '故障记录已移入回收站', 'fault_id': fault_id})


@app.route('/api/faults/<int:fault_id>/restore', methods=['POST'])
@require_admin
def restore_fault(fault_id):
    db = get_db()
    ensure_fault_report_soft_delete_schema(db)

    fault = db.execute("SELECT id, deleted_at FROM fault_reports WHERE id = ?", (fault_id,)).fetchone()
    if not fault:
        return api_error('故障记录不存在', 404)
    if not fault['deleted_at']:
        return api_success({'message': '故障记录无需恢复', 'fault_id': fault_id})

    db.execute(
        """
        UPDATE fault_reports
        SET deleted_at = NULL,
            deleted_by = NULL,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (fault_id,),
    )
    db.commit()
    logger.info(f"Fault restored from trash: id={fault_id}")

    return api_success({'message': '故障记录已恢复', 'fault_id': fault_id})


@app.route('/api/faults/<int:fault_id>', methods=['PUT'])
def update_fault(fault_id):
    db = get_db()
    ensure_ai_runtime_schema(db)
    fault_report_columns = ensure_fault_report_soft_delete_schema(db)
    payload = request.get_json(silent=True) or {}
    if not payload:
        return api_error('无可更新内容')

    select_fields = ["id", "status"]
    if 'project_id' in fault_report_columns:
        select_fields.append("project_id")
    deleted_clause = build_fault_deleted_clause(fault_report_columns, alias="", mode="active")
    fault = db.execute(
        f"SELECT {', '.join(select_fields)} FROM fault_reports WHERE id = ?{deleted_clause}",
        (fault_id,),
    ).fetchone()
    if not fault:
        return api_error('fault not found', 404)

    if projects_enabled(db) and 'project_id' in fault_report_columns and fault['project_id']:
        project_row = db.execute(
            "SELECT code FROM projects WHERE id = ?",
            (fault['project_id'],),
        ).fetchone()
        if project_row and not ensure_project_write_access(db, project_row['code']):
            return project_access_denied()

    editable_fields = [
        ('fault_type', 'fault_type'),
        ('description', 'description'),
        ('camera_location_text', 'camera_location_text'),
        ('reporter_name', 'reporter_name'),
        ('reporter_contact', 'reporter_contact'),
        ('handler_name', 'handler_name'),
        ('handler_note', 'handler_note'),
        ('equipment_type', 'equipment_type'),
        ('equipment_quantity', 'equipment_quantity'),
        ('fault_owner_type', 'fault_owner_type'),
        ('root_cause_type', 'root_cause_type'),
        ('is_batch_impact', 'is_batch_impact'),
    ]

    update_fields = []
    update_params = []
    for request_key, column_name in editable_fields:
        if request_key not in payload or column_name not in fault_report_columns:
            continue
        value = payload.get(request_key)
        if column_name in ('equipment_quantity', 'is_batch_impact'):
            try:
                value = int(value or 0)
            except (TypeError, ValueError):
                return api_error(f'{column_name} must be an integer')
        elif value is None:
            value = ''
        elif isinstance(value, str):
            value = value.strip()
        if column_name in ('fault_owner_type', 'root_cause_type') and value:
            if value not in VALID_OWNER_TYPES:
                return api_error(f'{column_name} 值无效')
        update_fields.append(f"{column_name} = ?")
        update_params.append(value)

    if 'created_at' in payload and 'created_at' in fault_report_columns:
        raw_created_at = payload.get('created_at')
        created_at_value = None
        if raw_created_at not in (None, ''):
            raw_created_at_text = str(raw_created_at).strip()
            parsed_created_at = None
            for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S'):
                try:
                    parsed_created_at = datetime.strptime(raw_created_at_text, fmt)
                    break
                except ValueError:
                    continue
            if parsed_created_at is None:
                return api_error('created_at 格式不正确，请使用正确的日期时间')
            created_at_value = parsed_created_at.strftime('%Y-%m-%d %H:%M:%S')
        else:
            return api_error('故障上报时间不能为空')

        update_fields.append("created_at = ?")
        update_params.append(created_at_value)

    if 'closed_at' in payload and 'closed_at' in fault_report_columns:
        raw_closed_at = payload.get('closed_at')
        closed_at_value = None
        if raw_closed_at not in (None, ''):
            raw_closed_at_text = str(raw_closed_at).strip()
            parsed_closed_at = None
            for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S'):
                try:
                    parsed_closed_at = datetime.strptime(raw_closed_at_text, fmt)
                    break
                except ValueError:
                    continue
            if parsed_closed_at is None:
                return api_error('closed_at 格式不正确，请使用正确的日期时间')
            closed_at_value = parsed_closed_at.strftime('%Y-%m-%d %H:%M:%S')
        elif fault['status'] == 'closed':
            return api_error('已关闭故障必须保留闭环时间')

        if fault['status'] != 'closed' and closed_at_value:
            return api_error('只有已关闭故障才能填写闭环时间')

        update_fields.append("closed_at = ?")
        update_params.append(closed_at_value)


    if 'fault_type' in payload and 'fault_type_label_snapshot' in fault_report_columns:
        update_fields.append("fault_type_label_snapshot = ?")
        update_params.append(str(payload.get('fault_type') or '').strip())

    if 'fault_type_code' in payload and 'fault_type_code' in fault_report_columns:
        update_fields.append("fault_type_code = ?")
        update_params.append(str(payload.get('fault_type_code') or '').strip())

    if not update_fields:
        return api_error('没有可更新字段')

    update_fields.append("updated_at = CURRENT_TIMESTAMP")
    update_params.append(fault_id)
    db.execute(
        f"UPDATE fault_reports SET {', '.join(update_fields)} WHERE id = ?",
        update_params,
    )
    db.commit()
    return api_success({'message': '故障记录已更新', 'fault_id': fault_id})

# ============================================================
# API: 故障详情（GET）
# ============================================================

@app.route('/api/faults/<int:fault_id>/detail', methods=['GET'])
def get_fault_detail(fault_id):
    """获取故障完整详情"""
    db = get_db()
    ensure_ai_runtime_schema(db)
    fault_report_columns = ensure_fault_report_soft_delete_schema(db)
    deleted_mode = get_fault_deleted_mode()
    has_camera_slots = table_exists(db, "camera_slots")
    camera_area_expr = "COALESCE(NULLIF(TRIM(c.area), ''), '')"
    camera_location_expr = "COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(f.camera_location_text), ''), '')"
    camera_slot_join = ""
    if has_camera_slots:
        camera_area_expr = "COALESCE(NULLIF(TRIM(c.area), ''), NULLIF(TRIM(cs.area), ''), '')"
        camera_location_expr = (
            "COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(cs.location_desc), ''), "
            "NULLIF(TRIM(f.camera_location_text), ''), '')"
        )
        camera_slot_join = "LEFT JOIN camera_slots cs ON f.camera_slot_id = cs.id"
    deleted_clause = build_fault_deleted_clause(fault_report_columns, alias="f", mode=deleted_mode)
    fault = db.execute("""
        SELECT f.*, s.name as station_name,
               c.camera_index,
               {camera_area_expr} as camera_area,
               {camera_location_expr} as camera_location,
               c.ip_address as camera_ip
        FROM fault_reports f
        LEFT JOIN stations s ON f.station_id = s.id
        LEFT JOIN cameras c ON f.camera_id = c.id
        {camera_slot_join}
        WHERE f.id = ?{deleted_clause}
    """.format(
        camera_area_expr=camera_area_expr,
        camera_location_expr=camera_location_expr,
        camera_slot_join=camera_slot_join,
        deleted_clause=deleted_clause,
    ), (fault_id,)).fetchone()

    if not fault:
        return api_error('故障记录不存在', 404)

    return api_success({'fault': enrich_fault_camera_location(dict(fault))})


# ============================================================
# API: 照片查询与分组
# ============================================================

@app.route('/api/photos', methods=['GET'])
def get_photos():
    """获取照片平铺列表"""
    db = get_db()

    station_id = request.args.get('station_id', type=int)
    county = request.args.get('county', '').strip()
    status = request.args.get('status', '').strip()
    keyword = request.args.get('keyword', '').strip()
    page = max(request.args.get('page', default=1, type=int), 1)
    page_size = request.args.get('page_size', default=60, type=int)
    page_size = min(max(page_size, 1), 200)
    offset = (page - 1) * page_size

    where = ["1=1"]
    params = []

    if station_id:
        where.append("p.station_id = ?")
        params.append(station_id)

    if county:
        where.append("(p.county_hint = ? OR s.county = ?)")
        params.extend([county, county])

    if status in ('matched', 'unmatched', 'ignored'):
        where.append("p.match_status = ?")
        params.append(status)

    if keyword:
        where.append("(p.filename LIKE ? OR p.rel_path LIKE ? OR p.station_hint LIKE ?)")
        kw = f"%{keyword}%"
        params.extend([kw, kw, kw])

    fault_only_clause, fault_only_params = _build_photo_fault_only_clause("p")
    if fault_only_clause:
        where.append(fault_only_clause)
        params.extend(fault_only_params)

    where_sql = " AND ".join(where)

    total_row = db.execute(
        f"""
        SELECT COUNT(*) AS total
        FROM photos p
        LEFT JOIN stations s ON p.station_id = s.id
        WHERE {where_sql}
        """,
        params
    ).fetchone()

    rows = db.execute(
        f"""
        SELECT p.*, s.name AS station_name, s.county AS station_county, s.voltage_level
        FROM photos p
        LEFT JOIN stations s ON p.station_id = s.id
        WHERE {where_sql}
        ORDER BY p.file_mtime DESC, p.id DESC
        LIMIT ? OFFSET ?
        """,
        params + [page_size, offset]
    ).fetchall()

    return api_success({
        'photos': [normalize_photo_row(row) for row in rows],
        'total': total_row['total'] if total_row else 0,
        'page': page,
        'page_size': page_size,
    })


@app.route('/api/photos/groups', methods=['GET'])
def get_photo_groups():
    """按变电站分组返回照片"""
    db = get_db()

    county = request.args.get('county', '').strip()
    station_id = request.args.get('station_id', type=int)
    status = request.args.get('status', '').strip()
    keyword = request.args.get('keyword', '').strip()
    limit_per_group = request.args.get('limit_per_group', default=120, type=int)
    limit_per_group = min(max(limit_per_group, 1), 300)

    where = ["1=1"]
    params = []

    if county:
        where.append("(p.county_hint = ? OR s.county = ?)")
        params.extend([county, county])

    if station_id:
        where.append("p.station_id = ?")
        params.append(station_id)

    if status in ('matched', 'unmatched', 'ignored'):
        where.append("p.match_status = ?")
        params.append(status)

    if keyword:
        where.append("(p.filename LIKE ? OR p.rel_path LIKE ? OR p.station_hint LIKE ?)")
        kw = f"%{keyword}%"
        params.extend([kw, kw, kw])

    fault_only_clause, fault_only_params = _build_photo_fault_only_clause("p")
    if fault_only_clause:
        where.append(fault_only_clause)
        params.extend(fault_only_params)

    where_sql = " AND ".join(where)

    rows = db.execute(
        f"""
        SELECT p.*, s.name AS station_name, s.county AS station_county
        FROM photos p
        LEFT JOIN stations s ON p.station_id = s.id
        WHERE {where_sql}
        ORDER BY p.match_status DESC, s.name, p.file_mtime DESC, p.id DESC
        """,
        params
    ).fetchall()

    grouped = {}
    unmatched = []

    for row in rows:
        photo = normalize_photo_row(row)
        if photo['match_status'] == 'matched' and photo.get('station_id'):
            key = str(photo['station_id'])
            if key not in grouped:
                grouped[key] = {
                    'station_id': photo['station_id'],
                    'station_name': photo.get('station_name') or '未知变电站',
                    'county': photo.get('station_county') or photo.get('county_hint') or '',
                    'photos': []
                }
            if len(grouped[key]['photos']) < limit_per_group:
                grouped[key]['photos'].append(photo)
        elif photo['match_status'] == 'unmatched':
            if len(unmatched) < limit_per_group:
                unmatched.append(photo)

    groups = sorted(grouped.values(), key=lambda g_: (g_['county'], g_['station_name']))

    return api_success({
        'groups': groups,
        'unmatched': unmatched,
        'group_count': len(groups),
        'unmatched_count': len([r for r in rows if r['match_status'] == 'unmatched'])
    })


@app.route('/photos/file/<int:photo_id>', methods=['GET'])
def get_photo_file(photo_id):
    """受控图片访问端点（photo_id映射 + root约束，防止路径穿越）"""
    # 需要登录才能访问照片
    if 'user_id' not in session:
        return api_error('请先登录', 401)

    db = get_db()
    row, photo_columns = _fetch_photo_asset_row(db, photo_id)
    if not row:
        return api_error('照片不存在', 404)

    ext = (row['ext'] or '').lower()
    if ext not in IMAGE_EXTENSIONS:
        return api_error('文件类型不支持', 400)

    root = get_photo_root()
    file_path = Path(row['abs_path']).resolve()

    # 必须位于PHOTO_ROOT_PATH内，防止path traversal
    if not is_path_under_root(file_path, root):
        logger.warning(f"Blocked photo traversal attempt: photo_id={photo_id}, path={file_path}")
        return api_error('非法路径访问', 403)

    if not file_path.exists() or not file_path.is_file():
        return _send_photo_thumbnail(db, row, photo_columns)

    return send_file(str(file_path), conditional=True)


@app.route('/photos/thumb/<int:photo_id>', methods=['GET'])
def get_photo_thumbnail(photo_id):
    if 'user_id' not in session:
        return api_error('请先登录', 401)

    db = get_db()
    row, photo_columns = _fetch_photo_asset_row(db, photo_id)
    if not row:
        return api_error('照片不存在', 404)

    ext = (row['ext'] or '').lower()
    if ext not in IMAGE_EXTENSIONS:
        return api_error('文件类型不支持', 400)

    return _send_photo_thumbnail(db, row, photo_columns)


def get_stats_scoped():
    db = get_db()
    year = request.args.get('year', type=int)
    payload, error = _build_statistics_payload(db, year, request.args.get('project'))
    if error:
        return error
    return api_success(payload)


def get_stations_scoped():
    db = get_db()
    project_scope, error = build_project_scope(db, request.args.get('project'))
    if error:
        return error

    county = request.args.get('county')
    query = """
        SELECT DISTINCT s.id, s.name, s.voltage_level, s.county, s.location, s.latitude, s.longitude
        FROM stations s
        WHERE 1=1
    """
    params = []
    if county:
        query += " AND s.county = ?"
        params.append(county)

    camera_columns = get_table_columns(db, "cameras")
    fault_report_columns = get_table_columns(db, "fault_reports")
    visibility_checks = []
    if project_scope['enabled'] and 'project_id' in camera_columns:
        project_sql, project_params = build_project_in_clause("c", project_scope['project_ids'])
        status_sql = " AND c.status = 'active'" if 'status' in camera_columns else ""
        visibility_checks.append(
            f"EXISTS (SELECT 1 FROM cameras c WHERE c.station_id = s.id{project_sql}{status_sql})"
        )
        params.extend(project_params)
    if project_scope['enabled'] and 'project_id' in fault_report_columns:
        project_sql, project_params = build_project_in_clause("f", project_scope['project_ids'])
        visibility_checks.append(
            f"EXISTS (SELECT 1 FROM fault_reports f WHERE f.station_id = s.id{project_sql})"
        )
        params.extend(project_params)
    if visibility_checks:
        query += " AND (" + " OR ".join(visibility_checks) + ")"

    query += " ORDER BY s.county, s.name"
    rows = db.execute(query, params).fetchall()
    return api_success({
        'stations': [dict(row) for row in rows],
        'total': len(rows)
    })


def get_station_scoped(station_id):
    db = get_db()
    project_scope, error = build_project_scope(db, request.args.get('project'))
    if error:
        return error

    station = db.execute("SELECT * FROM stations WHERE id = ?", (station_id,)).fetchone()
    if not station:
        return api_error('鍙樼數绔欎笉瀛樺湪', 404)

    station_payload = dict(station)
    camera_columns = get_table_columns(db, "cameras")
    fault_report_columns = get_table_columns(db, "fault_reports")

    camera_query = "SELECT * FROM cameras c WHERE c.station_id = ?"
    camera_params = [station_id]
    if project_scope['enabled'] and 'project_id' in camera_columns:
        project_sql, project_params = build_project_in_clause("c", project_scope['project_ids'])
        camera_query += project_sql
        camera_params.extend(project_params)
    if 'status' in camera_columns:
        camera_query += " AND c.status = 'active'"
    camera_query += """
        ORDER BY
            CASE
                WHEN c.channel_number IS NOT NULL THEN c.channel_number
                WHEN TRIM(COALESCE(c.camera_index, '')) GLOB '[0-9]*' THEN CAST(TRIM(c.camera_index) AS INTEGER)
                ELSE 2147483647
            END,
            CASE
                WHEN TRIM(COALESCE(c.camera_index, '')) = '' THEN 1
                ELSE 0
            END,
            c.camera_index,
            c.id
    """
    cameras = db.execute(camera_query, camera_params).fetchall()

    fault_count_query = "SELECT COUNT(*) AS count FROM fault_reports f WHERE f.station_id = ?"
    fault_count_params = [station_id]
    if project_scope['enabled'] and 'project_id' in fault_report_columns:
        project_sql, project_params = build_project_in_clause("f", project_scope['project_ids'])
        fault_count_query += project_sql
        fault_count_params.extend(project_params)
    fault_count = db.execute(fault_count_query, fault_count_params).fetchone()['count']

    station_payload['camera_count'] = len(cameras)
    station_payload['fault_count'] = fault_count
    if project_scope['enabled'] and station_payload['camera_count'] == 0 and station_payload['fault_count'] == 0:
        return api_error('鍙樼數绔欎笉瀛樺湪', 404)

    return api_success({
        'station': station_payload,
        'cameras': [dict(camera) for camera in cameras]
    })


def get_station_slots_scoped(station_id):
    db = get_db()
    project_scope, error = build_project_scope(db, request.args.get('project'))
    if error:
        return error

    station = db.execute("SELECT * FROM stations WHERE id = ?", (station_id,)).fetchone()
    if not station:
        return api_error('变电站不存在', 404)

    fault_report_columns = get_table_columns(db, "fault_reports")
    camera_columns = get_table_columns(db, "cameras")
    recorders = build_station_recorders_payload(db, station_id, project_scope)
    slots = build_station_slots_payload(db, station_id, project_scope)

    if project_scope['enabled']:
        visible_camera = False
        visible_fault = False
        if 'project_id' in camera_columns:
            camera_query = "SELECT 1 FROM cameras c WHERE c.station_id = ?"
            camera_params = [station_id]
            project_sql, project_params = build_project_in_clause("c", project_scope['project_ids'])
            camera_query += project_sql
            camera_params.extend(project_params)
            if 'status' in camera_columns:
                camera_query += " AND c.status = 'active'"
            visible_camera = db.execute(camera_query + " LIMIT 1", camera_params).fetchone() is not None
        if 'project_id' in fault_report_columns:
            fault_query = "SELECT 1 FROM fault_reports f WHERE f.station_id = ?"
            fault_params = [station_id]
            project_sql, project_params = build_project_in_clause("f", project_scope['project_ids'])
            fault_query += project_sql
            fault_params.extend(project_params)
            visible_fault = db.execute(fault_query + " LIMIT 1", fault_params).fetchone() is not None
        if not slots and not visible_camera and not visible_fault:
            return api_error('变电站不存在', 404)

    return api_success({
        'station': dict(station),
        'slots': slots,
        'recorders': recorders,
        'total': len(slots),
    })


def get_cameras_scoped():
    db = get_db()
    project_scope, error = build_project_scope(db, request.args.get('project'))
    if error:
        return error

    station_id = request.args.get('station_id', type=int)
    query = """
        SELECT c.*, s.name as station_name, s.voltage_level
        FROM cameras c
        JOIN stations s ON c.station_id = s.id
        WHERE 1=1
    """
    params = []
    if station_id:
        query += " AND c.station_id = ?"
        params.append(station_id)

    camera_columns = get_table_columns(db, "cameras")
    if project_scope['enabled'] and 'project_id' in camera_columns:
        project_sql, project_params = build_project_in_clause("c", project_scope['project_ids'])
        query += project_sql
        params.extend(project_params)
    if 'status' in camera_columns:
        query += " AND c.status = 'active'"

    query += " ORDER BY s.county, s.name, c.camera_index, c.channel_number"
    rows = db.execute(query, params).fetchall()
    return api_success({
        'cameras': [dict(row) for row in rows],
        'total': len(rows)
    })


def get_camera_by_ip_scoped():
    ip = request.args.get('ip', '').strip()
    if not ip:
        return api_error('鏈彁渚汭P鍦板潃')

    db = get_db()
    camera_columns = get_table_columns(db, "cameras")
    if not (projects_enabled(db) and 'project_id' in camera_columns):
        return get_camera_by_ip()
    project_scope, error = build_project_scope(db, request.args.get('project'))
    if error:
        return error

    query = """
        SELECT c.*, s.name as station_name, s.voltage_level, s.county
        FROM cameras c
        JOIN stations s ON c.station_id = s.id
        WHERE c.ip_address = ?
    """
    params = [ip]
    if project_scope['enabled'] and 'project_id' in camera_columns:
        project_sql, project_params = build_project_in_clause("c", project_scope['project_ids'])
        query += project_sql
        params.extend(project_params)
    if 'status' in camera_columns:
        query += " AND c.status = 'active'"

    camera = db.execute(query, params).fetchone()
    if not camera:
        return api_error('璇P鏆傛湭褰曞叆绯荤粺锛岃閫夋嫨鍙樼數绔欐墜鍔ㄦ姤淇�', 404)
    return api_success({'camera': dict(camera)})


def create_fault_scoped():
    data = request.get_json()
    if not data:
        return api_error('璇锋眰浣撴棤鏁�')

    required = ['station_id', 'reporter_name']
    for field in required:
        if not data.get(field):
            return api_error(f'缂哄皯蹇呭～瀛楁: {field}')

    db = get_db()
    fault_report_columns = get_table_columns(db, "fault_reports")
    camera_columns = get_table_columns(db, "cameras")
    project_code = (data.get('project') or '').strip()

    if project_code:
        project = get_project_by_code(db, project_code, include_inactive=False)
        if not project:
            return api_error('椤圭洰涓嶅瓨鍦�', 404)
        if not ensure_project_write_access(db, project['code']):
            return project_access_denied()
    elif 'project_id' in fault_report_columns and projects_enabled(db):
        try:
            camera_ids = normalize_camera_ids(data.get('camera_ids'))
            if not camera_ids and data.get('camera_id') not in (None, ""):
                camera_ids = normalize_camera_ids([data.get('camera_id')])
        except ValueError:
            return api_error('摄像头参数无效', 400)
        if camera_ids and 'project_id' in camera_columns:
            rows = fetch_camera_rows_by_ids(db, camera_ids, camera_columns)
            if not rows:
                return api_error('摄像头不存在', 404)
            project_ids = {row['project_id'] for row in rows if row['project_id'] is not None}
            if len(project_ids) > 1:
                return api_error('摄像头不属于同一项目', 400)
            if project_ids:
                project_row = db.execute(
                    "SELECT code FROM projects WHERE id = ?",
                    (next(iter(project_ids)),),
                ).fetchone()
                if project_row and not ensure_project_write_access(db, project_row['code']):
                    return project_access_denied()
        else:
            visible_projects = get_visible_projects(
                db,
                user_id=session.get('user_id'),
                role=session.get('role') or 'admin',
                include_inactive=False,
            )
            default_project_code = get_default_project_code(visible_projects) or 'unified'
            if not ensure_project_write_access(db, default_project_code):
                return project_access_denied()

    return create_fault()


def get_faults_scoped():
    db = get_db()
    ensure_ai_runtime_schema(db)
    deleted_mode = get_fault_deleted_mode()
    status = request.args.get('status')
    station_id = request.args.get('station_id', type=int)
    year = request.args.get('year', type=int)
    project_code = request.args.get('project', '').strip()
    source_type = request.args.get('source_type', '').strip()
    page = max(request.args.get('page', default=1, type=int), 1)
    page_size = request.args.get('page_size', default=50, type=int)
    page_size = min(max(page_size, 1), 200)
    offset = (page - 1) * page_size

    fault_report_columns = ensure_fault_report_soft_delete_schema(db)
    has_camera_slots = table_exists(db, "camera_slots")
    project_scope, error = build_project_scope(db, project_code)
    if error:
        return error

    where_clause = " WHERE 1=1"
    params = []
    if status:
        where_clause += " AND f.status = ?"
        params.append(status)
    if station_id:
        where_clause += " AND f.station_id = ?"
        params.append(station_id)
    if year:
        where_clause += " AND strftime('%Y', f.created_at) = ?"
        params.append(str(year))
    if source_type:
        if 'source_type' not in fault_report_columns:
            return api_error('source_type filter unavailable', 409)
        where_clause += " AND f.source_type = ?"
        params.append(source_type)
    if project_scope['enabled'] and 'project_id' in fault_report_columns:
        project_sql, project_params = build_project_in_clause("f", project_scope['project_ids'])
        where_clause += project_sql
        params.extend(project_params)
    where_clause += build_fault_deleted_clause(fault_report_columns, alias="f", mode=deleted_mode)

    count_query = f"""
        SELECT COUNT(*) as total
        FROM fault_reports f
        JOIN stations s ON f.station_id = s.id
        {where_clause}
    """
    total_row = db.execute(count_query, list(params)).fetchone()
    total = total_row['total'] if total_row else 0

    camera_area_expr = "COALESCE(NULLIF(TRIM(c.area), ''), '')"
    camera_location_expr = "COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(f.camera_location_text), ''), '')"
    camera_slot_join = ""
    if has_camera_slots:
        camera_area_expr = "COALESCE(NULLIF(TRIM(c.area), ''), NULLIF(TRIM(cs.area), ''), '')"
        camera_location_expr = (
            "COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(cs.location_desc), ''), "
            "NULLIF(TRIM(f.camera_location_text), ''), '')"
        )
        camera_slot_join = "LEFT JOIN camera_slots cs ON f.camera_slot_id = cs.id"

    query = f"""
        SELECT f.*, s.name as station_name, s.voltage_level,
               {camera_area_expr} as camera_area,
               {camera_location_expr} as camera_location
        FROM fault_reports f
        JOIN stations s ON f.station_id = s.id
        LEFT JOIN cameras c ON f.camera_id = c.id
        {camera_slot_join}
        {where_clause}
        ORDER BY {FAULT_STATUS_SORT_SQL}, f.created_at DESC, f.id DESC
        LIMIT ? OFFSET ?
    """
    rows = db.execute(query, params + [page_size, offset]).fetchall()
    fault_items = []
    detail_table_exists = table_exists(db, "fault_report_cameras")
    detail_map = {}
    if detail_table_exists:
        ensure_fault_report_camera_detail_schema(db)
        detail_map = fetch_fault_camera_details_map(db, [int(row['id']) for row in rows])
    for row in rows:
        fault_item = enrich_fault_camera_location(dict(row))
        fault_item['camera_details'] = detail_map.get(int(fault_item['id']), []) if detail_table_exists else []
        attach_fault_camera_detail_summary(fault_item)
        fault_items.append(fault_item)
    return api_success({
        'faults': fault_items,
        'total': total,
        'page': page,
        'page_size': page_size,
        'deleted_mode': deleted_mode,
    })




def update_fault_status_scoped(fault_id):
    data = request.get_json()
    if not data or 'status' not in data:
        return api_error('未提供状态')

    new_status = data['status']
    valid_statuses = ['open', 'handling', 'closed']
    if new_status not in valid_statuses:
        return api_error(f'无效状态，可选值: {valid_statuses}')

    db = get_db()
    fault_report_columns = ensure_fault_report_multi_camera_schema(db)
    fault_report_columns = ensure_fault_report_soft_delete_schema(db)
    if not (projects_enabled(db) and 'project_id' in fault_report_columns):
        return update_fault_status(fault_id)

    select_fields = ["id", "status"]
    if 'project_id' in fault_report_columns:
        select_fields.append("project_id")
    fault = db.execute(
        f"SELECT {', '.join(select_fields)} FROM fault_reports WHERE id = ?{build_fault_deleted_clause(fault_report_columns, alias='', mode='active')}",
        (fault_id,),
    ).fetchone()
    if not fault:
        return api_error('故障记录不存在', 404)

    if 'project_id' in fault_report_columns and fault['project_id']:
        project_row = db.execute(
            "SELECT code FROM projects WHERE id = ?",
            (fault['project_id'],),
        ).fetchone()
        if project_row and not ensure_project_write_access(db, project_row['code']):
            return project_access_denied()

    return update_fault_status(fault_id)



def get_fault_detail_scoped(fault_id):
    db = get_db()
    ensure_ai_runtime_schema(db)
    fault_report_columns = ensure_fault_report_soft_delete_schema(db)
    deleted_mode = get_fault_deleted_mode()
    has_camera_slots = table_exists(db, "camera_slots")
    assigned_user_select = ""
    assigned_user_join = ""
    if 'assigned_to' in fault_report_columns and table_exists(db, "users"):
        assigned_user_select = ", u.username as assigned_to_username"
        assigned_user_join = " LEFT JOIN users u ON f.assigned_to = u.id"
    camera_area_expr = "COALESCE(NULLIF(TRIM(c.area), ''), '')"
    camera_location_expr = "COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(f.camera_location_text), ''), '')"
    camera_slot_join = ""
    if has_camera_slots:
        camera_area_expr = "COALESCE(NULLIF(TRIM(c.area), ''), NULLIF(TRIM(cs.area), ''), '')"
        camera_location_expr = (
            "COALESCE(NULLIF(TRIM(c.location_desc), ''), NULLIF(TRIM(cs.location_desc), ''), "
            "NULLIF(TRIM(f.camera_location_text), ''), '')"
        )
        camera_slot_join = " LEFT JOIN camera_slots cs ON f.camera_slot_id = cs.id"
    fault = db.execute(
        f"""
        SELECT f.*, s.name as station_name,
               c.camera_index,
               {camera_area_expr} as camera_area,
               {camera_location_expr} as camera_location,
               c.ip_address as camera_ip
               {assigned_user_select}
        FROM fault_reports f
        LEFT JOIN stations s ON f.station_id = s.id
        LEFT JOIN cameras c ON f.camera_id = c.id
        {camera_slot_join}
        {assigned_user_join}
        WHERE f.id = ?{build_fault_deleted_clause(fault_report_columns, alias='f', mode=deleted_mode)}
        """,
        (fault_id,),
    ).fetchone()
    if not fault:
        return api_error('鏁呴殰璁板綍涓嶅瓨鍦�', 404)

    if 'project_id' in fault_report_columns and fault['project_id']:
        project_row = db.execute("SELECT code FROM projects WHERE id = ?", (fault['project_id'],)).fetchone()
        if project_row and not ensure_project_read_access(db, project_row['code']):
            return project_access_denied()

    payload = dict(fault)
    enrich_fault_camera_location(payload)
    detail_table_exists = table_exists(db, "fault_report_cameras")
    if detail_table_exists:
        ensure_fault_report_camera_detail_schema(db)
        payload['camera_details'] = fetch_fault_camera_details(db, fault_id)
    else:
        payload['camera_details'] = []
    attach_fault_camera_detail_summary(payload)
    if 'tags_json' in fault_report_columns:
        payload['tags'] = parse_tags_json(payload.get('tags_json'))

    payload['slot_history'] = []
    payload['slot_history_count'] = 0
    if 'camera_slot_id' in fault_report_columns and payload.get('camera_slot_id'):
        history_query = """
            SELECT
                f.id,
                f.camera_slot_id,
                COALESCE(f.fault_type_label_snapshot, f.fault_type) AS fault_label,
                f.status,
                f.created_at,
                f.closed_at
            FROM fault_reports f
            WHERE f.camera_slot_id = ?
              AND f.id != ?
        """
        history_params = [payload['camera_slot_id'], fault_id]
        if 'project_id' in fault_report_columns and payload.get('project_id'):
            history_query += " AND f.project_id = ?"
            history_params.append(payload['project_id'])
        history_query += build_fault_deleted_clause(fault_report_columns, alias="f", mode="active")
        history_query += " ORDER BY f.created_at DESC, f.id DESC LIMIT 5"
        slot_history = [dict(row) for row in db.execute(history_query, history_params).fetchall()]
        payload['slot_history'] = slot_history
        payload['slot_history_count'] = len(slot_history)
    return api_success({'fault': payload})


@app.route('/api/fault-tags', methods=['GET'])
def get_fault_tag_suggestions():
    db = get_db()
    fault_report_columns = ensure_fault_report_soft_delete_schema(db)
    if 'tags_json' not in fault_report_columns:
        return api_success({'tags': []})

    project_scope, error = build_project_scope(db, request.args.get('project'))
    if error:
        return error

    query = "SELECT tags_json FROM fault_reports f WHERE f.tags_json IS NOT NULL AND TRIM(f.tags_json) != ''"
    params = []
    query += build_fault_deleted_clause(fault_report_columns, alias="f", mode="active")
    if project_scope['enabled'] and 'project_id' in fault_report_columns:
        project_sql, project_params = build_project_in_clause("f", project_scope['project_ids'])
        query += project_sql
        params.extend(project_params)

    search = (request.args.get('q') or '').strip().lower()
    tags = []
    seen = set()
    for row in db.execute(query, params).fetchall():
        row_tags = parse_tags_json(row['tags_json'] if hasattr(row, 'keys') else row[0])
        for tag in row_tags:
            if search and search not in tag.lower():
                continue
            lowered = tag.lower()
            if lowered in seen:
                continue
            seen.add(lowered)
            tags.append(tag)
    tags.sort(key=lambda item: item.lower())
    return api_success({'tags': tags})


@app.route('/api/faults/<int:fault_id>/tags', methods=['PUT'])
def update_fault_tags(fault_id):
    db = get_db()
    fault_report_columns = ensure_fault_report_soft_delete_schema(db)
    if 'tags_json' not in fault_report_columns:
        return api_error('tags feature unavailable', 409)

    payload = request.get_json(silent=True) or {}
    tags = normalize_tags_payload(payload.get('tags', []))

    fault = db.execute(
        f"SELECT id, project_id FROM fault_reports WHERE id = ?{build_fault_deleted_clause(fault_report_columns, alias='', mode='active')}",
        (fault_id,),
    ).fetchone()
    if not fault:
        return api_error('fault not found', 404)

    if projects_enabled(db) and 'project_id' in fault_report_columns and fault['project_id']:
        project_row = db.execute(
            "SELECT code FROM projects WHERE id = ?",
            (fault['project_id'],),
        ).fetchone()
        if project_row and not ensure_project_write_access(db, project_row['code']):
            return project_access_denied()

    db.execute(
        "UPDATE fault_reports SET tags_json = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (json.dumps(tags, ensure_ascii=False), fault_id),
    )
    db.commit()
    return api_success({'fault_id': fault_id, 'tags': tags})


app.view_functions['get_stats'] = get_stats_scoped
app.view_functions['get_stations'] = get_stations_scoped
app.view_functions['get_station'] = get_station_scoped
app.view_functions['get_station_slots'] = get_station_slots_scoped
app.view_functions['get_cameras'] = get_cameras_scoped
app.view_functions['get_camera_by_ip'] = get_camera_by_ip_scoped
app.view_functions['create_fault'] = create_fault_scoped
app.view_functions['get_faults'] = get_faults_scoped
app.view_functions['update_fault_status'] = update_fault_status_scoped
app.view_functions['get_fault_detail'] = get_fault_detail_scoped


def _photos_project_scope_enabled(db):
    photo_columns = get_table_columns(db, "photos")
    return projects_enabled(db) and 'project_id' in photo_columns, photo_columns


def _build_photo_fault_only_clause(photo_alias: str, project_scope: dict | None = None):
    has_fault = request.args.get('has_fault', '').strip().lower() in ('1', 'true', 'yes')
    if not has_fault:
        return "", []

    db = get_db()
    fault_report_columns = get_table_columns(db, "fault_reports")
    fault_scope_sql = ""
    fault_scope_params = []
    if project_scope and project_scope.get('enabled') and 'project_id' in fault_report_columns:
        fault_scope_sql, fault_scope_params = build_project_in_clause("f", project_scope['project_ids'])

    return (
        f"({photo_alias}.match_status = 'unmatched' "
        f"OR EXISTS (SELECT 1 FROM fault_reports f WHERE f.station_id = {photo_alias}.station_id"
        f"{fault_scope_sql} AND COALESCE(f.status, 'open') IN ('open', 'handling')))"
    ), fault_scope_params


def _apply_photo_project_filter(db, base_where, params, requested_project_code):
    enabled, photo_columns = _photos_project_scope_enabled(db)
    if not enabled:
        return base_where, params, None, None

    project_scope, error = build_project_scope(db, requested_project_code)
    if error:
        return None, None, None, error

    where = list(base_where)
    scoped_params = list(params)
    if project_scope['enabled']:
        project_sql, project_params = build_project_in_clause("p", project_scope['project_ids'])
        if project_sql:
            where.append(project_sql.strip().removeprefix("AND ").strip())
            scoped_params.extend(project_params)
    return where, scoped_params, project_scope, None


def get_photos_scoped():
    db = get_db()
    enabled, _ = _photos_project_scope_enabled(db)
    if not enabled:
        return get_photos()

    station_id = request.args.get('station_id', type=int)
    county = request.args.get('county', '').strip()
    status = request.args.get('status', '').strip()
    keyword = request.args.get('keyword', '').strip()
    page = max(request.args.get('page', default=1, type=int), 1)
    page_size = request.args.get('page_size', default=60, type=int)
    page_size = min(max(page_size, 1), 200)
    offset = (page - 1) * page_size

    where = ["1=1"]
    params = []
    if station_id:
        where.append("p.station_id = ?")
        params.append(station_id)
    if county:
        where.append("(p.county_hint = ? OR s.county = ?)")
        params.extend([county, county])
    if status in ('matched', 'unmatched', 'ignored'):
        where.append("p.match_status = ?")
        params.append(status)
    if keyword:
        kw = f"%{keyword}%"
        where.append("(p.filename LIKE ? OR p.rel_path LIKE ? OR p.station_hint LIKE ?)")
        params.extend([kw, kw, kw])

    where, params, project_scope, error = _apply_photo_project_filter(
        db,
        where,
        params,
        request.args.get('project'),
    )
    if error:
        return error

    fault_only_clause, fault_only_params = _build_photo_fault_only_clause("p", project_scope)
    if fault_only_clause:
        where.append(fault_only_clause)
        params.extend(fault_only_params)

    where_sql = " AND ".join(where)
    total_row = db.execute(
        f"""
        SELECT COUNT(*) AS total
        FROM photos p
        LEFT JOIN stations s ON p.station_id = s.id
        WHERE {where_sql}
        """,
        params,
    ).fetchone()
    rows = db.execute(
        f"""
        SELECT p.*, s.name AS station_name, s.county AS station_county, s.voltage_level
        FROM photos p
        LEFT JOIN stations s ON p.station_id = s.id
        WHERE {where_sql}
        ORDER BY p.file_mtime DESC, p.id DESC
        LIMIT ? OFFSET ?
        """,
        params + [page_size, offset],
    ).fetchall()
    return api_success({
        'photos': [normalize_photo_row(row) for row in rows],
        'total': total_row['total'] if total_row else 0,
        'page': page,
        'page_size': page_size,
    })


def get_photo_groups_scoped():
    db = get_db()
    enabled, _ = _photos_project_scope_enabled(db)
    if not enabled:
        return get_photo_groups()

    county = request.args.get('county', '').strip()
    station_id = request.args.get('station_id', type=int)
    status = request.args.get('status', '').strip()
    keyword = request.args.get('keyword', '').strip()
    limit_per_group = request.args.get('limit_per_group', default=120, type=int)
    limit_per_group = min(max(limit_per_group, 1), 300)

    where = ["1=1"]
    params = []
    if county:
        where.append("(p.county_hint = ? OR s.county = ?)")
        params.extend([county, county])
    if station_id:
        where.append("p.station_id = ?")
        params.append(station_id)
    if status in ('matched', 'unmatched', 'ignored'):
        where.append("p.match_status = ?")
        params.append(status)
    if keyword:
        kw = f"%{keyword}%"
        where.append("(p.filename LIKE ? OR p.rel_path LIKE ? OR p.station_hint LIKE ?)")
        params.extend([kw, kw, kw])

    where, params, project_scope, error = _apply_photo_project_filter(
        db,
        where,
        params,
        request.args.get('project'),
    )
    if error:
        return error

    fault_only_clause, fault_only_params = _build_photo_fault_only_clause("p", project_scope)
    if fault_only_clause:
        where.append(fault_only_clause)
        params.extend(fault_only_params)

    rows = db.execute(
        f"""
        SELECT p.*, s.name AS station_name, s.county AS station_county
        FROM photos p
        LEFT JOIN stations s ON p.station_id = s.id
        WHERE {' AND '.join(where)}
        ORDER BY p.match_status DESC, s.name, p.file_mtime DESC, p.id DESC
        """,
        params,
    ).fetchall()

    grouped = {}
    unmatched = []
    for row in rows:
        photo = normalize_photo_row(row)
        if photo['match_status'] == 'matched' and photo.get('station_id'):
            key = str(photo['station_id'])
            if key not in grouped:
                grouped[key] = {
                    'station_id': photo['station_id'],
                    'station_name': photo.get('station_name') or '未知变电站',
                    'county': photo.get('station_county') or photo.get('county_hint') or '',
                    'photos': []
                }
            if len(grouped[key]['photos']) < limit_per_group:
                grouped[key]['photos'].append(photo)
        elif photo['match_status'] == 'unmatched':
            if len(unmatched) < limit_per_group:
                unmatched.append(photo)

    groups = sorted(grouped.values(), key=lambda item: (item['county'], item['station_name']))
    return api_success({
        'groups': groups,
        'unmatched': unmatched,
        'group_count': len(groups),
        'unmatched_count': len([row for row in rows if row['match_status'] == 'unmatched'])
    })


def get_photo_file_scoped(photo_id):
    db = get_db()
    enabled, _ = _photos_project_scope_enabled(db)
    if not enabled:
        return get_photo_file(photo_id)

    if 'user_id' not in session:
        return api_error('请先登录', 401)

    row, _ = _fetch_photo_asset_row(db, photo_id, include_project=True)
    if not row:
        return api_error('照片不存在', 404)

    if row['project_id']:
        project_row = db.execute("SELECT code FROM projects WHERE id = ?", (row['project_id'],)).fetchone()
        if project_row and not ensure_project_read_access(db, project_row['code']):
            return project_access_denied()

    return get_photo_file(photo_id)


def get_photo_thumbnail_scoped(photo_id):
    db = get_db()
    enabled, _ = _photos_project_scope_enabled(db)
    if not enabled:
        return get_photo_thumbnail(photo_id)

    if 'user_id' not in session:
        return api_error('请先登录', 401)

    row, _ = _fetch_photo_asset_row(db, photo_id, include_project=True)
    if not row:
        return api_error('照片不存在', 404)

    if row['project_id']:
        project_row = db.execute("SELECT code FROM projects WHERE id = ?", (row['project_id'],)).fetchone()
        if project_row and not ensure_project_read_access(db, project_row['code']):
            return project_access_denied()

    return get_photo_thumbnail(photo_id)


def export_statistics_scoped():
    db = get_db()
    fault_report_columns = get_table_columns(db, "fault_reports")
    camera_columns = get_table_columns(db, "cameras")
    if not (
        projects_enabled(db)
        and 'project_id' in fault_report_columns
        and 'project_id' in camera_columns
    ):
        return export_statistics()

    year = request.args.get('year', type=int)
    project_scope, error = build_project_scope(db, request.args.get('project'))
    if error:
        return error

    payload, error = _build_statistics_payload(db, year, request.args.get('project'))
    if error:
        return error
    detail_rows = _build_statistics_detail_rows(db, project_scope, year)

    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Border, Font, PatternFill, Side

        wb = Workbook()
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2f67f6', end_color='2f67f6', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws1 = wb.active
        ws1.title = '概览'
        overview = [
            ('项目范围', payload['project_scope']['requested_project']),
            ('统计年份', payload['target_year']),
            ('变电站总数', payload['stations']),
            ('当前设备总数', payload['cameras']),
            ('故障总数', payload['faults']),
            ('故障率', f"{payload['fault_rate']:.2f}%"),
            ('本月故障数', payload['faults_this_month']),
            ('本年故障数', payload['faults_this_year']),
            ('更换摄像头数', payload['kpi']['camera_replacement_count']),
        ]
        ws1.append(['指标', '值'])
        for key, value in overview:
            ws1.append([key, value])

        ws2 = wb.create_sheet('KPI')
        ws2.append(['指标', '值'])
        ws2.append(['打开中', payload['kpi']['open_count']])
        ws2.append(['处理中', payload['kpi']['handling_count']])
        ws2.append(['已关闭', payload['kpi']['closed_count']])
        ws2.append(['响应样本数', payload['kpi']['response_sample_count']])
        ws2.append(['关闭样本数', payload['kpi']['close_sample_count']])
        ws2.append(['平均响应时长(秒)', payload['kpi']['avg_response_seconds']])
        ws2.append(['平均关闭时长(秒)', payload['kpi']['avg_close_seconds']])
        ws2.append(['更换摄像头数量', payload['kpi']['camera_replacement_count']])
        ws2.append(['更换摄像头记录数', payload['kpi']['camera_replacement_record_count']])
        ws2.append(['未填数量按1台估算记录数', payload['kpi']['camera_replacement_inferred_record_count']])

        ws3 = wb.create_sheet('月度趋势')
        ws3.append(['月份', '故障数量'])
        for item in payload['monthly_trend']:
            ws3.append([item['month'], item['count']])

        ws4 = wb.create_sheet('故障语义统计')
        ws4.append(['semantic_group', '故障类型', '数量'])
        for item in payload['fault_type_distribution']:
            ws4.append([item['semantic_group'], item['fault_label'], item['count']])

        ws5 = wb.create_sheet('县区统计')
        ws5.append(['县区', '故障数量'])
        for item in payload['county_distribution']:
            ws5.append([item['county'], item['count']])

        ws6 = wb.create_sheet('电压等级统计')
        ws6.append(['电压等级', '故障数量'])
        for item in payload['voltage_distribution']:
            ws6.append([item['voltage_level'], item['count']])

        ws7 = wb.create_sheet('故障明细')
        common_columns = [
            ('ID', 'id'),
            ('项目编码', 'project_code'),
            ('项目名称', 'project_name'),
            ('变电站', 'station_name'),
            ('电压等级', 'voltage_level'),
            ('县区', 'county'),
            ('摄像头位置', 'camera_display_text'),
            ('摄像头明细', 'camera_locations_text'),
            ('逐路恢复摘要', 'camera_recovery_text'),
            ('自恢复路数', 'self_recovered_camera_count'),
            ('项目设备编号', 'project_device_code'),
            ('semantic_group', 'semantic_group'),
            ('故障类型编码', 'fault_type_code'),
            ('故障类型名称', 'fault_label'),
            ('描述', 'description'),
            ('状态', 'status'),
            ('报修人', 'reporter_name'),
            ('联系方式', 'reporter_contact'),
            ('创建时间', 'created_at'),
            ('开始处理时间', 'handling_started_at'),
            ('关闭时间', 'closed_at'),
            ('故障归属', 'fault_owner_type'),
            ('根因确认', 'root_cause_type'),
            ('共因标记', 'is_batch_impact'),
            ('影响摄像头数', 'impact_camera_count'),
            ('是否聚合单', 'is_aggregated'),
            ('计入统计摄像头数', 'statistics_camera_count'),
        ]
        admin_columns = [
            ('槽位ID', 'camera_slot_id'),
            ('来源类型', 'source_type'),
            ('来源批次', 'source_batch_id'),
            ('来源幂等键', 'source_record_key'),
            ('原始时间', 'source_time_raw'),
            ('原始时区', 'source_timezone'),
            ('遗留系统类型', 'system_type'),
            ('处理人ID', 'assigned_to'),
            ('处理人账号', 'assigned_to_username'),
            ('历史处理人', 'handler_name'),
            ('处理备注', 'handler_note'),
            ('标签JSON', 'tags_json'),
        ]
        export_columns = common_columns + (admin_columns if session.get('role') == 'admin' else [])
        ws7.append([header for header, _ in export_columns])
        for row in detail_rows:
            export_row = []
            for header, key in export_columns:
                value = row.get(key, '')
                if key in ('fault_owner_type', 'root_cause_type') and value:
                    value = ROOT_CAUSE_LABELS.get(value, value)
                export_row.append(value)
            ws7.append(export_row)

        for sheet in [ws1, ws2, ws3, ws4, ws5, ws6, ws7]:
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            for column in sheet.columns:
                sheet.column_dimensions[column[0].column_letter].width = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.exception('export_statistics_scoped failed')
        return api_error(f'导出失败: {e}', 500)


app.view_functions['export_statistics'] = export_statistics_scoped
app.view_functions['get_photos'] = get_photos_scoped
app.view_functions['get_photo_groups'] = get_photo_groups_scoped
app.view_functions['get_photo_file'] = get_photo_file_scoped
app.view_functions['get_photo_thumbnail'] = get_photo_thumbnail_scoped


# ============================================================
# 服务端搜索 API（替代前端全量加载搜索）
# ============================================================

@app.route('/api/search', methods=['GET'])
def search():
    """服务端搜索：按关键词匹配站点和故障，返回前 10 条结果。"""
    q = (request.args.get('q') or '').strip()
    if not q or len(q) < 1:
        return api_success({'results': [], 'total': 0})

    if 'user_id' not in session:
        return api_error('请先登录', 401)

    db = get_db()
    project_scope, error = build_project_scope(db, request.args.get('project'))
    if error:
        return error

    like_pattern = f'%{q}%'
    results = []
    limit = 10

    # 搜索站点
    station_params = [like_pattern, like_pattern, like_pattern]
    station_project_clause = ''
    if project_scope['enabled']:
        camera_columns = get_table_columns(db, "cameras")
        if 'project_id' in camera_columns:
            station_project_clause = (
                " AND EXISTS (SELECT 1 FROM cameras c WHERE c.station_id = s.id"
                " AND c.project_id IN ({}) )".format(
                    ','.join(['?'] * len(project_scope['project_ids']))
                )
            )
            station_params.extend(project_scope['project_ids'])

    stations = db.execute(
        f"""
        SELECT DISTINCT s.id, s.name, s.voltage_level, s.county
        FROM stations s
        WHERE (s.name LIKE ? OR s.county LIKE ? OR s.location LIKE ?)
        {station_project_clause}
        ORDER BY s.county, s.name
        LIMIT ?
        """,
        station_params + [limit],
    ).fetchall()

    for s in stations:
        results.append({
            'type': 'station',
            'title': s['name'],
            'subtitle': f"{s['county'] or ''} · {s['voltage_level'] or ''}",
            'url': f"/design/style2/stations?project={project_scope.get('project_code', '')}",
        })

    # 搜索故障
    fault_report_columns = get_table_columns(db, "fault_reports")
    deleted_clause = build_fault_deleted_clause(fault_report_columns, alias="f")
    fault_project_clause = ''
    fault_params = [like_pattern, like_pattern, like_pattern]
    if project_scope['enabled'] and 'project_id' in fault_report_columns:
        fault_project_clause = (
            " AND f.project_id IN ({})".format(
                ','.join(['?'] * len(project_scope['project_ids']))
            )
        )
        fault_params.extend(project_scope['project_ids'])

    faults = db.execute(
        f"""
        SELECT f.id, f.fault_type, f.status, f.station_id,
               s.name AS station_name
        FROM fault_reports f
        LEFT JOIN stations s ON f.station_id = s.id
        WHERE (f.description LIKE ? OR f.fault_type LIKE ? OR s.name LIKE ?)
        {deleted_clause}
        {fault_project_clause}
        ORDER BY f.created_at DESC
        LIMIT ?
        """,
        fault_params + [limit],
    ).fetchall()

    status_labels = {'open': '待处理', 'handling': '处理中', 'closed': '已关闭'}
    for f in faults:
        results.append({
            'type': 'fault',
            'title': f"#{f['id']} {f['station_name'] or '-'}",
            'subtitle': f"{f['fault_type'] or '-'} · {status_labels.get(f['status'], f['status'])}",
            'url': f"/design/style2/faults?project={project_scope.get('project_code', '')}",
        })

    results.sort(key=lambda r: r.get('title', ''))
    return api_success({
        'results': results[:limit],
        'total': len(results),
        'query': q,
    })


# ============================================================
# 健康检查
# ============================================================

@app.route('/health', methods=['GET'])
def health():
    """健康检查"""
    return api_success({'status': 'ok'})

# ============================================================
# 页面路由
# ============================================================

@app.route('/')
def index():
    """首页"""
    return render_template('index.html')

@app.route('/stations')
def stations():
    """变电站列表页"""
    return render_template('stations.html')

@app.route('/fault/new')
def fault_new():
    """故障报修页"""
    return render_template('fault_new.html')

@app.route('/faults')
def faults():
    """故障记录页"""
    return render_template(
        'faults.html',
        is_admin=session.get('role') == 'admin',
        can_edit_tags=bool(session.get('user_id')),
    )

@app.route('/statistics')
def statistics():
    """统计页面"""
    return render_template('statistics.html')

@app.route('/map')
def map_page():
    """地图页面"""
    return render_template('map.html')

@app.route('/photos')
def photos_page():
    """照片分类页"""
    return render_template('photos.html')

@app.route('/login')
def login_page():
    """登录页面"""
    return render_template('login.html')

@app.route('/admin')
def admin_page():
    """管理后台页面"""
    if session.get('role') != 'admin':
        return redirect('/login')
    return render_template('admin.html')

# ============================================================
# 设计变体路由
# ============================================================

@app.route('/design')
def design_variants():
    """设计变体选择页"""
    return render_template('design_variants/index.html')

@app.route('/design/<style>')
def design_variant_index(style):
    """设计变体首页 - 三种风格预览"""
    valid_styles = ['style1', 'style2', 'style3']
    if style not in valid_styles:
        return "Invalid style", 404

    template_map = {
        'style1': 'design_variants/index_style1.html',
        'style2': 'index.html',
        'style3': 'design_variants/index_style3.html'
    }

    return render_template(template_map[style])

@app.route('/design/<style>/stations')
def design_variant_stations(style):
    """设计变体 - 变电站列表页"""
    if style not in ['style1', 'style2', 'style3']:
        return "Invalid style", 404
    return render_template('stations.html')

@app.route('/design/<style>/fault/new')
def design_variant_fault_new(style):
    """设计变体 - 故障报修页"""
    if style not in ['style1', 'style2', 'style3']:
        return "Invalid style", 404
    return render_template('fault_new.html')

@app.route('/design/<style>/faults')
def design_variant_faults(style):
    """设计变体 - 故障记录页"""
    if style not in ['style1', 'style2', 'style3']:
        return "Invalid style", 404
    return render_template(
        'faults.html',
        is_admin=session.get('role') == 'admin',
        can_edit_tags=bool(session.get('user_id')),
    )

@app.route('/design/<style>/statistics')
def design_variant_statistics(style):
    """设计变体 - 统计报表页"""
    if style not in ['style1', 'style2', 'style3']:
        return "Invalid style", 404
    return render_template('statistics.html')

@app.route('/design/<style>/map')
def design_variant_map(style):
    """设计变体 - 地图页"""
    if style not in ['style1', 'style2', 'style3']:
        return "Invalid style", 404
    return render_template('map.html')

@app.route('/design/<style>/photos')
def design_variant_photos(style):
    """设计变体 - 照片页"""
    if style not in ['style1', 'style2', 'style3']:
        return "Invalid style", 404
    return render_template('photos.html')

# ============================================================
# 启动
# ============================================================

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=Config.DEBUG)
