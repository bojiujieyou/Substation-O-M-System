# create_coordinate_template.py — 创建坐标导入模板
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "坐标导入"

# 表头样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

# 写入表头
headers = ["变电站名称", "电压等级", "纬度", "经度"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 设置列宽
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12

# 示例数据（丽水地区部分变电站的近似坐标，实际需要从地图获取准确值）
sample_data = [
    # 名称, 电压等级, 纬度, 经度
    ["110kV丽水变", "110kV", 28.4672, 119.9222],
    ["110kV水阁变", "110kV", 28.4205, 119.8901],
    ["110kV岩都变", "110kV", 28.3567, 119.8534],
    ["110kV杨千变", "110kV", 28.4891, 119.9756],
    ["110kV恒泽变", "110kV", 28.4456, 119.9456],
    ["35kV水阁变", "35kV", 28.4205, 119.8901],
    # 示例结束，请删除上述行并填入真实坐标
]

for row_idx, data in enumerate(sample_data, start=2):
    for col_idx, value in enumerate(data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 添加提示行
ws.cell(row=len(sample_data) + 3, column=1, value="提示：请使用高德地图或百度地图获取准确坐标")
ws.cell(row=len(sample_data) + 4, column=1, value="高德坐标拾取: https://lbs.amap.com/tool/mappicker")

output_path = os.path.join(os.path.dirname(__file__), 'coordinates_template.xlsx')
wb.save(output_path)
print(f"模板已创建: {output_path}")
print("请用高德/百度地图获取准确坐标后填写")
