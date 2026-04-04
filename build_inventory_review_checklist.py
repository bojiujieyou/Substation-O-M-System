from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from parse_excel import (
    ExcelParseError,
    _is_flat_inventory_format,
    _normalize_header_value,
    parse_station_excel,
)


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _compact_text(value: Any) -> str:
    text = _normalize_text(value)
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff#-]+", "", text)
    return text.lower()


def _is_placeholder_text(value: Any) -> bool:
    compact = _compact_text(value)
    return compact in {"", "-", "--", "na", "n-a", "null", "none"}


def _load_flat_inventory_row_metadata(filepath: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows or not _is_flat_inventory_format(rows):
        return []

    header_row = rows[0] or []
    header_index: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        header = _normalize_header_value(cell)
        if header:
            header_index[header] = index

    station_col = header_index.get("变电站")
    device_name_col = header_index.get("设备名称")
    if station_col is None or device_name_col is None:
        return []

    row_metadata: list[dict[str, Any]] = []
    for excel_row, row in enumerate(rows[1:], start=2):
        if not row or len(row) <= device_name_col:
            continue
        station_name = ""
        if len(row) > station_col and row[station_col]:
            station_name = str(row[station_col]).strip()
        device_name = str(row[device_name_col]).strip() if row[device_name_col] else ""
        if not station_name or not device_name:
            continue
        row_metadata.append(
            {
                "excel_row": excel_row,
                "raw_station_name": station_name,
                "raw_device_name": device_name,
            }
        )
    return row_metadata


def _build_review_rows(filepath: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    parsed = parse_station_excel(str(filepath))
    cameras = parsed.get("cameras", [])
    flat_metadata = _load_flat_inventory_row_metadata(filepath)

    review_rows: list[dict[str, Any]] = []
    for index, camera in enumerate(cameras):
        metadata = flat_metadata[index] if index < len(flat_metadata) else {}
        row = {
            "excel_row": metadata.get("excel_row"),
            "raw_station_name": metadata.get("raw_station_name"),
            "raw_device_name": metadata.get("raw_device_name"),
            "project_camera_code": camera.get("project_camera_code") or metadata.get("raw_device_name") or "",
            "camera_index": _normalize_text(camera.get("camera_index")),
            "channel_number": camera.get("channel_number"),
            "area": _normalize_text(camera.get("area")),
            "location_desc": _normalize_text(camera.get("location_desc") or camera.get("location")),
        }
        review_rows.append(row)
    return parsed, review_rows


def _build_slot_signature_label(row: dict[str, Any]) -> str:
    location = _normalize_text(row.get("location_desc"))
    area = _normalize_text(row.get("area")) or "-"
    channel = row.get("channel_number") if row.get("channel_number") is not None else row.get("camera_index") or "-"
    if _is_placeholder_text(location):
        location = "位置缺失"
    return f"{location} | {area} | CH{channel}"


def _collect_group_examples(
    groups: dict[str, list[dict[str, Any]]],
    label_builder,
    value_key: str,
) -> list[dict[str, Any]]:
    ordered = sorted(groups.items(), key=lambda item: (-len(item[1]), item[0]))
    examples: list[dict[str, Any]] = []
    for _, rows in ordered[:3]:
        sample_rows = [
            {
                "excel_row": row.get("excel_row"),
                "project_camera_code": row.get("project_camera_code"),
                "location_desc": row.get("location_desc"),
                "area": row.get("area"),
                "camera_index": row.get("camera_index"),
                "channel_number": row.get("channel_number"),
            }
            for row in rows[:5]
        ]
        examples.append(
            {
                value_key: label_builder(rows[0]),
                "count": len(rows),
                "rows": sample_rows,
            }
        )
    return examples


def _analyze_station(row_report: dict[str, Any]) -> dict[str, Any]:
    filepath = Path(row_report["filepath"])
    parsed, review_rows = _build_review_rows(filepath)

    duplicate_device_code_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    duplicate_slot_signature_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    location_collision_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in review_rows:
        device_code = _compact_text(row.get("project_camera_code"))
        if device_code:
            duplicate_device_code_groups[device_code].append(row)

        slot_signature = "|".join(
            [
                _compact_text(row.get("location_desc")),
                _compact_text(row.get("area")),
                str(row.get("channel_number") if row.get("channel_number") is not None else row.get("camera_index") or ""),
            ]
        )
        duplicate_slot_signature_groups[slot_signature].append(row)

        location_key = _compact_text(row.get("location_desc"))
        if location_key and not _is_placeholder_text(row.get("location_desc")):
            location_collision_groups[location_key].append(row)

    duplicate_device_code_groups = {
        key: rows
        for key, rows in duplicate_device_code_groups.items()
        if key and len(rows) > 1
    }
    duplicate_slot_signature_groups = {
        key: rows
        for key, rows in duplicate_slot_signature_groups.items()
        if key and len(rows) > 1
    }

    filtered_location_collision_groups: dict[str, list[dict[str, Any]]] = {}
    for key, rows in location_collision_groups.items():
        if len(rows) < 2:
            continue
        distinct_variants = {
            (
                _compact_text(item.get("area")),
                str(item.get("channel_number") if item.get("channel_number") is not None else item.get("camera_index") or ""),
            )
            for item in rows
        }
        if len(distinct_variants) > 1:
            filtered_location_collision_groups[key] = rows

    issue_types: list[str] = []
    suggested_actions: list[str] = []

    if duplicate_device_code_groups:
        issue_types.append("duplicate_device_code")
        suggested_actions.append("核对样本内是否存在同设备编号重复台账行，必要时先在源文件去重。")
    if duplicate_slot_signature_groups:
        issue_types.append("duplicate_slot_signature")
        suggested_actions.append("核对同槽位记录是否应视为同一设备，确认 area/location/channel 的命名是否应规范化。")
    if filtered_location_collision_groups:
        issue_types.append("location_collision")
        suggested_actions.append("核对同位置下不同视频类型或不同通道是否应拆分为独立槽位。")
    if not issue_types:
        issue_types.append("naming_or_business_review")
        suggested_actions.append("结合现场命名和台账口径人工确认这些更新是否合理，无需先修改导入链路。")

    change_volume = (
        int(row_report.get("cameras_updated", 0))
        + int(row_report.get("cameras_replaced", 0))
        + int(row_report.get("cameras_retired", 0))
    )
    if change_volume >= 5:
        review_priority = "high"
    elif change_volume >= 2:
        review_priority = "medium"
    else:
        review_priority = "low"

    return {
        "station": row_report.get("station") or parsed["station"]["name"],
        "county": row_report.get("county") or parsed["station"].get("county"),
        "filepath": str(filepath),
        "camera_rows": int(row_report.get("camera_rows", len(review_rows))),
        "cameras_updated": int(row_report.get("cameras_updated", 0)),
        "cameras_replaced": int(row_report.get("cameras_replaced", 0)),
        "cameras_retired": int(row_report.get("cameras_retired", 0)),
        "review_priority": review_priority,
        "issue_types": issue_types,
        "suggested_actions": suggested_actions,
        "signals": {
            "duplicate_device_code_groups": len(duplicate_device_code_groups),
            "duplicate_slot_signature_groups": len(duplicate_slot_signature_groups),
            "location_collision_groups": len(filtered_location_collision_groups),
        },
        "examples": {
            "duplicate_device_code": _collect_group_examples(
                duplicate_device_code_groups,
                lambda row: row.get("project_camera_code") or "",
                "project_camera_code",
            ),
            "duplicate_slot_signature": _collect_group_examples(
                duplicate_slot_signature_groups,
                _build_slot_signature_label,
                "location_desc",
            ),
            "location_collision": _collect_group_examples(
                filtered_location_collision_groups,
                lambda row: row.get("location_desc") or "",
                "location_desc",
            ),
        },
    }


def build_review_checklist(
    *,
    dry_run_report_path: str | Path,
    json_output_path: str | Path | None = None,
    markdown_output_path: str | Path | None = None,
    csv_output_path: str | Path | None = None,
) -> dict[str, Any]:
    report_path = Path(dry_run_report_path)
    report = json.loads(report_path.read_text(encoding="utf-8"))

    target_rows = [
        row
        for row in report.get("rows", [])
        if row.get("status") != "imported"
        or int(row.get("cameras_updated", 0)) > 0
        or int(row.get("cameras_replaced", 0)) > 0
        or int(row.get("cameras_retired", 0)) > 0
    ]

    stations: list[dict[str, Any]] = []
    for row in target_rows:
        try:
            stations.append(_analyze_station(row))
        except (FileNotFoundError, ExcelParseError) as exc:
            stations.append(
                {
                    "station": row.get("station") or row.get("file"),
                    "county": row.get("county"),
                    "filepath": row.get("filepath"),
                    "camera_rows": int(row.get("camera_rows", 0)),
                    "cameras_updated": int(row.get("cameras_updated", 0)),
                    "cameras_replaced": int(row.get("cameras_replaced", 0)),
                    "cameras_retired": int(row.get("cameras_retired", 0)),
                    "review_priority": "high",
                    "issue_types": ["report_parse_failed"],
                    "suggested_actions": [f"源文件无法复核，请先修复样本文件可读性：{exc}"],
                    "signals": {
                        "duplicate_device_code_groups": 0,
                        "duplicate_slot_signature_groups": 0,
                        "location_collision_groups": 0,
                    },
                    "examples": {
                        "duplicate_device_code": [],
                        "duplicate_slot_signature": [],
                        "location_collision": [],
                    },
                }
            )

    summary = {
        "station_count": len(stations),
        "updated_total": sum(item["cameras_updated"] for item in stations),
        "replaced_total": sum(item["cameras_replaced"] for item in stations),
        "retired_total": sum(item["cameras_retired"] for item in stations),
        "high_priority_stations": sum(1 for item in stations if item["review_priority"] == "high"),
        "stations_with_duplicate_device_code": sum(
            1 for item in stations if item["signals"]["duplicate_device_code_groups"] > 0
        ),
        "stations_with_duplicate_slot_signature": sum(
            1 for item in stations if item["signals"]["duplicate_slot_signature_groups"] > 0
        ),
        "stations_with_location_collision": sum(
            1 for item in stations if item["signals"]["location_collision_groups"] > 0
        ),
    }

    checklist = {
        "source_report": str(report_path),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": summary,
        "stations": stations,
    }

    if json_output_path:
        output_path = Path(json_output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(checklist, ensure_ascii=False, indent=2), encoding="utf-8")

    if csv_output_path:
        output_path = Path(csv_output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(
                fh,
                fieldnames=[
                    "station",
                    "county",
                    "review_priority",
                    "cameras_updated",
                    "cameras_replaced",
                    "cameras_retired",
                    "issue_types",
                    "duplicate_device_code_groups",
                    "duplicate_slot_signature_groups",
                    "location_collision_groups",
                    "filepath",
                ],
            )
            writer.writeheader()
            for item in stations:
                writer.writerow(
                    {
                        "station": item["station"],
                        "county": item["county"],
                        "review_priority": item["review_priority"],
                        "cameras_updated": item["cameras_updated"],
                        "cameras_replaced": item["cameras_replaced"],
                        "cameras_retired": item["cameras_retired"],
                        "issue_types": ",".join(item["issue_types"]),
                        "duplicate_device_code_groups": item["signals"]["duplicate_device_code_groups"],
                        "duplicate_slot_signature_groups": item["signals"]["duplicate_slot_signature_groups"],
                        "location_collision_groups": item["signals"]["location_collision_groups"],
                        "filepath": item["filepath"],
                    }
                )

    if markdown_output_path:
        output_path = Path(markdown_output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        lines = [
            "# Inventory Residual Review Checklist",
            "",
            f"- Source report: `{report_path}`",
            f"- Stations to review: `{summary['station_count']}`",
            f"- Updated rows to confirm: `{summary['updated_total']}`",
            f"- Replaced rows to confirm: `{summary['replaced_total']}`",
            f"- Retired rows to confirm: `{summary['retired_total']}`",
            "",
            "## Station Checklist",
            "",
        ]
        for item in stations:
            lines.extend(
                [
                    f"### {item['station']}",
                    f"- Priority: `{item['review_priority']}`",
                    f"- Change summary: updated `{item['cameras_updated']}`, replaced `{item['cameras_replaced']}`, retired `{item['cameras_retired']}`",
                    f"- Issue types: `{', '.join(item['issue_types'])}`",
                    f"- Source file: `{item['filepath']}`",
                    "- Suggested actions:",
                ]
            )
            for action in item["suggested_actions"]:
                lines.append(f"  - {action}")
            if item["examples"]["duplicate_device_code"]:
                lines.append("- Duplicate device code examples:")
                for example in item["examples"]["duplicate_device_code"]:
                    lines.append(
                        f"  - `{example['project_camera_code']}` x {example['count']}"
                    )
            if item["examples"]["duplicate_slot_signature"]:
                lines.append("- Duplicate slot signature examples:")
                for example in item["examples"]["duplicate_slot_signature"]:
                    lines.append(
                        f"  - `{example['location_desc']}` x {example['count']}"
                    )
            if item["examples"]["location_collision"]:
                lines.append("- Location collision examples:")
                for example in item["examples"]["location_collision"]:
                    lines.append(
                        f"  - `{example['location_desc']}` x {example['count']}"
                    )
            lines.append("")
        output_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")

    return checklist


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据 inventory dry-run 报告生成残留复核清单")
    parser.add_argument("--dry-run-report", required=True, help="import_excel dry-run 报告路径")
    parser.add_argument("--json-out", help="输出 JSON 清单")
    parser.add_argument("--md-out", help="输出 Markdown 清单")
    parser.add_argument("--csv-out", help="输出 CSV 清单")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    checklist = build_review_checklist(
        dry_run_report_path=args.dry_run_report,
        json_output_path=args.json_out,
        markdown_output_path=args.md_out,
        csv_output_path=args.csv_out,
    )
    print(f"待复核站点: {checklist['summary']['station_count']}")
    print(f"updated 合计: {checklist['summary']['updated_total']}")
    print(f"replaced 合计: {checklist['summary']['replaced_total']}")
    print(f"retired 合计: {checklist['summary']['retired_total']}")
    if args.json_out:
        print(f"JSON: {args.json_out}")
    if args.md_out:
        print(f"Markdown: {args.md_out}")
    if args.csv_out:
        print(f"CSV: {args.csv_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
